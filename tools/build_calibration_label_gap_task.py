from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from tools.build_linked_workbook import build_workbook


LABEL_GAP_PREFIX_FIELDS = [
    "label_priority",
    "label_target_decisive_rows",
    "label_decisive_rows_needed",
    "label_goal",
    "label_question",
    "label_decision_hint",
    "label_decision_impact",
    "label_evidence_source_kind",
    "label_evidence_source_path",
]


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Build a focused calibration label-gap review task.")
    parser.add_argument("--status-json", required=True, help="calibration_status.json from run_calibration_cycle.py.")
    parser.add_argument("--sample-csv", help="Optional sample CSV. Defaults to artifacts.sample_csv in status JSON.")
    parser.add_argument("--filled-sample", help="Optional filled CSV/XLSX rows to exclude from the next gap task.")
    parser.add_argument("--output-csv", required=True)
    parser.add_argument("--output-xlsx")
    parser.add_argument(
        "--priority",
        action="append",
        choices=["high", "medium", "normal", "low"],
        help="Priority to include. Repeatable. Default includes all priorities with a positive gap.",
    )
    args = parser.parse_args(argv)

    summary = build_calibration_label_gap_task(
        status_json=args.status_json,
        sample_csv=args.sample_csv,
        filled_sample=args.filled_sample,
        output_csv=args.output_csv,
        output_xlsx=args.output_xlsx,
        priorities=args.priority,
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


def build_calibration_label_gap_task(
    *,
    status_json: str | Path,
    output_csv: str | Path,
    sample_csv: str | Path | None = None,
    filled_sample: str | Path | None = None,
    output_xlsx: str | Path | None = None,
    priorities: list[str] | None = None,
) -> dict:
    status = _read_json(Path(status_json))
    sample_value = sample_csv or status.get("artifacts", {}).get("sample_csv") or ""
    sample_path = Path(sample_value) if sample_value else Path()
    sample_rows = _read_rows(sample_path) if sample_value and sample_path.is_file() else []
    filled_keys = _filled_row_keys(Path(filled_sample)) if filled_sample else set()
    headers = list(sample_rows[0].keys()) if sample_rows else []
    status_summary = status.get("summary", {})
    allowed_priorities = set(priorities or ["high", "medium", "normal", "low"])
    selected: list[dict[str, str]] = []
    target_summaries = []

    for target in status.get("label_targets", []):
        priority = str(target.get("priority") or "")
        if priority not in allowed_priorities:
            continue
        needed = _to_int(target.get("decisive_rows_needed"))
        if needed <= 0:
            continue
        reason = str(target.get("review_reason") or "")
        candidates = [
            row
            for row in sample_rows
            if row.get("review_reason") == reason and not str(row.get("manual_decision") or "").strip()
            and _row_key(row) not in filled_keys
        ]
        picked = candidates[:needed]
        for row in picked:
            selected.append(_label_gap_row(row, target, status_summary))
        target_summaries.append(
            {
                "review_reason": reason,
                "priority": priority,
                "needed": needed,
                "selected": len(picked),
                "available_unlabeled": len(candidates),
            }
        )

    fields = _fields(headers)
    output_csv_path = Path(output_csv)
    _write_rows(output_csv_path, selected, fields)
    xlsx_summary = {}
    if output_xlsx:
        xlsx_summary = build_workbook([("Label_Gap_Task", output_csv_path)], output_xlsx)
    return {
        "status_json": str(status_json),
        "sample_csv": str(sample_path),
        "filled_sample": str(filled_sample or ""),
        "output_csv": str(output_csv_path),
        "output_xlsx": str(output_xlsx or ""),
        "task_rows": len(selected),
        "priority_counts": dict(Counter(row.get("label_priority", "") for row in selected)),
        "reason_counts": dict(Counter(row.get("review_reason", "") for row in selected)),
        "targets": target_summaries,
        "xlsx": xlsx_summary,
    }


def _label_gap_row(row: dict[str, str], target: dict, status_summary: dict) -> dict[str, str]:
    out = dict(row)
    review_reason = str(row.get("review_reason") or target.get("review_reason") or "")
    out["label_priority"] = str(target.get("priority") or "")
    out["label_target_decisive_rows"] = str(target.get("target_decisive_rows") or "")
    out["label_decisive_rows_needed"] = str(target.get("decisive_rows_needed") or "")
    out["label_goal"] = str(target.get("label_goal") or "")
    out["label_question"] = _label_question(review_reason)
    out["label_decision_hint"] = _label_decision_hint(review_reason)
    out["label_decision_impact"] = _label_decision_impact(review_reason, target)
    out["label_evidence_source_kind"] = _label_evidence_source_kind(review_reason, status_summary)
    out["label_evidence_source_path"] = _label_evidence_source_path(review_reason, status_summary)
    return out


def _label_question(review_reason: str) -> str:
    if review_reason == "recall_unresolved_top_candidate":
        return "Should this unresolved candidate be accepted as the provider's official website?"
    if review_reason == "precision_calibrated_pattern_release":
        return "Is this pattern-released official_url truly the provider's independent official website?"
    if review_reason.startswith("precision_"):
        return "Is the shown official_url truly the provider's independent official website?"
    return "What is the correct manual decision for this candidate?"


def _label_decision_hint(review_reason: str) -> str:
    base = "Use accept if the shown URL is correct; replace with manual_url if another official site is correct; reject if it is wrong; unsure only when evidence is insufficient."
    if review_reason == "precision_calibrated_pattern_release":
        return base + " A reject/replace here blocks wider automatic release for this pattern until retested."
    if review_reason == "precision_second_pass_accepted_lt70":
        return base + " This low-confidence second-pass lane controls whether sub-70 accepts stay manual-only."
    if review_reason == "recall_unresolved_top_candidate":
        return base + " An accept here measures recoverable recall from unresolved rows."
    return base


def _label_decision_impact(review_reason: str, target: dict) -> str:
    target_rows = _to_int(target.get("target_decisive_rows"))
    needed_rows = _to_int(target.get("decisive_rows_needed"))
    sample_goal = f"{target_rows or needed_rows} decisive labels" if target_rows or needed_rows else "the required decisive labels"
    remaining_goal = f"{needed_rows} remaining decisive labels" if needed_rows else sample_goal
    if review_reason == "precision_second_pass_accepted_lt70":
        return (
            f"If {remaining_goal} are correct with zero reject/replace blockers, this lane can become a candidate for narrower review downgrade. "
            "Any reject/replace keeps sub-70 second-pass accepts protected and adds a regression case; unsure does not count as decisive evidence."
        )
    if review_reason == "precision_calibrated_pattern_release":
        return (
            "A reject/replace blocks wider automatic release for this pattern and turns the row into a regression case. "
            f"If the spot-check reaches {sample_goal} with zero blockers, the pattern remains a guarded release candidate."
        )
    if review_reason == "recall_unresolved_top_candidate":
        return (
            "Accept/replace labels measure recoverable recall from unresolved rows, but they do not by themselves lower global thresholds. "
            "Reject/unsure keeps the candidate manual-only and helps tune unresolved review priority."
        )
    if review_reason in {
        "precision_generic_identity_term_risk",
        "precision_low_confidence_auto_match",
        "precision_second_pass_accepted_70_84",
        "precision_slug_extension_identity_risk",
    }:
        return (
            f"If {sample_goal} are clean with no reject/replace blockers, this protected lane can be considered for a narrow routing downgrade after regression tests. "
            "Any wrong candidate keeps the lane protected and records the exact failure pattern for AgentA rules."
        )
    if review_reason.startswith("precision_"):
        return (
            "Wrong labels keep this precision lane protected and become regression examples. "
            "Clean decisive labels may support a narrow rule or routing change only after repeated evidence."
        )
    return (
        "This label updates calibration evidence. Decisive accept/replace/reject decisions affect reports; unsure preserves manual review until stronger evidence exists."
    )


def _label_evidence_source_kind(review_reason: str, status_summary: dict) -> str:
    if review_reason == "precision_calibrated_pattern_release":
        return str(status_summary.get("pattern_release_source_kind") or "")
    return ""


def _label_evidence_source_path(review_reason: str, status_summary: dict) -> str:
    if review_reason == "precision_calibrated_pattern_release":
        return str(status_summary.get("pattern_release_source_path") or "")
    return ""


def _fields(sample_headers: list[str]) -> list[str]:
    return [*LABEL_GAP_PREFIX_FIELDS, *[field for field in sample_headers if field not in LABEL_GAP_PREFIX_FIELDS]]


def _filled_row_keys(path: Path) -> set[tuple[str, str]]:
    if not path.exists():
        return set()
    return {
        _row_key(row)
        for row in _read_table(path)
        if _row_key(row) and str(row.get("manual_decision") or row.get("your_decision") or row.get("decision") or "").strip()
    }


def _row_key(row: dict[str, str]) -> tuple[str, str]:
    provider_id = str(row.get("provider_id") or "").strip()
    review_reason = str(row.get("review_reason") or "").strip()
    if not provider_id or not review_reason:
        return ("", "")
    return (provider_id, review_reason)


def _read_json(path: Path) -> dict:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def _read_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _read_table(path: Path) -> list[dict[str, str]]:
    if path.suffix.casefold() == ".xlsx":
        return _read_xlsx(path)
    return _read_rows(path)


def _read_xlsx(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    workbook = load_workbook(path, data_only=False, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows())
    if not rows:
        return []
    headers = [_cell_text(cell.value) for cell in rows[0]]
    out = []
    for cells in rows[1:]:
        row = {headers[idx]: _cell_text(cells[idx].value) for idx in range(len(headers)) if headers[idx]}
        if any(row.values()):
            out.append(row)
    return out


def _cell_text(value: object) -> str:
    return str(value or "").strip()


def _write_rows(path: Path, rows: list[dict[str, str]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    effective_fields = fields or [*LABEL_GAP_PREFIX_FIELDS, "message"]
    effective_rows = rows or []
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=effective_fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(effective_rows)


def _to_int(value: object) -> int:
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return 0


if __name__ == "__main__":
    raise SystemExit(main())
