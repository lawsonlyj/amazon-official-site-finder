from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from tools.build_linked_workbook import build_workbook


PROTECTED_LANE_PREFIX_FIELDS = [
    "protected_lane_priority",
    "protected_lane_recommendation",
    "protected_lane_label_gap_closed",
    "protected_lane_target_decisive_rows",
    "protected_lane_decisive_rows",
    "protected_lane_goal",
    "review_instruction",
    "optimization_use",
    "if_accept",
    "if_replace_or_reject",
    "if_unsure",
]

PROTECTED_LANE_RECOMMENDATIONS = {
    "keep_review_lane",
    "needs_more_labels",
    "candidate_for_review_downgrade",
    "candidate_for_narrow_recall_rule",
    "spot_check_candidate",
}


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Build the next protected-lane calibration review task.")
    parser.add_argument("--status-json", required=True, help="calibration_status.json from run_calibration_cycle.py.")
    parser.add_argument("--sample-csv", help="Optional sample CSV. Defaults to artifacts.sample_csv in status JSON.")
    parser.add_argument(
        "--filled-sample",
        action="append",
        default=[],
        help="Filled CSV/XLSX rows to exclude from this next task. Repeatable.",
    )
    parser.add_argument("--output-csv", required=True)
    parser.add_argument("--output-xlsx")
    parser.add_argument("--output-json", help="Optional summary JSON path.")
    parser.add_argument("--max-rows", type=int, default=50)
    parser.add_argument("--max-per-reason", type=int, default=12)
    parser.add_argument(
        "--include-validated-pattern-release",
        action="store_true",
        help="Include already validated precision_calibrated_pattern_release rows.",
    )
    args = parser.parse_args(argv)

    summary = build_protected_lane_review_task(
        status_json=args.status_json,
        sample_csv=args.sample_csv,
        filled_sample=args.filled_sample,
        output_csv=args.output_csv,
        output_xlsx=args.output_xlsx,
        output_json=args.output_json,
        max_rows=args.max_rows,
        max_per_reason=args.max_per_reason,
        include_validated_pattern_release=args.include_validated_pattern_release,
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


def build_protected_lane_review_task(
    *,
    status_json: str | Path,
    output_csv: str | Path,
    sample_csv: str | Path | None = None,
    filled_sample: str | Path | list[str | Path] | None = None,
    output_xlsx: str | Path | None = None,
    output_json: str | Path | None = None,
    max_rows: int = 50,
    max_per_reason: int = 12,
    include_validated_pattern_release: bool = False,
) -> dict:
    status_path = Path(status_json)
    status = _read_json(status_path)
    status_summary = status.get("summary", {})
    sample_value = sample_csv or status.get("artifacts", {}).get("sample_csv") or ""
    sample_path = Path(sample_value) if sample_value else Path()
    sample_rows = _read_rows(sample_path) if sample_value and sample_path.is_file() else []
    headers = list(sample_rows[0].keys()) if sample_rows else []
    filled_paths = _as_paths(filled_sample)
    filled_provider_ids, filled_row_keys = _filled_keys(filled_paths)
    target_by_reason = {
        str(target.get("review_reason") or ""): target
        for target in status.get("label_targets", [])
        if str(target.get("review_reason") or "")
    }
    include_reasons = {
        reason
        for reason, target in target_by_reason.items()
        if _include_target(target, status_summary, include_validated_pattern_release=include_validated_pattern_release)
    }

    candidates: list[dict[str, str]] = []
    excluded_filled = 0
    available_by_reason: Counter[str] = Counter()
    for row in sample_rows:
        reason = str(row.get("review_reason") or "").strip()
        if reason not in include_reasons:
            continue
        if _row_has_manual_decision(row) or _row_key(row) in filled_row_keys or str(row.get("provider_id") or "").strip() in filled_provider_ids:
            excluded_filled += 1
            continue
        target = target_by_reason.get(reason, {})
        available_by_reason[reason] += 1
        candidates.append(_protected_lane_row(row, target))

    selected = _select_rows(candidates, max_rows=max_rows, max_per_reason=max_per_reason)
    fields = _fields(headers)
    output_csv_path = Path(output_csv)
    _write_rows(output_csv_path, selected, fields)
    xlsx_summary = {}
    if output_xlsx:
        xlsx_summary = build_workbook([("Protected_Lanes", output_csv_path)], output_xlsx)

    target_summaries = _target_summaries(
        status.get("label_targets", []),
        include_reasons,
        available_by_reason,
        Counter(row.get("review_reason", "") for row in selected),
    )
    summary = {
        "status_json": str(status_path),
        "sample_csv": str(sample_path),
        "filled_samples": [str(path) for path in filled_paths],
        "output_csv": str(output_csv_path),
        "output_xlsx": str(output_xlsx or ""),
        "task_rows": len(selected),
        "max_rows": max_rows,
        "max_per_reason": max_per_reason,
        "excluded_filled_rows": excluded_filled,
        "filled_provider_ids": len(filled_provider_ids),
        "filled_row_keys": len(filled_row_keys),
        "reason_counts": dict(Counter(row.get("review_reason", "") for row in selected)),
        "priority_counts": dict(Counter(row.get("protected_lane_priority", "") for row in selected)),
        "agent_b_decision_counts": dict(Counter(row.get("agent_b_decision", "") for row in selected)),
        "targets": target_summaries,
        "xlsx": xlsx_summary,
    }
    if output_json:
        path = Path(output_json)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def _include_target(target: dict, status_summary: dict, *, include_validated_pattern_release: bool) -> bool:
    reason = str(target.get("review_reason") or "")
    if not reason:
        return False
    recommendation = str(target.get("recommendation") or "")
    priority = str(target.get("priority") or "")
    if (
        reason == "precision_calibrated_pattern_release"
        and not include_validated_pattern_release
        and str(status_summary.get("pattern_release_status") or "") == "current_guarded_candidate"
        and _to_int(target.get("decisive_rows_needed")) == 0
        and _to_int(target.get("blocking_rows")) == 0
    ):
        return False
    if recommendation in PROTECTED_LANE_RECOMMENDATIONS:
        return True
    return priority in {"high", "medium"} and (reason.startswith("precision_") or reason == "recall_unresolved_top_candidate")


def _protected_lane_row(row: dict[str, str], target: dict) -> dict[str, str]:
    out = dict(row)
    reason = str(row.get("review_reason") or target.get("review_reason") or "")
    decisive_needed = _to_int(target.get("decisive_rows_needed"))
    out["protected_lane_priority"] = str(target.get("priority") or "")
    out["protected_lane_recommendation"] = str(target.get("recommendation") or "")
    out["protected_lane_label_gap_closed"] = "yes" if decisive_needed == 0 else "no"
    out["protected_lane_target_decisive_rows"] = str(target.get("target_decisive_rows") or "")
    out["protected_lane_decisive_rows"] = str(target.get("decisive_rows") or "")
    out["protected_lane_goal"] = str(target.get("label_goal") or "")
    out["review_instruction"] = (
        "Fill manual_decision only: accept if the shown URL is correct, replace with manual_url if another official site is correct, "
        "reject if the shown URL is wrong/unproven, unsure only when evidence conflicts."
    )
    out["optimization_use"] = _optimization_use(reason, str(target.get("recommendation") or ""))
    out["if_accept"] = str(target.get("if_clean_action") or "")
    out["if_replace_or_reject"] = str(target.get("if_blocked_action") or "")
    out["if_unsure"] = str(target.get("if_unsure_action") or "")
    out["manual_decision"] = ""
    out["manual_url"] = ""
    out["notes"] = ""
    return out


def _optimization_use(review_reason: str, recommendation: str) -> str:
    if review_reason == "recall_unresolved_top_candidate":
        return (
            "Accept/replace labels are used to mine narrow recall patterns; reject/unsure labels keep this unresolved lane manual-only."
        )
    if review_reason == "precision_calibrated_pattern_release":
        return "Reject/replace blocks wider guarded pattern release; clean labels keep it as a sampled spot-check candidate."
    if recommendation == "candidate_for_review_downgrade":
        return "Clean labels become a downgrade candidate only after regression tests; wrong labels keep the lane protected."
    if review_reason.startswith("precision_"):
        return (
            "Reject/replace labels become precision regression fixtures and scoring/risk-rule evidence; clean labels may narrow review only after repeated evidence."
        )
    return "Labels update calibration evidence; single rows do not change thresholds or rules by themselves."


def _select_rows(rows: list[dict[str, str]], *, max_rows: int, max_per_reason: int) -> list[dict[str, str]]:
    selected: list[dict[str, str]] = []
    counts: Counter[str] = Counter()
    for row in sorted(rows, key=_sort_key):
        if max_rows > 0 and len(selected) >= max_rows:
            break
        reason = row.get("review_reason", "")
        if max_per_reason > 0 and counts[reason] >= max_per_reason:
            continue
        selected.append(row)
        counts[reason] += 1
    return selected


def _sort_key(row: dict[str, str]) -> tuple:
    priority_order = {"high": 0, "medium": 1, "normal": 2, "low": 3}
    recommendation_order = {
        "keep_review_lane": 0,
        "needs_more_labels": 1,
        "candidate_for_review_downgrade": 2,
        "candidate_for_narrow_recall_rule": 3,
        "spot_check_candidate": 4,
    }
    return (
        priority_order.get(row.get("protected_lane_priority", ""), 9),
        recommendation_order.get(row.get("protected_lane_recommendation", ""), 9),
        -_to_int(row.get("sample_priority")),
        row.get("review_reason", ""),
        row.get("provider_name", ""),
        row.get("provider_id", ""),
    )


def _target_summaries(
    targets: list[dict],
    include_reasons: set[str],
    available_by_reason: Counter[str],
    selected_by_reason: Counter[str],
) -> list[dict]:
    out = []
    for target in targets:
        reason = str(target.get("review_reason") or "")
        if reason not in include_reasons:
            continue
        out.append(
            {
                "review_reason": reason,
                "priority": str(target.get("priority") or ""),
                "recommendation": str(target.get("recommendation") or ""),
                "target_decisive_rows": _to_int(target.get("target_decisive_rows")),
                "decisive_rows": _to_int(target.get("decisive_rows")),
                "decisive_rows_needed": _to_int(target.get("decisive_rows_needed")),
                "available_unfilled_rows": available_by_reason.get(reason, 0),
                "selected_rows": selected_by_reason.get(reason, 0),
            }
        )
    return out


def _filled_keys(paths: list[Path]) -> tuple[set[str], set[tuple[str, str]]]:
    provider_ids: set[str] = set()
    row_keys: set[tuple[str, str]] = set()
    for path in paths:
        if not path.exists():
            continue
        for row in _read_table(path):
            if not _row_has_manual_decision(row):
                continue
            provider_id = str(row.get("provider_id") or "").strip()
            if provider_id:
                provider_ids.add(provider_id)
            key = _row_key(row)
            if key:
                row_keys.add(key)
    return provider_ids, row_keys


def _row_has_manual_decision(row: dict[str, str]) -> bool:
    return bool(str(row.get("manual_decision") or row.get("your_decision") or row.get("decision") or "").strip())


def _row_key(row: dict[str, str]) -> tuple[str, str]:
    provider_id = str(row.get("provider_id") or "").strip()
    review_reason = str(row.get("review_reason") or "").strip()
    if not provider_id or not review_reason:
        return ("", "")
    return (provider_id, review_reason)


def _fields(sample_headers: list[str]) -> list[str]:
    headers = list(sample_headers)
    for field in [
        "provider_id",
        "provider_name",
        "provider_detail_url",
        "official_url",
        "candidate_url",
        "review_reason",
        "manual_decision",
        "manual_url",
        "notes",
    ]:
        if field not in headers:
            headers.append(field)
    return [*PROTECTED_LANE_PREFIX_FIELDS, *[field for field in headers if field not in PROTECTED_LANE_PREFIX_FIELDS]]


def _read_json(path: Path) -> dict:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def _read_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _read_table(path: Path) -> list[dict[str, str]]:
    if path.suffix.casefold() == ".xlsx":
        return _read_xlsx(path)
    return _read_rows(path)


def _read_xlsx(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    workbook = load_workbook(path, data_only=False, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows())
    if not rows:
        return []
    headers = [_cell_text(cell.value) for cell in rows[0]]
    out = []
    for cells in rows[1:]:
        row = {headers[idx]: _cell_text(cells[idx].value) for idx in range(len(headers)) if headers[idx]}
        if any(row.values()):
            out.append(row)
    return out


def _cell_text(value: object) -> str:
    return str(value or "").strip()


def _write_rows(path: Path, rows: list[dict[str, str]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    effective_fields = fields or [*PROTECTED_LANE_PREFIX_FIELDS, "message"]
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=effective_fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _as_paths(value: str | Path | list[str | Path] | None) -> list[Path]:
    if value is None:
        return []
    if isinstance(value, (str, Path)):
        return [Path(value)]
    return [Path(item) for item in value if str(item)]


def _to_int(value: object) -> int:
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return 0


if __name__ == "__main__":
    raise SystemExit(main())
