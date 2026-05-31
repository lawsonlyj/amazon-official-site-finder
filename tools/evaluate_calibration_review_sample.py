from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from urllib.parse import urlparse

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))


DETAIL_FIELDS = [
    "provider_id",
    "provider_name",
    "sample_reason",
    "pattern_scope",
    "pattern_match",
    "review_reason",
    "agent_b_decision",
    "reason_for_unsure",
    "official_url",
    "candidate_url",
    "manual_decision",
    "manual_url",
    "normalized_decision",
    "normalized_manual_url",
    "decision_quality_issue",
    "lane_kind",
    "calibration_outcome",
]


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Evaluate filled high-value calibration review labels.")
    parser.add_argument("--sample", required=True, help="Filled calibration sample CSV/XLSX.")
    parser.add_argument("--output-json")
    parser.add_argument("--output-md")
    parser.add_argument("--output-csv", help="Optional row-level normalized calibration outcomes.")
    args = parser.parse_args(argv)

    report = evaluate_calibration_review_sample(
        sample=args.sample,
        output_json=args.output_json,
        output_md=args.output_md,
        output_csv=args.output_csv,
    )
    print(json.dumps(report["summary"], ensure_ascii=False, indent=2))
    return 0


def evaluate_calibration_review_sample(
    *,
    sample: str | Path,
    output_json: str | Path | None = None,
    output_md: str | Path | None = None,
    output_csv: str | Path | None = None,
) -> dict:
    rows = _read_table(Path(sample))
    details = [_detail(row) for row in rows]
    labeled = _valid_labeled(details)
    decisive = _valid_decisive(details)
    summary = {
        "sample_rows": len(details),
        "labeled_rows": len(labeled),
        "decisive_rows": len(decisive),
        "manual_decision_counts": dict(Counter(row["normalized_decision"] for row in labeled)),
        "invalid_manual_decision_rows": sum(1 for row in details if row["decision_quality_issue"] == "invalid_manual_decision"),
        "replace_missing_manual_url_rows": sum(1 for row in details if row["decision_quality_issue"] == "replace_missing_manual_url"),
        "decision_quality_issue_rows": sum(1 for row in details if row["decision_quality_issue"]),
        "candidate_correct_rows": sum(1 for row in details if row["calibration_outcome"] == "candidate_correct"),
        "candidate_incorrect_rows": sum(1 for row in details if row["calibration_outcome"] == "candidate_incorrect"),
        "recall_useful_rows": sum(1 for row in details if row["calibration_outcome"] == "recall_candidate_useful"),
        "recall_not_useful_rows": sum(1 for row in details if row["calibration_outcome"] == "recall_candidate_not_useful"),
        "unsure_rows": sum(1 for row in details if row["calibration_outcome"] == "manual_unsure"),
    }
    pattern_recommendations = _pattern_recommendations(details)
    pattern_rule_candidates = _pattern_rule_candidates(pattern_recommendations)
    lane_recommendations = _lane_recommendations(details)
    summary["pattern_rule_candidate_rows"] = len(pattern_rule_candidates["candidate_for_rule"])
    summary["pattern_rejected_rows"] = len(pattern_rule_candidates["reject_pattern"])
    summary["pattern_needs_more_label_rows"] = len(pattern_rule_candidates["needs_more_labels"])
    lane_counts = Counter(row["recommendation"] for row in lane_recommendations)
    summary["lane_keep_review_rows"] = lane_counts.get("keep_review_lane", 0)
    summary["lane_candidate_for_change_rows"] = lane_counts.get("candidate_for_review_downgrade", 0) + lane_counts.get(
        "candidate_for_narrow_recall_rule", 0
    )
    summary["lane_needs_more_label_rows"] = lane_counts.get("needs_more_labels", 0)
    report = {
        "summary": summary,
        "by_sample_reason": _group_stats(details, "sample_reason"),
        "by_review_reason": _group_stats(details, "review_reason"),
        "by_agent_b_decision": _group_stats(details, "agent_b_decision"),
        "by_pattern_match": _group_stats(details, "pattern_match"),
        "lane_recommendations": lane_recommendations,
        "pattern_recommendations": pattern_recommendations,
        "pattern_rule_candidates": pattern_rule_candidates,
        "recommendations": _recommendations(details),
        "details": details,
        "inputs": {"sample": str(sample)},
    }
    if output_json:
        path = Path(output_json)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    if output_md:
        path = Path(output_md)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(_render_markdown(report), encoding="utf-8")
    if output_csv:
        _write_rows(Path(output_csv), details, DETAIL_FIELDS)
    return report


def _detail(row: dict[str, str]) -> dict[str, str]:
    decision = _decision(row)
    lane_kind = _lane_kind(row)
    manual_decision = _first(row, "manual_decision", "your_decision", "decision")
    normalized_manual_url = _normalize_url(_first(row, "manual_url", "your_true_official_url", "true_official_url"))
    issue = _decision_quality_issue(manual_decision, decision, normalized_manual_url)
    outcome = "fill_quality_issue" if issue else _outcome(decision, lane_kind)
    return {
        "provider_id": _first(row, "provider_id"),
        "provider_name": _first(row, "provider_name"),
        "sample_reason": _first(row, "sample_reason"),
        "pattern_scope": _first(row, "pattern_scope"),
        "pattern_match": _first(row, "pattern_match"),
        "review_reason": _first(row, "review_reason"),
        "agent_b_decision": _first(row, "agent_b_decision"),
        "reason_for_unsure": _first(row, "reason_for_unsure"),
        "official_url": _normalize_url(_first(row, "official_url")),
        "candidate_url": _normalize_url(_first(row, "candidate_url", "current_or_candidate_url", "official_url")),
        "manual_decision": manual_decision,
        "manual_url": _first(row, "manual_url", "your_true_official_url", "true_official_url"),
        "normalized_decision": decision,
        "normalized_manual_url": normalized_manual_url,
        "decision_quality_issue": issue,
        "lane_kind": lane_kind,
        "calibration_outcome": outcome,
    }


def _lane_kind(row: dict[str, str]) -> str:
    sample_reason = _first(row, "sample_reason")
    review_reason = _first(row, "review_reason")
    if review_reason == "recall_unresolved_top_candidate" or sample_reason == "recall_candidate_label":
        return "recall"
    if review_reason.startswith("precision_") or sample_reason in {
        "agent_b_accept_risky_lane",
        "agent_b_reject_check",
        "generic_identity_label",
        "slug_extension_label",
        "second_pass_threshold_label",
        "low_confidence_label",
        "timeout_needs_manual",
    }:
        return "precision"
    return "general"


def _outcome(decision: str, lane_kind: str) -> str:
    if not decision:
        return "unlabeled"
    if decision == "unsure":
        return "manual_unsure"
    if lane_kind == "recall":
        if decision in {"accept", "replace"}:
            return "recall_candidate_useful"
        return "recall_candidate_not_useful"
    if decision == "accept":
        return "candidate_correct"
    if decision in {"replace", "reject"}:
        return "candidate_incorrect"
    return "unlabeled"


def _group_stats(rows: list[dict[str, str]], field: str) -> dict[str, dict]:
    out: dict[str, dict] = {}
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        grouped[row.get(field, "") or "(blank)"].append(row)
    for key, items in sorted(grouped.items()):
        labeled = _valid_labeled(items)
        decisive = _valid_decisive(items)
        decisions = Counter(row["normalized_decision"] for row in labeled)
        outcomes = Counter(row["calibration_outcome"] for row in items)
        out[key] = {
            "rows": len(items),
            "labeled_rows": len(labeled),
            "decisive_rows": len(decisive),
            "decision_counts": dict(decisions),
            "outcome_counts": dict(outcomes),
            "candidate_correct_rate": _ratio(outcomes.get("candidate_correct", 0), len(decisive)),
            "candidate_incorrect_rate": _ratio(outcomes.get("candidate_incorrect", 0), len(decisive)),
            "recall_useful_rate": _ratio(outcomes.get("recall_candidate_useful", 0), len(decisive)),
        }
    return out


def _valid_labeled(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    return [row for row in rows if row["normalized_decision"] and not row["decision_quality_issue"]]


def _valid_decisive(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    return [row for row in _valid_labeled(rows) if row["normalized_decision"] != "unsure"]


def _recommendations(details: list[dict[str, str]]) -> list[str]:
    recommendations: list[str] = []
    quality_issues = [row for row in details if row["decision_quality_issue"]]
    if quality_issues:
        recommendations.append(
            "Fix calibration fill-quality issues before applying threshold or rule changes; invalid decisions and replace rows without manual_url can distort calibration."
        )
    labeled = _valid_labeled(details)
    decisive = _valid_decisive(details)
    if not labeled:
        recommendations.append(
            "No valid filled calibration labels yet. Fill manual_decision, manual_url, and notes before changing thresholds or review lanes."
        )
        return recommendations
    precision_rows = [row for row in decisive if row["lane_kind"] == "precision"]
    precision_bad = [row for row in precision_rows if row["calibration_outcome"] == "candidate_incorrect"]
    recall_rows = [row for row in decisive if row["lane_kind"] == "recall"]
    recall_useful = [row for row in recall_rows if row["calibration_outcome"] == "recall_candidate_useful"]
    agent_b_accepts = [
        row for row in decisive if row["agent_b_decision"] == "accept" or row["sample_reason"] == "agent_b_accept_risky_lane"
    ]
    risky_accept_bad = [row for row in agent_b_accepts if row["calibration_outcome"] in {"candidate_incorrect", "recall_candidate_not_useful"}]
    if risky_accept_bad:
        recommendations.append(
            "Keep Check and Suggestion risky accepts in manual review; human labels still show incorrect accepted candidates in risky lanes."
        )
    elif len(agent_b_accepts) >= 10:
        recommendations.append(
            "Check and Suggestion risky accepts had no labeled corrections in this sample; consider a narrow release rule only for the exact evidence pattern."
        )

    if precision_bad:
        recommendations.append(
            "Do not globally lower acceptance thresholds yet; precision lanes still contain bad official-site candidates."
        )
    elif len(precision_rows) >= 10:
        recommendations.append(
            "Precision-lane labels show no bad candidates in this sample; consider narrowing the reviewed lane rather than changing the global threshold."
        )

    generic_or_slug = [
        row
        for row in decisive
        if row["review_reason"] in {"precision_generic_identity_term_risk", "precision_slug_extension_identity_risk"}
    ]
    generic_or_slug_bad = [row for row in generic_or_slug if row["calibration_outcome"] == "candidate_incorrect"]
    generic_or_slug_good = [row for row in generic_or_slug if row["calibration_outcome"] == "candidate_correct"]
    if generic_or_slug_bad:
        recommendations.append("Keep generic-name and slug-extension identity constraints; the sample still has same-name/domain-shape mistakes.")
    elif len(generic_or_slug_good) >= 5:
        recommendations.append(
            "Generic-name and slug-extension labels are mostly correct; consider requiring only manual review when service/country evidence is also weak."
        )

    if recall_useful:
        recommendations.append(
            "Add recall examples from accepted/replaced unresolved rows to query and low-score strong-identity tests instead of lowering the global threshold."
        )
    if recall_rows and _ratio(len(recall_useful), len(recall_rows)) is not None and len(recall_useful) < len(recall_rows):
        recommendations.append("Keep unresolved recall rows as human/Check and Suggestion evidence only; not every top candidate is useful.")

    timeout_rows = [row for row in decisive if row["sample_reason"] == "timeout_needs_manual" or row["reason_for_unsure"] == "agent_b_row_timeout"]
    timeout_useful = [
        row
        for row in timeout_rows
        if row["calibration_outcome"] in {"candidate_correct", "recall_candidate_useful"}
    ]
    if timeout_rows and _ratio(len(timeout_useful), len(timeout_rows)) and _ratio(len(timeout_useful), len(timeout_rows)) >= 0.5:
        recommendations.append("Retry Check and Suggestion timeout rows with resume/longer timeout before manual review; many timed-out rows are useful candidates.")
    elif timeout_rows:
        recommendations.append("Keep timeout rows in manual review priority; current labels do not justify auto-accepting timed-out candidates.")

    if not recommendations:
        recommendations.append("Labels are mixed or sparse; keep current threshold and collect more calibration rows before changing rules.")
    return recommendations


def _pattern_recommendations(details: list[dict[str, str]]) -> list[dict]:
    groups: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in details:
        if row.get("pattern_match"):
            groups[row["pattern_match"]].append(row)
    out = []
    for pattern, rows in sorted(groups.items()):
        labeled = _valid_labeled(rows)
        decisive = _valid_decisive(rows)
        good = [
            row
            for row in decisive
            if row["calibration_outcome"] in {"candidate_correct", "recall_candidate_useful"}
        ]
        bad = [
            row
            for row in decisive
            if row["calibration_outcome"] in {"candidate_incorrect", "recall_candidate_not_useful"}
        ]
        if bad:
            recommendation = "reject_pattern"
            reason = "Human labels found at least one wrong candidate for this pattern."
        elif len(good) >= 5:
            recommendation = "candidate_for_rule"
            reason = "Five or more decisive labels supported this pattern with no wrong candidates."
        elif good:
            recommendation = "needs_more_labels"
            reason = "Current labels support this pattern, but support is still too small for a production rule."
        else:
            recommendation = "unlabeled"
            reason = "No decisive human labels for this pattern yet."
        out.append(
            {
                "pattern": pattern,
                "pattern_scope": _first(rows[0], "pattern_scope"),
                "rows": len(rows),
                "labeled_rows": len(labeled),
                "decisive_rows": len(decisive),
                "supporting_rows": len(good),
                "blocking_rows": len(bad),
                "support_rate": _ratio(len(good), len(decisive)),
                "support_rate_wilson_lower_80": _wilson_interval(len(good), len(decisive))["lower"],
                "blocking_rate_wilson_upper_80": _wilson_interval(len(bad), len(decisive))["upper"],
                "evidence_strength": _evidence_strength(len(decisive), len(good), len(bad)),
                "recommendation": recommendation,
                "reason": reason,
            }
        )
    out.sort(
        key=lambda row: (
            {"reject_pattern": 0, "candidate_for_rule": 1, "needs_more_labels": 2, "unlabeled": 3}.get(
                row["recommendation"], 9
            ),
            -row["decisive_rows"],
            row["pattern"],
        )
    )
    return out


def _lane_recommendations(details: list[dict[str, str]]) -> list[dict]:
    groups: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in details:
        if row.get("review_reason"):
            groups[row["review_reason"]].append(row)
    out = []
    for reason, rows in sorted(groups.items()):
        labeled = _valid_labeled(rows)
        decisive = _valid_decisive(rows)
        outcomes = Counter(row["calibration_outcome"] for row in decisive)
        lane_kind = _first(rows[0], "lane_kind")
        support_rows = _lane_support_rows(lane_kind, outcomes)
        blocking_rows = _lane_blocking_rows(lane_kind, outcomes)
        support_interval = _wilson_interval(support_rows, len(decisive))
        blocking_interval = _wilson_interval(blocking_rows, len(decisive))
        recommendation, explanation, required_action = _lane_decision(reason, lane_kind, outcomes, len(decisive))
        out.append(
            {
                "review_reason": reason,
                "lane_kind": lane_kind,
                "rows": len(rows),
                "labeled_rows": len(labeled),
                "decisive_rows": len(decisive),
                "candidate_correct_rows": outcomes.get("candidate_correct", 0),
                "candidate_incorrect_rows": outcomes.get("candidate_incorrect", 0),
                "recall_useful_rows": outcomes.get("recall_candidate_useful", 0),
                "recall_not_useful_rows": outcomes.get("recall_candidate_not_useful", 0),
                "manual_unsure_rows": Counter(row["calibration_outcome"] for row in labeled).get("manual_unsure", 0),
                "support_rows": support_rows,
                "blocking_rows": blocking_rows,
                "support_rate": _ratio(support_rows, len(decisive)),
                "support_rate_wilson_lower_80": support_interval["lower"],
                "blocking_rate_wilson_upper_80": blocking_interval["upper"],
                "evidence_strength": _evidence_strength(len(decisive), support_rows, blocking_rows),
                "recommendation": recommendation,
                "reason": explanation,
                "required_action": required_action,
            }
        )
    out.sort(
        key=lambda row: (
            {"keep_review_lane": 0, "candidate_for_review_downgrade": 1, "candidate_for_narrow_recall_rule": 2, "needs_more_labels": 3}.get(
                row["recommendation"], 9
            ),
            -row["decisive_rows"],
            row["review_reason"],
        )
    )
    return out


def _lane_support_rows(lane_kind: str, outcomes: Counter[str]) -> int:
    if lane_kind == "recall":
        return outcomes.get("recall_candidate_useful", 0)
    return outcomes.get("candidate_correct", 0)


def _lane_blocking_rows(lane_kind: str, outcomes: Counter[str]) -> int:
    if lane_kind == "recall":
        return outcomes.get("recall_candidate_not_useful", 0)
    return outcomes.get("candidate_incorrect", 0)


def _lane_decision(reason: str, lane_kind: str, outcomes: Counter[str], decisive_rows: int) -> tuple[str, str, str]:
    candidate_bad = outcomes.get("candidate_incorrect", 0)
    candidate_good = outcomes.get("candidate_correct", 0)
    recall_useful = outcomes.get("recall_candidate_useful", 0)
    recall_bad = outcomes.get("recall_candidate_not_useful", 0)
    if decisive_rows == 0:
        return (
            "needs_more_labels",
            "No decisive labels for this review lane yet.",
            "Keep this lane in calibration samples before changing routing.",
        )
    if lane_kind == "recall":
        if recall_bad:
            return (
                "keep_review_lane",
                "At least one labeled recall candidate was not useful.",
                "Keep this recall lane manual-only and use good rows as evidence, not broad auto-release.",
            )
        if recall_useful >= 5:
            return (
                "candidate_for_narrow_recall_rule",
                "Five or more decisive recall labels were useful with no blockers.",
                "Mine the exact supporting evidence pattern and add regression tests before any recall release rule.",
            )
        return (
            "needs_more_labels",
            "Recall labels are supportive but below the five-label promotion threshold.",
            "Keep this lane in calibration samples until it reaches five useful decisive labels with zero blockers.",
        )
    if candidate_bad:
        return (
            "keep_review_lane",
            "At least one labeled candidate in this precision lane was wrong.",
            "Keep this precision lane in manual review and add blocking regression fixtures for wrong rows.",
        )
    if candidate_good >= 5:
        action = "Consider downgrading this lane to spot-check only after adding regression tests for the clean evidence pattern."
        if reason == "precision_calibrated_pattern_release":
            action = "Keep this as a sampled spot-check lane until it stays clean across another labeled batch."
        return (
            "candidate_for_review_downgrade",
            "Five or more decisive precision labels were correct with no blockers.",
            action,
        )
    return (
        "needs_more_labels",
        "Precision labels are clean so far but below the five-label promotion threshold.",
        "Keep this lane in calibration samples before reducing manual review.",
    )


def _pattern_rule_candidates(pattern_recommendations: list[dict]) -> dict[str, list[dict]]:
    buckets = {
        "candidate_for_rule": [],
        "needs_more_labels": [],
        "reject_pattern": [],
    }
    for item in pattern_recommendations:
        recommendation = item.get("recommendation", "")
        if recommendation not in buckets:
            continue
        enriched = dict(item)
        enriched["required_action"] = _required_action_for_pattern(item)
        buckets[recommendation].append(enriched)
    return buckets


def _required_action_for_pattern(item: dict) -> str:
    recommendation = item.get("recommendation", "")
    scope = item.get("pattern_scope", "")
    if recommendation == "candidate_for_rule":
        if scope == "recall":
            return "Add regression tests, then consider a narrow recall recovery rule for this exact evidence pattern; treat five clean labels as minimum support, not final proof."
        if scope == "precision":
            return "Add regression tests, then consider narrowing manual-review routing for this exact precision pattern; treat five clean labels as minimum support, not final proof."
        return "Add regression tests before any production rule change; treat five clean labels as minimum support, not final proof."
    if recommendation == "needs_more_labels":
        return "Keep this pattern in calibration samples until it reaches five decisive supporting labels with zero blockers."
    if recommendation == "reject_pattern":
        return "Do not release this pattern automatically; keep matching rows in manual review or add a blocking regression fixture."
    return "No action."


def _decision(row: dict[str, str]) -> str:
    raw = _first(row, "manual_decision", "your_decision", "decision").casefold()
    aliases = {
        "accept": "accept",
        "accepted": "accept",
        "approve": "accept",
        "approved": "accept",
        "correct": "accept",
        "yes": "accept",
        "true": "accept",
        "正确": "accept",
        "对": "accept",
        "replace": "replace",
        "replacement": "replace",
        "修正": "replace",
        "替换": "replace",
        "reject": "reject",
        "rejected": "reject",
        "no": "reject",
        "false": "reject",
        "wrong": "reject",
        "incorrect": "reject",
        "错误": "reject",
        "错": "reject",
        "无官网": "reject",
        "unsure": "unsure",
        "unknown": "unsure",
        "uncertain": "unsure",
        "不确定": "unsure",
    }
    return aliases.get(raw, raw if raw in {"accept", "replace", "reject", "unsure"} else "")


def _decision_quality_issue(manual_decision: str, normalized_decision: str, normalized_manual_url: str) -> str:
    if manual_decision.strip() and not normalized_decision:
        return "invalid_manual_decision"
    if normalized_decision == "replace" and not normalized_manual_url:
        return "replace_missing_manual_url"
    return ""


def _read_table(path: Path) -> list[dict[str, str]]:
    if path.suffix.casefold() == ".xlsx":
        return _read_xlsx(path)
    return _read_rows(path)


def _read_xlsx(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    workbook = load_workbook(path, data_only=False, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows())
    if not rows:
        return []
    headers = [_cell_text(cell.value) for cell in rows[0]]
    out = []
    for cells in rows[1:]:
        row = {headers[idx]: _cell_text(cells[idx].value) for idx in range(len(headers)) if headers[idx]}
        if any(row.values()):
            out.append(row)
    return out


def _read_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _write_rows(path: Path, rows: list[dict[str, str]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _render_markdown(report: dict) -> str:
    summary = report["summary"]
    lines = [
        "# Calibration Review Evaluation",
        "",
        "## Summary",
        "",
        f"- Sample rows: {summary['sample_rows']}",
        f"- Labeled rows: {summary['labeled_rows']}",
        f"- Decisive rows: {summary['decisive_rows']}",
        f"- Decision quality issue rows: {summary['decision_quality_issue_rows']}",
        f"- Invalid manual decisions: {summary['invalid_manual_decision_rows']}",
        f"- Replace rows missing manual_url: {summary['replace_missing_manual_url_rows']}",
        f"- Candidate correct rows: {summary['candidate_correct_rows']}",
        f"- Candidate incorrect rows: {summary['candidate_incorrect_rows']}",
        f"- Recall useful rows: {summary['recall_useful_rows']}",
        f"- Recall not useful rows: {summary['recall_not_useful_rows']}",
        "",
        "## Recommendations",
        "",
    ]
    for item in report["recommendations"]:
        lines.append(f"- {item}")
    lines.extend(["", "## Sample Reasons", ""])
    for reason, stats in report["by_sample_reason"].items():
        decisions = ", ".join(f"{key}={value}" for key, value in stats["decision_counts"].items()) or "none"
        outcomes = ", ".join(f"{key}={value}" for key, value in stats["outcome_counts"].items()) or "none"
        lines.append(f"- {reason}: rows={stats['rows']}, labeled={stats['labeled_rows']}, decisions=({decisions}), outcomes=({outcomes})")
    lines.extend(["", "## Review Reasons", ""])
    for reason, stats in report["by_review_reason"].items():
        outcomes = ", ".join(f"{key}={value}" for key, value in stats["outcome_counts"].items()) or "none"
        lines.append(f"- {reason}: rows={stats['rows']}, labeled={stats['labeled_rows']}, outcomes=({outcomes})")
    if report.get("lane_recommendations"):
        lines.extend(["", "## Review Lane Guidance", ""])
        for row in report["lane_recommendations"]:
            lines.append(
                "- {recommendation}: rows={rows}, decisive={decisive}, good={good}, bad={bad}, recall_useful={recall_useful}, recall_bad={recall_bad}, strength={strength}, support_lower80={support_lower}, block_upper80={block_upper} :: {reason}".format(
                    recommendation=row["recommendation"],
                    rows=row["rows"],
                    decisive=row["decisive_rows"],
                    good=row["candidate_correct_rows"],
                    bad=row["candidate_incorrect_rows"],
                    recall_useful=row["recall_useful_rows"],
                    recall_bad=row["recall_not_useful_rows"],
                    strength=row.get("evidence_strength", ""),
                    support_lower=row.get("support_rate_wilson_lower_80"),
                    block_upper=row.get("blocking_rate_wilson_upper_80"),
                    reason=row["review_reason"],
                )
            )
            lines.append(f"  Action: {row['required_action']}")
    if report.get("pattern_recommendations"):
        lines.extend(["", "## Pattern Validation", ""])
        for row in report["pattern_recommendations"][:25]:
            lines.append(
                "- {recommendation}: scope={scope}, rows={rows}, decisive={decisive}, support={support}, block={block}, strength={strength}, support_lower80={support_lower}, block_upper80={block_upper} :: {pattern}".format(
                    recommendation=row["recommendation"],
                    scope=row.get("pattern_scope", ""),
                    rows=row["rows"],
                    decisive=row["decisive_rows"],
                    support=row["supporting_rows"],
                    block=row["blocking_rows"],
                    strength=row.get("evidence_strength", ""),
                    support_lower=row.get("support_rate_wilson_lower_80"),
                    block_upper=row.get("blocking_rate_wilson_upper_80"),
                    pattern=row["pattern"],
                )
            )
    rule_candidates = report.get("pattern_rule_candidates", {})
    if rule_candidates:
        lines.extend(["", "## Candidate Rule Export", ""])
        for key, title in [
            ("candidate_for_rule", "Candidate For Rule"),
            ("needs_more_labels", "Needs More Labels"),
            ("reject_pattern", "Rejected Pattern"),
        ]:
            items = rule_candidates.get(key) or []
            lines.append(f"### {title}")
            if not items:
                lines.append("- None")
                continue
            for row in items[:10]:
                lines.append(
                    "- scope={scope}, support={support}, block={block}: {pattern} -- {action}".format(
                        scope=row.get("pattern_scope", ""),
                        support=row.get("supporting_rows"),
                        block=row.get("blocking_rows"),
                        pattern=row.get("pattern"),
                        action=row.get("required_action", ""),
                    )
                )
    lines.append("")
    return "\n".join(lines)


def _first(row: dict[str, str], *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip():
            return _cell_text(value)
    return ""


def _cell_text(value: object) -> str:
    text = str(value or "").strip()
    if text.startswith("="):
        text = text[1:]
    if text.upper().startswith("HYPERLINK("):
        match = re.search(r'HYPERLINK\("([^"]+)"', text, flags=re.IGNORECASE)
        return match.group(1).strip() if match else ""
    return text


def _normalize_url(value: object) -> str:
    raw = str(value or "").strip().replace("\xa0", "").rstrip(".,);]")
    if not raw:
        return ""
    parsed = urlparse(raw if "://" in raw else f"https://{raw}")
    if not parsed.netloc:
        return ""
    return f"{parsed.scheme or 'https'}://{parsed.netloc}{parsed.path or ''}".rstrip("/")


def _ratio(num: int, den: int) -> float | None:
    return round(num / den, 4) if den else None


def _wilson_interval(successes: int, total: int, z: float = 1.2815515655446004) -> dict[str, float | None]:
    if total <= 0:
        return {"lower": None, "upper": None}
    p = successes / total
    denom = 1 + (z * z / total)
    center = (p + (z * z / (2 * total))) / denom
    margin = (z * ((p * (1 - p) / total + z * z / (4 * total * total)) ** 0.5)) / denom
    return {
        "lower": round(max(0.0, center - margin), 4),
        "upper": round(min(1.0, center + margin), 4),
    }


def _evidence_strength(decisive_rows: int, support_rows: int, blocking_rows: int) -> str:
    if decisive_rows <= 0:
        return "unlabeled"
    if blocking_rows:
        return "blocked"
    if support_rows < 5:
        return "thin_support"
    if decisive_rows < 10:
        return "minimum_support"
    return "strong_support"


if __name__ == "__main__":
    raise SystemExit(main())
