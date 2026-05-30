from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from urllib.parse import urlparse

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

DETAIL_FIELDS = [
    "provider_id",
    "provider_name",
    "candidate_policy_action",
    "candidate_policy_pattern",
    "candidate_policy_source",
    "candidate_url",
    "manual_decision",
    "manual_url",
    "normalized_decision",
    "normalized_manual_url",
    "decision_quality_issue",
    "policy_outcome",
    "policy_supports_rule",
    "policy_blocks_rule",
    "review_reason",
    "agent_b_decision",
    "known_label_status",
    "notes",
]


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Evaluate a filled policy validation task.")
    parser.add_argument("--task", required=True, help="Filled policy_validation_task CSV/XLSX.")
    parser.add_argument("--output-json")
    parser.add_argument("--output-md")
    parser.add_argument("--output-csv")
    args = parser.parse_args(argv)

    report = evaluate_policy_validation_task(
        task=args.task,
        output_json=args.output_json,
        output_md=args.output_md,
        output_csv=args.output_csv,
    )
    print(json.dumps(report["summary"], ensure_ascii=False, indent=2))
    return 0


def evaluate_policy_validation_task(
    *,
    task: str | Path,
    output_json: str | Path | None = None,
    output_md: str | Path | None = None,
    output_csv: str | Path | None = None,
) -> dict:
    rows = _read_table(Path(task))
    details = [_detail(row) for row in rows]
    decisive = _decisive(details)
    summary = {
        "task_rows": len(details),
        "labeled_rows": len(_labeled(details)),
        "decisive_rows": len(decisive),
        "manual_decision_counts": dict(Counter(row["normalized_decision"] for row in _labeled(details))),
        "decision_quality_issue_rows": sum(1 for row in details if row["decision_quality_issue"]),
        "invalid_manual_decision_rows": sum(1 for row in details if row["decision_quality_issue"] == "invalid_manual_decision"),
        "replace_missing_manual_url_rows": sum(1 for row in details if row["decision_quality_issue"] == "replace_missing_manual_url"),
        "support_rows": sum(1 for row in details if row["policy_supports_rule"] == "1"),
        "blocking_rows": sum(1 for row in details if row["policy_blocks_rule"] == "1"),
        "holdout_supported_rows": sum(1 for row in details if row["policy_outcome"] == "holdout_supported"),
        "holdout_blocked_rows": sum(1 for row in details if row["policy_outcome"] == "holdout_blocked"),
        "release_supported_rows": sum(1 for row in details if row["policy_outcome"] == "release_supported"),
        "release_blocked_rows": sum(1 for row in details if row["policy_outcome"] == "release_blocked"),
        "manual_unsure_rows": sum(1 for row in details if row["policy_outcome"] == "manual_unsure"),
    }
    pattern_recommendations = _pattern_recommendations(details)
    rule_candidates = _rule_candidate_buckets(pattern_recommendations)
    summary["candidate_for_rule_rows"] = len(rule_candidates["candidate_for_rule"])
    summary["needs_more_labels_rows"] = len(rule_candidates["needs_more_labels"])
    summary["reject_pattern_rows"] = len(rule_candidates["reject_pattern"])
    report = {
        "summary": summary,
        "by_action": _group_stats(details, "candidate_policy_action"),
        "by_pattern": _group_stats(details, "candidate_policy_pattern"),
        "pattern_recommendations": pattern_recommendations,
        "policy_rule_candidates": rule_candidates,
        "recommendations": _recommendations(details, pattern_recommendations),
        "details": details,
        "inputs": {"task": str(task)},
    }
    if output_json:
        path = Path(output_json)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    if output_md:
        path = Path(output_md)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(_render_markdown(report), encoding="utf-8")
    if output_csv:
        _write_rows(Path(output_csv), details, DETAIL_FIELDS)
    return report


def _detail(row: dict[str, str]) -> dict[str, str]:
    decision = _decision(row)
    manual_decision = _first(row, "manual_decision", "your_decision", "decision")
    manual_url = _first(row, "manual_url", "your_true_official_url", "true_official_url")
    normalized_manual_url = _normalize_url(manual_url)
    issue = _decision_quality_issue(manual_decision, decision, normalized_manual_url)
    outcome = "fill_quality_issue" if issue else _policy_outcome(row, decision, normalized_manual_url)
    supports = outcome in {"holdout_supported", "release_supported"}
    blocks = outcome in {"holdout_blocked", "release_blocked"}
    return {
        "provider_id": _first(row, "provider_id"),
        "provider_name": _first(row, "provider_name"),
        "candidate_policy_action": _first(row, "candidate_policy_action"),
        "candidate_policy_pattern": _first(row, "candidate_policy_pattern"),
        "candidate_policy_source": _first(row, "candidate_policy_source"),
        "candidate_url": _normalize_url(_first(row, "candidate_url", "current_official_url", "official_url")),
        "manual_decision": manual_decision,
        "manual_url": manual_url,
        "normalized_decision": decision,
        "normalized_manual_url": normalized_manual_url,
        "decision_quality_issue": issue,
        "policy_outcome": outcome,
        "policy_supports_rule": "1" if supports else "",
        "policy_blocks_rule": "1" if blocks else "",
        "review_reason": _first(row, "review_reason"),
        "agent_b_decision": _first(row, "agent_b_decision"),
        "known_label_status": _first(row, "known_label_status"),
        "notes": _first(row, "notes"),
    }


def _policy_outcome(row: dict[str, str], decision: str, normalized_manual_url: str) -> str:
    if not decision:
        return "unlabeled"
    if decision == "unsure":
        return "manual_unsure"
    action = _first(row, "candidate_policy_action").casefold()
    if action == "holdout":
        if decision == "accept":
            return "holdout_blocked"
        return "holdout_supported"
    if action == "release":
        candidate_url = _normalize_url(_first(row, "candidate_url", "current_official_url", "official_url"))
        if decision == "accept":
            return "release_supported"
        if decision == "replace" and _same_site(candidate_url, normalized_manual_url):
            return "release_supported"
        return "release_blocked"
    return "unlabeled"


def _pattern_recommendations(details: list[dict[str, str]]) -> list[dict]:
    groups: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in details:
        pattern = row.get("candidate_policy_pattern", "")
        if pattern:
            groups[pattern].append(row)
    out = []
    for pattern, rows in sorted(groups.items()):
        decisive = _decisive(rows)
        support = [row for row in decisive if row["policy_supports_rule"] == "1"]
        blocking = [row for row in decisive if row["policy_blocks_rule"] == "1"]
        if blocking:
            recommendation = "reject_pattern"
            reason = "At least one filled validation row blocks this policy pattern."
        elif len(support) >= 5:
            recommendation = "candidate_for_rule"
            reason = "Five or more decisive labels support this policy pattern with no blockers."
        elif support:
            recommendation = "needs_more_labels"
            reason = "Current labels support this policy pattern, but support is still below the five-label promotion threshold."
        else:
            recommendation = "unlabeled"
            reason = "No decisive validation labels for this policy pattern yet."
        action_counts = Counter(row["candidate_policy_action"] for row in rows)
        out.append(
            {
                "pattern": pattern,
                "actions": dict(action_counts),
                "rows": len(rows),
                "labeled_rows": len(_labeled(rows)),
                "decisive_rows": len(decisive),
                "supporting_rows": len(support),
                "blocking_rows": len(blocking),
                "support_rate": _ratio(len(support), len(decisive)),
                "support_rate_wilson_lower_80": _wilson_interval(len(support), len(decisive))["lower"],
                "blocking_rate_wilson_upper_80": _wilson_interval(len(blocking), len(decisive))["upper"],
                "evidence_strength": _evidence_strength(len(decisive), len(support), len(blocking)),
                "recommendation": recommendation,
                "reason": reason,
                "supporting_provider_ids": [row["provider_id"] for row in support],
                "blocking_provider_ids": [row["provider_id"] for row in blocking],
            }
        )
    out.sort(
        key=lambda row: (
            {"reject_pattern": 0, "candidate_for_rule": 1, "needs_more_labels": 2, "unlabeled": 3}.get(
                row["recommendation"], 9
            ),
            -row["decisive_rows"],
            row["pattern"],
        )
    )
    return out


def _rule_candidate_buckets(pattern_recommendations: list[dict]) -> dict[str, list[dict]]:
    buckets = {"candidate_for_rule": [], "needs_more_labels": [], "reject_pattern": []}
    for item in pattern_recommendations:
        recommendation = item.get("recommendation", "")
        if recommendation in buckets:
            enriched = dict(item)
            enriched["required_action"] = _required_action(item)
            buckets[recommendation].append(enriched)
    return buckets


def _required_action(item: dict) -> str:
    recommendation = item.get("recommendation", "")
    if recommendation == "candidate_for_rule":
        return "Add regression tests for every supporting row, then apply this exact policy pattern only under existing risky-URL guards."
    if recommendation == "needs_more_labels":
        return "Keep this pattern in policy-validation tasks until it reaches five decisive supporting labels with zero blockers."
    if recommendation == "reject_pattern":
        return "Do not automate this policy pattern; add blocking regression fixtures for the blocking rows."
    return "No action."


def _recommendations(details: list[dict[str, str]], pattern_recommendations: list[dict]) -> list[str]:
    out: list[str] = []
    quality_issues = [row for row in details if row["decision_quality_issue"]]
    if quality_issues:
        out.append("Fix invalid decisions or replace rows missing manual_url before using this validation file for rule decisions.")
    blockers = [row for row in details if row["policy_blocks_rule"] == "1"]
    if blockers:
        out.append("At least one filled row blocks a candidate policy pattern; do not apply blocked patterns.")
    ready = [row for row in pattern_recommendations if row.get("recommendation") == "candidate_for_rule"]
    if ready:
        out.append("One or more policy patterns reached the minimum five clean decisive labels; add regression tests before enabling them.")
    thin = [row for row in pattern_recommendations if row.get("recommendation") == "needs_more_labels"]
    if thin:
        out.append("Some policy patterns are directionally supported but still thin; keep collecting targeted labels instead of widening automation.")
    if not out:
        out.append("No decisive policy validation evidence yet; keep thresholds and routing unchanged.")
    return out


def _group_stats(rows: list[dict[str, str]], field: str) -> dict[str, dict]:
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        grouped[row.get(field, "") or "(blank)"].append(row)
    out = {}
    for key, items in sorted(grouped.items()):
        decisive = _decisive(items)
        out[key] = {
            "rows": len(items),
            "labeled_rows": len(_labeled(items)),
            "decisive_rows": len(decisive),
            "supporting_rows": sum(1 for row in items if row["policy_supports_rule"] == "1"),
            "blocking_rows": sum(1 for row in items if row["policy_blocks_rule"] == "1"),
            "outcome_counts": dict(Counter(row["policy_outcome"] for row in items)),
        }
    return out


def _labeled(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    return [row for row in rows if row["normalized_decision"] and not row["decision_quality_issue"]]


def _decisive(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    return [row for row in _labeled(rows) if row["normalized_decision"] != "unsure"]


def _decision(row: dict[str, str]) -> str:
    raw = _first(row, "manual_decision", "your_decision", "decision").casefold()
    aliases = {
        "accept": "accept",
        "accepted": "accept",
        "approve": "accept",
        "approved": "accept",
        "correct": "accept",
        "yes": "accept",
        "true": "accept",
        "正确": "accept",
        "对": "accept",
        "replace": "replace",
        "replacement": "replace",
        "修正": "replace",
        "替换": "replace",
        "reject": "reject",
        "rejected": "reject",
        "no": "reject",
        "false": "reject",
        "wrong": "reject",
        "incorrect": "reject",
        "错误": "reject",
        "错": "reject",
        "无官网": "reject",
        "unsure": "unsure",
        "unknown": "unsure",
        "uncertain": "unsure",
        "不确定": "unsure",
    }
    return aliases.get(raw, raw if raw in {"accept", "replace", "reject", "unsure"} else "")


def _decision_quality_issue(manual_decision: str, normalized_decision: str, normalized_manual_url: str) -> str:
    if manual_decision.strip() and not normalized_decision:
        return "invalid_manual_decision"
    if normalized_decision == "replace" and not normalized_manual_url:
        return "replace_missing_manual_url"
    return ""


def _read_table(path: Path) -> list[dict[str, str]]:
    if path.suffix.casefold() == ".xlsx":
        return _read_xlsx(path)
    return _read_rows(path)


def _read_xlsx(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    workbook = load_workbook(path, data_only=False, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows())
    if not rows:
        return []
    headers = [_cell_text(cell.value) for cell in rows[0]]
    out = []
    for cells in rows[1:]:
        row = {headers[idx]: _cell_text(cells[idx].value) for idx in range(len(headers)) if headers[idx]}
        if any(row.values()):
            out.append(row)
    return out


def _read_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _write_rows(path: Path, rows: list[dict[str, str]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _render_markdown(report: dict) -> str:
    summary = report["summary"]
    lines = [
        "# Policy Validation Evaluation",
        "",
        "## Summary",
        "",
        f"- Task rows: {summary['task_rows']}",
        f"- Labeled rows: {summary['labeled_rows']}",
        f"- Decisive rows: {summary['decisive_rows']}",
        f"- Decision quality issue rows: {summary['decision_quality_issue_rows']}",
        f"- Support rows: {summary['support_rows']}",
        f"- Blocking rows: {summary['blocking_rows']}",
        f"- Candidate-for-rule patterns: {summary['candidate_for_rule_rows']}",
        f"- Needs-more-labels patterns: {summary['needs_more_labels_rows']}",
        f"- Reject-pattern rows: {summary['reject_pattern_rows']}",
        "",
        "## Recommendations",
        "",
    ]
    for item in report["recommendations"]:
        lines.append(f"- {item}")
    lines.extend(["", "## Patterns", ""])
    if not report["pattern_recommendations"]:
        lines.append("- None")
    for row in report["pattern_recommendations"]:
        lines.append(
            "- {recommendation}: rows={rows}, decisive={decisive}, support={support}, block={block}, strength={strength}, support_lower80={support_lower}, block_upper80={block_upper} :: {pattern}".format(
                recommendation=row["recommendation"],
                rows=row["rows"],
                decisive=row["decisive_rows"],
                support=row["supporting_rows"],
                block=row["blocking_rows"],
                strength=row["evidence_strength"],
                support_lower=row["support_rate_wilson_lower_80"],
                block_upper=row["blocking_rate_wilson_upper_80"],
                pattern=row["pattern"],
            )
        )
        lines.append(f"  Action: {_required_action(row)}")
    lines.append("")
    return "\n".join(lines)


def _first(row: dict[str, str], *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip():
            return _cell_text(value)
    return ""


def _cell_text(value: object) -> str:
    text = str(value or "").strip()
    if text.startswith("="):
        text = text[1:]
    if text.upper().startswith("HYPERLINK("):
        match = re.search(r'HYPERLINK\("([^"]+)"', text, flags=re.IGNORECASE)
        return match.group(1).strip() if match else ""
    return text


def _normalize_url(value: object) -> str:
    raw = str(value or "").strip().replace("\xa0", "").rstrip(".,);]")
    if not raw:
        return ""
    parsed = urlparse(raw if "://" in raw else f"https://{raw}")
    if not parsed.netloc:
        return ""
    return f"{parsed.scheme or 'https'}://{parsed.netloc}{parsed.path or ''}".rstrip("/")


def _same_site(left: str, right: str) -> bool:
    if not left or not right:
        return False
    left_parsed = urlparse(left if "://" in left else f"https://{left}")
    right_parsed = urlparse(right if "://" in right else f"https://{right}")
    return _base_domain(left_parsed.netloc) == _base_domain(right_parsed.netloc)


def _base_domain(domain: str) -> str:
    labels = [item for item in domain.casefold().split(".") if item and item != "www"]
    return ".".join(labels[-2:]) if len(labels) >= 2 else ".".join(labels)


def _ratio(num: int, den: int) -> float | None:
    return round(num / den, 4) if den else None


def _wilson_interval(successes: int, total: int, z: float = 1.2815515655446004) -> dict[str, float | None]:
    if total <= 0:
        return {"lower": None, "upper": None}
    p = successes / total
    denom = 1 + (z * z / total)
    center = (p + (z * z / (2 * total))) / denom
    margin = (z * ((p * (1 - p) / total + z * z / (4 * total * total)) ** 0.5)) / denom
    return {
        "lower": round(max(0.0, center - margin), 4),
        "upper": round(min(1.0, center + margin), 4),
    }


def _evidence_strength(decisive_rows: int, support_rows: int, blocking_rows: int) -> str:
    if decisive_rows <= 0:
        return "unlabeled"
    if blocking_rows:
        return "blocked"
    if support_rows < 5:
        return "thin_support"
    if decisive_rows < 10:
        return "minimum_support"
    return "strong_support"


if __name__ == "__main__":
    raise SystemExit(main())
