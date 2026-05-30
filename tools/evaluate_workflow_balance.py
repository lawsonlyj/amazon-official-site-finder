from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
from pathlib import Path
from urllib.parse import urlparse

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from finder.text import domain_from_url


DETAIL_FIELDS = [
    "provider_id",
    "provider_name",
    "label_source",
    "expected_kind",
    "expected_domain",
    "expected_url",
    "output_status",
    "output_confidence",
    "output_domain",
    "output_url",
    "outcome",
    "manual_review_required",
    "manual_review_reason",
    "agent_b_checked",
    "agent_b_decision",
    "agent_b_candidate_domain",
    "agent_b_suggested_domain",
    "agent_b_confidence",
    "agent_b_candidate_score",
    "agent_b_reason_for_unsure",
]


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Evaluate precision/coverage balance from human review labels.")
    parser.add_argument("--baseline-final", help="Baseline final CSV. Non-reviewed rows are treated as correct labels.")
    parser.add_argument("--candidate-final", required=True, help="Candidate workflow final CSV to evaluate.")
    parser.add_argument("--human-review", help="Filled human review CSV/XLSX with corrected yellow rows.")
    parser.add_argument(
        "--labeled-details",
        help="Existing balance details CSV/JSON with expected_kind/expected_domain labels; useful after old baseline artifacts are cleaned.",
    )
    parser.add_argument("--run-dir", help="Optional candidate run dir, used to count review_task rows.")
    parser.add_argument(
        "--simulate-thresholds",
        default="",
        help="Comma-separated matched confidence thresholds to simulate by moving lower matched rows to unresolved.",
    )
    parser.add_argument("--output-json")
    parser.add_argument("--output-csv")
    args = parser.parse_args(argv)

    if args.labeled_details:
        summary = evaluate_balance_from_details(
            labeled_details=args.labeled_details,
            candidate_final=args.candidate_final,
            run_dir=args.run_dir,
            output_json=args.output_json,
            output_csv=args.output_csv,
            simulate_thresholds=args.simulate_thresholds,
        )
    else:
        if not args.baseline_final or not args.human_review:
            parser.error("--baseline-final and --human-review are required unless --labeled-details is provided.")
        summary = evaluate_balance(
            baseline_final=args.baseline_final,
            candidate_final=args.candidate_final,
            human_review=args.human_review,
            run_dir=args.run_dir,
            output_json=args.output_json,
            output_csv=args.output_csv,
            simulate_thresholds=args.simulate_thresholds,
        )
    print(json.dumps(summary["overall"], ensure_ascii=False, indent=2))
    return 0


def evaluate_balance(
    *,
    baseline_final: str | Path,
    candidate_final: str | Path,
    human_review: str | Path,
    run_dir: str | Path | None = None,
    output_json: str | Path | None = None,
    output_csv: str | Path | None = None,
    simulate_thresholds: str | list[int] | None = None,
) -> dict:
    baseline_rows = _read_rows(Path(baseline_final))
    candidate_rows = _index_rows(_read_rows(Path(candidate_final)))
    review_rows = _index_rows(_read_table(Path(human_review)))
    labels = [_label_from_row(row, review_rows.get(_row_key(row), {})) for row in baseline_rows]
    labels = [label for label in labels if label]
    return _evaluate_balance_from_labels(
        labels=labels,
        candidate_rows=candidate_rows,
        candidate_final=candidate_final,
        run_dir=run_dir,
        output_json=output_json,
        output_csv=output_csv,
        simulate_thresholds=simulate_thresholds,
        inputs={
            "baseline_final": str(baseline_final),
            "candidate_final": str(candidate_final),
            "human_review": str(human_review),
        },
    )


def evaluate_balance_from_details(
    *,
    labeled_details: str | Path,
    candidate_final: str | Path,
    run_dir: str | Path | None = None,
    output_json: str | Path | None = None,
    output_csv: str | Path | None = None,
    simulate_thresholds: str | list[int] | None = None,
) -> dict:
    detail_rows = _read_details(Path(labeled_details))
    labels = [_label_from_detail_row(row) for row in detail_rows]
    labels = [label for label in labels if label]
    candidate_rows = _index_rows(_read_rows(Path(candidate_final)))
    return _evaluate_balance_from_labels(
        labels=labels,
        candidate_rows=candidate_rows,
        candidate_final=candidate_final,
        run_dir=run_dir,
        output_json=output_json,
        output_csv=output_csv,
        simulate_thresholds=simulate_thresholds,
        inputs={
            "labeled_details": str(labeled_details),
            "candidate_final": str(candidate_final),
        },
    )


def _evaluate_balance_from_labels(
    *,
    labels: list[dict[str, str]],
    candidate_rows: dict[str, dict[str, str]],
    candidate_final: str | Path,
    run_dir: str | Path | None,
    output_json: str | Path | None,
    output_csv: str | Path | None,
    simulate_thresholds: str | list[int] | None,
    inputs: dict[str, str],
) -> dict:
    details = [_evaluate_label(label, _candidate_for_label(label, candidate_rows)) for label in labels]
    review_task_rows = None
    review_task_path = None
    agent_b_rows = None
    agent_b_path = None
    if run_dir:
        run_dir = Path(run_dir)
        review_task_path = _find_review_task(run_dir)
        if review_task_path:
            review_task_rows = _read_rows(review_task_path)
            details = _annotate_manual_review(details, review_task_rows)
        agent_b_path = _find_agent_b(run_dir)
        if agent_b_path:
            agent_b_rows = _read_rows(agent_b_path)
            details = _annotate_agent_b(details, agent_b_rows)
    overall = _summarize(details)
    agent_b_recall_release_simulations = []
    manual_review_lanes = []
    manual_review_lane_drop_simulations = []
    if run_dir:
        if review_task_rows is not None:
            overall.update(_summarize_manual_review_capture(details, len(review_task_rows)))
            manual_review_lanes = _summarize_manual_review_lanes(details, review_task_rows)
            manual_review_lane_drop_simulations = _manual_review_lane_drop_simulations(
                manual_review_lanes,
                len(review_task_rows),
            )
        if agent_b_rows is not None:
            overall.update(_summarize_agent_b_balance(details, agent_b_rows))
            agent_b_recall_release_simulations = _agent_b_recall_release_simulations(details)
        unresolved = _find_unresolved(run_dir)
        if unresolved.exists():
            overall["unresolved_rows"] = len(_read_rows(unresolved))
    summary = {
        "overall": overall,
        "threshold_simulations": _threshold_simulations(labels, candidate_rows, simulate_thresholds),
        "agent_b_recall_release_simulations": agent_b_recall_release_simulations,
        "manual_review_lanes": manual_review_lanes,
        "manual_review_lane_drop_simulations": manual_review_lane_drop_simulations,
        "inputs": {
            "candidate_final": str(candidate_final),
            **inputs,
            "run_dir": str(run_dir) if run_dir else "",
            "review_task": str(review_task_path) if review_task_path else "",
            "agent_b": str(agent_b_path) if agent_b_path else "",
        },
        "details": details,
    }
    if output_json:
        path = Path(output_json)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    if output_csv:
        _write_rows(Path(output_csv), details, DETAIL_FIELDS)
    return summary


def _threshold_simulations(
    labels: list[dict[str, str]],
    candidate_rows: dict[str, dict[str, str]],
    thresholds: str | list[int] | None,
) -> list[dict]:
    if not thresholds:
        return []
    if isinstance(thresholds, str):
        values = [item.strip() for item in thresholds.split(",") if item.strip()]
        threshold_values = []
        for value in values:
            try:
                threshold_values.append(int(float(value)))
            except ValueError:
                continue
    else:
        threshold_values = [int(value) for value in thresholds]
    out = []
    for threshold in threshold_values:
        rows = []
        for label in labels:
            candidate = dict(_candidate_for_label(label, candidate_rows))
            if _should_drop_for_threshold(candidate, threshold):
                candidate["official_url"] = ""
                candidate["official_domain"] = ""
                candidate["status"] = "unresolved"
                candidate["decision_source"] = f"threshold_sim:{threshold}"
            rows.append(_evaluate_label(label, candidate))
        overall = _summarize(rows)
        overall["threshold"] = threshold
        out.append(overall)
    return out


def _should_drop_for_threshold(row: dict[str, str], threshold: int) -> bool:
    if row.get("status") != "matched":
        return False
    if not row.get("official_url"):
        return False
    try:
        confidence = int(float(row.get("confidence") or 0))
    except (TypeError, ValueError):
        return True
    return confidence < threshold


def _label_from_row(row: dict[str, str], review_row: dict[str, str]) -> dict[str, str] | None:
    decision = _decision(review_row)
    provider_id = row.get("provider_id", "")
    provider_name = row.get("provider_name", "")
    if decision == "unsure":
        return None
    if decision == "replace":
        manual_url = _normalize_url(_first(review_row, "manual_url", "your_true_official_url", "true_official_url"))
        if manual_url:
            return _label(provider_id, provider_name, "human_replace", "official", manual_url)
        return _label(provider_id, provider_name, "human_replace_missing_url", "no_official", "")
    if decision == "reject":
        manual_url = _normalize_url(_first(review_row, "manual_url", "your_true_official_url", "true_official_url"))
        if manual_url:
            return _label(provider_id, provider_name, "human_reject_with_url", "official", manual_url)
        return _label(provider_id, provider_name, "human_reject", "no_official", "")
    if decision == "accept":
        manual_url = _normalize_url(_first(review_row, "manual_url", "your_true_official_url", "true_official_url"))
        accepted_url = manual_url or _normalize_url(_first(review_row, "official_url", "current_or_candidate_url", "candidate_url"))
        if accepted_url:
            return _label(provider_id, provider_name, "human_accept", "official", accepted_url)
        return _label(provider_id, provider_name, "human_accept_no_url", "no_official", "")
    baseline_url = _normalize_url(row.get("official_url", ""))
    if baseline_url:
        return _label(provider_id, provider_name, "baseline_unmarked_correct", "official", baseline_url)
    return _label(provider_id, provider_name, "baseline_unmarked_no_official", "no_official", "")


def _label(provider_id: str, provider_name: str, source: str, kind: str, url: str) -> dict[str, str]:
    return {
        "provider_id": provider_id,
        "provider_name": provider_name,
        "label_source": source,
        "expected_kind": kind,
        "expected_url": url,
        "expected_domain": domain_from_url(url) if url else "",
    }


def _label_from_detail_row(row: dict[str, str]) -> dict[str, str] | None:
    provider_id = row.get("provider_id", "")
    provider_name = row.get("provider_name", "")
    source = row.get("label_source", "") or "labeled_details"
    expected_url = _normalize_url(row.get("expected_url", ""))
    expected_domain = domain_from_url(row.get("expected_domain", "") or expected_url)
    expected_kind = (row.get("expected_kind") or "").strip().casefold()
    if expected_kind not in {"official", "no_official"}:
        expected_kind = "official" if expected_domain else "no_official"
    if expected_kind == "official" and not expected_domain:
        return None
    if expected_kind == "no_official":
        expected_url = ""
        expected_domain = ""
    return {
        "provider_id": provider_id,
        "provider_name": provider_name,
        "label_source": source,
        "expected_kind": expected_kind,
        "expected_url": expected_url,
        "expected_domain": expected_domain,
    }


def _evaluate_label(label: dict[str, str], row: dict[str, str]) -> dict[str, str]:
    output_url = _normalize_url(row.get("official_url", ""))
    output_domain = domain_from_url(row.get("official_domain", "") or output_url) if output_url else ""
    expected_kind = label["expected_kind"]
    expected_domain = label["expected_domain"]
    if expected_kind == "official":
        if output_domain and output_domain == expected_domain:
            outcome = "correct_official"
        elif output_domain:
            outcome = "false_official"
        else:
            outcome = "over_rejected"
    else:
        outcome = "false_official" if output_domain else "correct_no_official"
    return {
        **label,
        "output_status": row.get("status", ""),
        "output_confidence": row.get("confidence", ""),
        "output_domain": output_domain,
        "output_url": output_url,
        "outcome": outcome,
    }


def _candidate_for_label(label: dict[str, str], rows: dict[str, dict[str, str]]) -> dict[str, str]:
    provider_id = label.get("provider_id", "").strip()
    if provider_id and f"id:{provider_id}" in rows:
        return rows[f"id:{provider_id}"]
    provider_name = label.get("provider_name", "").strip().casefold()
    return rows.get(f"name:{provider_name}", {})


def _summarize(details: list[dict[str, str]]) -> dict:
    total = len(details)
    expected_official = sum(1 for row in details if row["expected_kind"] == "official")
    expected_no_official = total - expected_official
    correct_official = sum(1 for row in details if row["outcome"] == "correct_official")
    correct_no_official = sum(1 for row in details if row["outcome"] == "correct_no_official")
    false_official = sum(1 for row in details if row["outcome"] == "false_official")
    over_rejected = sum(1 for row in details if row["outcome"] == "over_rejected")
    official_outputs = correct_official + false_official
    return {
        "labeled_rows": total,
        "expected_official_rows": expected_official,
        "expected_no_official_rows": expected_no_official,
        "official_output_rows": official_outputs,
        "correct_official_rows": correct_official,
        "correct_no_official_rows": correct_no_official,
        "false_official_rows": false_official,
        "over_rejected_rows": over_rejected,
        "auto_precision": _ratio(correct_official, official_outputs),
        "official_recall": _ratio(correct_official, expected_official),
        "overall_accuracy": _ratio(correct_official + correct_no_official, total),
    }


def _annotate_manual_review(details: list[dict[str, str]], review_task_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    review_index = _index_rows(review_task_rows)
    out = []
    for row in details:
        review_row = review_index.get(_row_key(row), {})
        annotated = dict(row)
        annotated["manual_review_required"] = "yes" if review_row else ""
        annotated["manual_review_reason"] = _first(review_row, "review_reason") if review_row else ""
        out.append(annotated)
    return out


def _annotate_agent_b(details: list[dict[str, str]], agent_b_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    agent_b_index = _index_rows(agent_b_rows)
    out = []
    for row in details:
        agent_b_row = agent_b_index.get(_row_key(row), {})
        annotated = dict(row)
        annotated["agent_b_checked"] = "yes" if agent_b_row else ""
        annotated["agent_b_decision"] = _first(agent_b_row, "agent_b_decision", "manual_decision") if agent_b_row else ""
        annotated["agent_b_candidate_domain"] = _agent_b_candidate_domain(agent_b_row)
        annotated["agent_b_suggested_domain"] = _agent_b_suggested_domain(agent_b_row)
        annotated["agent_b_confidence"] = _first(agent_b_row, "confidence", "evidence_score") if agent_b_row else ""
        annotated["agent_b_candidate_score"] = _first(agent_b_row, "evidence_score", "confidence") if agent_b_row else ""
        annotated["agent_b_reason_for_unsure"] = _first(agent_b_row, "reason_for_unsure") if agent_b_row else ""
        out.append(annotated)
    return out


def _summarize_manual_review_capture(details: list[dict[str, str]], review_task_rows: int) -> dict:
    reviewed = [row for row in details if row.get("manual_review_required") == "yes"]
    false_official_total = sum(1 for row in details if row["outcome"] == "false_official")
    over_rejected_total = sum(1 for row in details if row["outcome"] == "over_rejected")
    false_official_reviewed = sum(1 for row in reviewed if row["outcome"] == "false_official")
    over_rejected_reviewed = sum(1 for row in reviewed if row["outcome"] == "over_rejected")
    correct_official_reviewed = sum(1 for row in reviewed if row["outcome"] == "correct_official")
    correct_no_official_reviewed = sum(1 for row in reviewed if row["outcome"] == "correct_no_official")
    return {
        "manual_review_rows": review_task_rows,
        "manual_review_labeled_rows": len(reviewed),
        "manual_review_false_official_rows": false_official_reviewed,
        "manual_review_missed_false_official_rows": false_official_total - false_official_reviewed,
        "manual_review_over_rejected_rows": over_rejected_reviewed,
        "manual_review_missed_over_rejected_rows": over_rejected_total - over_rejected_reviewed,
        "manual_review_correct_official_rows": correct_official_reviewed,
        "manual_review_correct_no_official_rows": correct_no_official_reviewed,
        "manual_review_false_official_capture_rate": _ratio(false_official_reviewed, false_official_total),
        "manual_review_over_rejected_capture_rate": _ratio(over_rejected_reviewed, over_rejected_total),
        "manual_review_false_official_share": _ratio(false_official_reviewed, len(reviewed)),
        "manual_review_correct_official_share": _ratio(correct_official_reviewed, len(reviewed)),
    }


def _summarize_manual_review_lanes(
    details: list[dict[str, str]],
    review_task_rows: list[dict[str, str]],
) -> list[dict]:
    review_counts = Counter(row.get("review_reason", "") for row in review_task_rows)
    reviewed_details = [row for row in details if row.get("manual_review_required") == "yes"]
    all_reasons = sorted(
        set(review_counts)
        | {row.get("manual_review_reason", "") for row in reviewed_details if row.get("manual_review_reason")}
    )
    out = []
    for reason in all_reasons:
        lane_details = [row for row in reviewed_details if row.get("manual_review_reason") == reason]
        outcome_counts = _counts(row["outcome"] for row in lane_details)
        labeled_rows = len(lane_details)
        risk_rows = outcome_counts.get("false_official", 0) + outcome_counts.get("over_rejected", 0)
        out.append(
            {
                "review_reason": reason,
                "review_task_rows": review_counts.get(reason, 0),
                "labeled_rows": labeled_rows,
                "false_official_rows": outcome_counts.get("false_official", 0),
                "over_rejected_rows": outcome_counts.get("over_rejected", 0),
                "correct_official_rows": outcome_counts.get("correct_official", 0),
                "correct_no_official_rows": outcome_counts.get("correct_no_official", 0),
                "risk_rows": risk_rows,
                "risk_share_of_labeled_lane": _ratio(risk_rows, labeled_rows),
                "correct_share_of_labeled_lane": _ratio(
                    outcome_counts.get("correct_official", 0) + outcome_counts.get("correct_no_official", 0),
                    labeled_rows,
                ),
            }
        )
    return sorted(
        out,
        key=lambda row: (
            -(row["risk_rows"] or 0),
            -(row["review_task_rows"] or 0),
            row["review_reason"],
        ),
    )


def _manual_review_lane_drop_simulations(lanes: list[dict], review_task_rows: int) -> list[dict]:
    out = []
    for lane in lanes:
        out.append(
            {
                "drop_review_reason": lane["review_reason"],
                "manual_review_rows_after_drop": review_task_rows - lane.get("review_task_rows", 0),
                "manual_review_rows_removed": lane.get("review_task_rows", 0),
                "known_false_official_missed_if_dropped": lane.get("false_official_rows", 0),
                "known_over_rejected_missed_if_dropped": lane.get("over_rejected_rows", 0),
                "known_correct_reviews_removed_if_dropped": lane.get("correct_official_rows", 0)
                + lane.get("correct_no_official_rows", 0),
            }
        )
    return sorted(
        out,
        key=lambda row: (
            row["known_false_official_missed_if_dropped"] > 0,
            row["known_over_rejected_missed_if_dropped"] > 0,
            -row["manual_review_rows_removed"],
            row["drop_review_reason"],
        ),
    )


def _summarize_agent_b_balance(details: list[dict[str, str]], agent_b_rows: list[dict[str, str]]) -> dict:
    checked = [row for row in details if row.get("agent_b_checked") == "yes"]
    decisions = _counts(row.get("agent_b_decision", "") for row in checked)
    false_official = [row for row in checked if row["outcome"] == "false_official"]
    false_official_accepted = [row for row in false_official if row.get("agent_b_decision") == "accept"]
    false_official_caught = [row for row in false_official if row.get("agent_b_decision") and row.get("agent_b_decision") != "accept"]
    correct_official = [row for row in checked if row["outcome"] == "correct_official"]
    correct_official_accepted = [row for row in correct_official if row.get("agent_b_decision") == "accept"]
    correct_no_official_expected = [row for row in checked if row["expected_kind"] == "no_official"]
    no_official_accepted_or_replaced = [
        row for row in correct_no_official_expected if row.get("agent_b_decision") in {"accept", "replace"}
    ]
    over_rejected = [row for row in checked if row["outcome"] == "over_rejected"]
    over_rejected_correct = [
        row for row in over_rejected if row.get("agent_b_suggested_domain") and row["agent_b_suggested_domain"] == row["expected_domain"]
    ]
    over_rejected_wrong = [
        row for row in over_rejected if row.get("agent_b_suggested_domain") and row["agent_b_suggested_domain"] != row["expected_domain"]
    ]
    over_rejected_hold = [row for row in over_rejected if not row.get("agent_b_suggested_domain")]
    return {
        "agent_b_rows": len(agent_b_rows),
        "agent_b_labeled_rows": len(checked),
        "agent_b_accept_rows": decisions.get("accept", 0),
        "agent_b_replace_rows": decisions.get("replace", 0),
        "agent_b_reject_rows": decisions.get("reject", 0),
        "agent_b_unsure_rows": decisions.get("unsure", 0),
        "agent_b_false_official_rows": len(false_official),
        "agent_b_false_official_caught_rows": len(false_official_caught),
        "agent_b_false_official_accept_rows": len(false_official_accepted),
        "agent_b_false_official_catch_rate": _ratio(len(false_official_caught), len(false_official)),
        "agent_b_false_official_accept_rate": _ratio(len(false_official_accepted), len(false_official)),
        "agent_b_correct_official_rows": len(correct_official),
        "agent_b_correct_official_accept_rows": len(correct_official_accepted),
        "agent_b_correct_official_non_accept_rows": len(correct_official) - len(correct_official_accepted),
        "agent_b_correct_official_accept_rate": _ratio(len(correct_official_accepted), len(correct_official)),
        "agent_b_expected_no_official_rows": len(correct_no_official_expected),
        "agent_b_expected_no_official_accept_or_replace_rows": len(no_official_accepted_or_replaced),
        "agent_b_expected_no_official_accept_or_replace_rate": _ratio(
            len(no_official_accepted_or_replaced), len(correct_no_official_expected)
        ),
        "agent_b_over_rejected_rows": len(over_rejected),
        "agent_b_over_rejected_correct_recovery_rows": len(over_rejected_correct),
        "agent_b_over_rejected_wrong_recovery_rows": len(over_rejected_wrong),
        "agent_b_over_rejected_hold_rows": len(over_rejected_hold),
        "agent_b_over_rejected_recovery_rate": _ratio(len(over_rejected_correct), len(over_rejected)),
    }


def _agent_b_suggested_domain(row: dict[str, str]) -> str:
    decision = _first(row, "agent_b_decision", "manual_decision")
    if decision == "accept":
        return domain_from_url(_first(row, "candidate_domain", "candidate_url"))
    if decision == "replace":
        return domain_from_url(_first(row, "manual_url", "replacement_url", "replacement_domain"))
    return ""


def _agent_b_candidate_domain(row: dict[str, str]) -> str:
    return domain_from_url(_first(row, "candidate_domain", "candidate_url", "replacement_domain", "replacement_url"))


def _agent_b_recall_release_simulations(details: list[dict[str, str]]) -> list[dict]:
    thresholds = [0, 30, 45, 50, 60, 70, 75, 80, 85]
    recall_rows = [
        row
        for row in details
        if row.get("manual_review_reason") == "recall_unresolved_top_candidate"
        and row.get("agent_b_checked") == "yes"
        and row.get("agent_b_candidate_domain")
    ]
    if not recall_rows:
        return []
    expected_official = [row for row in recall_rows if row.get("expected_kind") == "official"]
    out = []
    for threshold in thresholds:
        released = [row for row in recall_rows if _numeric(row.get("agent_b_candidate_score")) >= threshold]
        correct = [
            row
            for row in released
            if row.get("expected_kind") == "official"
            and row.get("agent_b_candidate_domain") == row.get("expected_domain")
        ]
        wrong = [
            row
            for row in released
            if row.get("expected_kind") != "official"
            or row.get("agent_b_candidate_domain") != row.get("expected_domain")
        ]
        held_correct = [
            row
            for row in recall_rows
            if row not in released
            and row.get("expected_kind") == "official"
            and row.get("agent_b_candidate_domain") == row.get("expected_domain")
        ]
        out.append(
            {
                "agent_b_evidence_threshold": threshold,
                "recall_rows": len(recall_rows),
                "expected_official_rows": len(expected_official),
                "release_rows": len(released),
                "correct_recovery_rows": len(correct),
                "wrong_release_rows": len(wrong),
                "held_correct_recovery_rows": len(held_correct),
                "release_precision": _ratio(len(correct), len(released)),
                "official_recovery_rate": _ratio(len(correct), len(expected_official)),
                "released_correct_provider_ids": [row.get("provider_id", "") for row in correct],
                "released_wrong_provider_ids": [row.get("provider_id", "") for row in wrong],
            }
        )
    return out


def _counts(values) -> dict[str, int]:
    counts: dict[str, int] = {}
    for value in values:
        key = str(value or "").strip()
        if key:
            counts[key] = counts.get(key, 0) + 1
    return counts


def _ratio(num: int, den: int) -> float | None:
    return round(num / den, 4) if den else None


def _numeric(value: object) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _find_review_task(run_dir: Path) -> Path | None:
    for name in ("review_task.csv", "manual_official_site_review_task.csv"):
        path = run_dir / name
        if path.exists():
            return path
    return None


def _find_unresolved(run_dir: Path) -> Path:
    for name in ("unresolved.csv", "provider_unresolved_second_pass.csv"):
        path = run_dir / name
        if path.exists():
            return path
    return run_dir / "unresolved.csv"


def _find_agent_b(run_dir: Path) -> Path | None:
    for name in ("agent_b/check.csv", "agent_b_verification_results.csv"):
        path = run_dir / name
        if path.exists():
            return path
    return None


def _read_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _read_table(path: Path) -> list[dict[str, str]]:
    if path.suffix.casefold() == ".xlsx":
        return _read_xlsx(path)
    return _read_rows(path)


def _read_details(path: Path) -> list[dict[str, str]]:
    if path.suffix.casefold() == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            rows = data.get("details", [])
            return rows if isinstance(rows, list) else []
        return data if isinstance(data, list) else []
    return _read_rows(path)


def _read_xlsx(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    workbook = load_workbook(path, data_only=False, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows())
    if not rows:
        return []
    headers = [_cell_text(cell.value) for cell in rows[0]]
    out = []
    for cells in rows[1:]:
        row = {headers[idx]: _cell_text(cells[idx].value) for idx in range(len(headers)) if headers[idx]}
        if any(row.values()):
            out.append(row)
    return out


def _index_rows(rows: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    return {_row_key(row): row for row in rows if _row_key(row)}


def _row_key(row: dict[str, str]) -> str:
    provider_id = str(row.get("provider_id") or "").strip()
    if provider_id:
        return f"id:{provider_id}"
    return f"name:{str(row.get('provider_name') or '').strip().casefold()}"


def _decision(row: dict[str, str]) -> str:
    raw = _first(row, "manual_decision", "your_decision", "decision").casefold()
    aliases = {
        "accept": "accept",
        "approve": "accept",
        "approved": "accept",
        "replace": "replace",
        "reject": "reject",
        "rejected": "reject",
        "unsure": "unsure",
    }
    return aliases.get(raw, raw)


def _first(row: dict[str, str], *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip():
            return _cell_text(value)
    return ""


def _cell_text(value: object) -> str:
    text = str(value or "").strip()
    if text.startswith("="):
        text = text[1:]
    if text.upper().startswith("HYPERLINK("):
        match = re.search(r'HYPERLINK\("([^"]+)"', text, flags=re.IGNORECASE)
        return match.group(1).strip() if match else ""
    return text


def _normalize_url(value: object) -> str:
    raw = str(value or "").strip().replace("\xa0", "").rstrip(".,);]")
    if not raw:
        return ""
    parsed = urlparse(raw if "://" in raw else f"https://{raw}")
    if not parsed.netloc:
        return ""
    path = parsed.path or ""
    return f"{parsed.scheme or 'https'}://{parsed.netloc}{path}".rstrip("/")


def _write_rows(path: Path, rows: list[dict[str, str]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


if __name__ == "__main__":
    raise SystemExit(main())
