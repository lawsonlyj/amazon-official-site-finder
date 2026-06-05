from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import zipfile
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from finder.text import domain_from_url
from tools.output_layout import (
    agent_b_suggestion_paths,
    first_existing,
)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Generate check-and-suggestion optimization suggestions from check results and review learning.")
    parser.add_argument("--run-dir", required=True)
    parser.add_argument("--check-csv", dest="agent_b_csv")
    parser.add_argument("--agent-b-csv", dest="agent_b_csv")
    parser.add_argument("--learning-summary")
    parser.add_argument("--human-review", help="Optional filled human review CSV/XLSX to convert into recommendations and fixtures.")
    parser.add_argument("--output-json")
    parser.add_argument("--output-md")
    args = parser.parse_args(argv)

    summary = run_agent_b_recommendations(
        run_dir=args.run_dir,
        agent_b_csv=args.agent_b_csv,
        learning_summary=args.learning_summary,
        output_json=args.output_json,
        output_md=args.output_md,
        human_review=args.human_review,
    )
    print(json.dumps(summary["overall"], ensure_ascii=False, indent=2))
    return 0


def run_agent_b_recommendations(
    *,
    run_dir: str | Path,
    agent_b_csv: str | Path | None = None,
    learning_summary: str | Path | None = None,
    human_review: str | Path | None = None,
    output_json: str | Path | None = None,
    output_md: str | Path | None = None,
) -> dict:
    run_dir = Path(run_dir)
    agent_b_path = Path(agent_b_csv) if agent_b_csv else (
        first_existing(run_dir, "check_suggestion/check.csv", "agent_b/check.csv", "agent_b_verification_results.csv")
        or run_dir / "check_suggestion/check.csv"
    )
    learning_path = Path(learning_summary) if learning_summary else (
        first_existing(run_dir, "reviewed/learning.json", "manual_review_learning_summary.json")
        or run_dir / "reviewed/learning.json"
    )
    canonical = agent_b_suggestion_paths(run_dir)
    output_json_path = Path(output_json) if output_json else canonical["json"]
    output_md_path = Path(output_md) if output_md else canonical["md"]

    agent_b_rows = _read_rows(agent_b_path)
    learning = _read_json(learning_path)
    human_review_rows = _read_table(Path(human_review)) if human_review else []
    recommendations = analyze_recommendations(agent_b_rows, learning, human_review_rows=human_review_rows)
    summary = {
        "overall": {
            "agent_b_rows": len(agent_b_rows),
            "human_review_rows": len(human_review_rows),
            "recommendation_count": len(recommendations),
            "safe_config_action_count": sum(1 for item in recommendations if item.get("safe_to_apply")),
        },
        "recommendations": recommendations,
        "inputs": {
            "agent_b_csv": str(agent_b_path),
            "learning_summary": str(learning_path),
            "human_review": str(human_review) if human_review else "",
        },
        "outputs": {"json": str(output_json_path), "md": str(output_md_path)},
    }
    output_json_path.parent.mkdir(parents=True, exist_ok=True)
    output_md_path.parent.mkdir(parents=True, exist_ok=True)
    output_json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    _write_markdown(output_md_path, summary)
    _update_manifest(run_dir / "manifest.json", summary)
    return summary


def analyze_recommendations(
    agent_b_rows: list[dict[str, str]],
    learning: dict,
    *,
    human_review_rows: list[dict[str, str]] | None = None,
) -> list[dict]:
    recommendations = []
    rejected_domains = Counter()
    replace_queries = Counter()
    low_score_accepts = []
    same_name_conflicts = []

    for row in agent_b_rows:
        decision = row.get("agent_b_decision") or row.get("manual_decision")
        candidate_domain = domain_from_url(row.get("candidate_domain") or row.get("candidate_url", ""))
        if decision == "reject" and candidate_domain:
            rejected_domains[candidate_domain] += 1
        if decision == "replace":
            for query in [item.strip() for item in row.get("independent_search_queries", "").split(";") if item.strip()]:
                replace_queries[query] += 1
        if decision == "accept" and _to_int(row.get("evidence_score")) < 70:
            low_score_accepts.append(row)
        counter = row.get("counter_evidence", "").casefold()
        unsure = row.get("reason_for_unsure", "").casefold()
        if "name" in counter or "conflict" in unsure or decision == "unsure":
            same_name_conflicts.append(row)

    for domain, count in rejected_domains.items():
        if count >= 2 or _looks_like_bad_domain(domain):
            recommendations.append(
                {
                    "type": "excluded_domain",
                    "title": "Repeated rejected bad domain",
                    "domain": domain,
                    "count": count,
                    "safe_to_apply": True,
                    "action": "add_to_excluded_domains",
                    "reason": "Multiple check-and-suggestion rejects or known directory/platform domain.",
                }
            )

    repeated_replace_queries = [query for query, count in replace_queries.items() if count >= 2]
    if repeated_replace_queries:
        recommendations.append(
            {
                "type": "second_pass_query",
                "title": "Repeated replacement query pattern",
                "queries": repeated_replace_queries[:10],
                "count": sum(replace_queries[query] for query in repeated_replace_queries),
                "safe_to_apply": False,
                "action": "review_second_pass_query_templates",
                "reason": "Several replacements were found by the same independent query shape.",
            }
        )

    if len(low_score_accepts) >= 2:
        recommendations.append(
            {
                "type": "threshold_review",
                "title": "Low-score accepts confirmed by check-and-suggestion",
                "count": len(low_score_accepts),
                "safe_to_apply": False,
                "action": "evaluate_strong_evidence_rule_or_threshold",
                "reason": "Confirmed low-score accepts should become labels before any threshold change.",
            }
        )

    if len(same_name_conflicts) >= 3:
        recommendations.append(
            {
                "type": "identity_constraint",
                "title": "Same-name or insufficient identity conflicts",
                "count": len(same_name_conflicts),
                "safe_to_apply": False,
                "safe_artifact": True,
                "action": "write_identity_regression_fixtures",
                "reason": "Repeated uncertainty suggests identity constraints, not automatic URL rules.",
                "examples": [
                    {
                        "provider_id": row.get("provider_id", ""),
                        "provider_name": row.get("provider_name", ""),
                        "candidate_url": row.get("candidate_url", ""),
                        "candidate_domain": row.get("candidate_domain", ""),
                        "agent_b_decision": row.get("agent_b_decision", ""),
                        "evidence_score": row.get("evidence_score", ""),
                        "counter_evidence": row.get("counter_evidence", ""),
                        "reason_for_unsure": row.get("reason_for_unsure", ""),
                    }
                    for row in same_name_conflicts[:50]
                ],
            }
            )

    human_recommendations = _human_review_recommendations(human_review_rows or [])
    recommendations.extend(human_recommendations)

    learning_opt = (learning or {}).get("optimization", {})
    for domain in learning_opt.get("safe_excluded_domain_candidates", []) or []:
        if not any(item.get("domain") == domain for item in recommendations):
            recommendations.append(
                {
                    "type": "excluded_domain",
                    "title": "Manual review safe excluded-domain candidate",
                    "domain": domain,
                    "count": 1,
                    "safe_to_apply": True,
                    "action": "add_to_excluded_domains",
                    "reason": "Manual review learning marked this as a safe repeated/config candidate.",
                }
            )

    if not recommendations and (agent_b_rows or learning):
        recommendations.append(
            {
                "type": "labels_only",
                "title": "No repeated safe rule change",
                "safe_to_apply": False,
                "action": "keep_as_labels_and_evidence",
                "reason": "Patterns are single-case or not safe enough for automatic config changes.",
            }
        )
    return recommendations


def _human_review_recommendations(rows: list[dict[str, str]]) -> list[dict]:
    if not rows:
        return []
    normalized = [_normalize_human_review_row(row) for row in rows]
    normalized = [row for row in normalized if row["provider_id"] or row["provider_name"]]
    if not normalized:
        return []
    decision_counts = Counter(row["manual_decision"] for row in normalized)
    tag_counts = Counter(tag for row in normalized for tag in row["note_tags"])
    recommendations: list[dict] = [
        {
            "type": "human_review_regression",
            "title": "Human review labels should become regression fixtures",
            "count": len(normalized),
            "decision_counts": dict(decision_counts),
            "safe_to_apply": False,
            "safe_artifact": True,
            "action": "write_human_review_regression_fixtures",
            "reason": "Filled human review rows provide accept/replace/reject/unsure labels for future workflow comparisons.",
            "examples": normalized[:200],
        }
    ]
    if tag_counts.get("candidate_unreachable") or tag_counts.get("domain_variant_fix"):
        recommendations.append(
            {
                "type": "url_reachability",
                "title": "Candidate URL reachability and canonical variant checks",
                "count": tag_counts.get("candidate_unreachable", 0) + tag_counts.get("domain_variant_fix", 0),
                "safe_to_apply": False,
                "safe_artifact": True,
                "action": "verify_url_variants_before_accept",
                "reason": "Human review repeatedly found unreachable candidates, including cases where a www/protocol variant is the usable official URL.",
                "examples": [row for row in normalized if {"candidate_unreachable", "domain_variant_fix"} & set(row["note_tags"])][:50],
            }
        )
    if tag_counts.get("platform_profile_only"):
        recommendations.append(
            {
                "type": "platform_profile_evidence_only",
                "title": "Platform profiles should be evidence, not official URLs",
                "count": tag_counts["platform_profile_only"],
                "safe_to_apply": True,
                "action": "add_to_excluded_domains",
                "domain": "indiamart.com",
                "reason": "Human review flagged IndiaMART-style profiles as platform pages or uncertain evidence rather than independent official websites.",
            }
        )
    if (
        tag_counts.get("wrong_company")
        or tag_counts.get("service_mismatch")
        or tag_counts.get("name_or_logo_mismatch")
        or tag_counts.get("region_mismatch")
    ):
        recommendations.append(
            {
                "type": "identity_constraint",
                "title": "Human review confirms identity and service consistency constraints",
                "count": tag_counts.get("wrong_company", 0)
                + tag_counts.get("service_mismatch", 0)
                + tag_counts.get("name_or_logo_mismatch", 0)
                + tag_counts.get("region_mismatch", 0),
                "safe_to_apply": False,
                "safe_artifact": True,
                "action": "write_identity_regression_fixtures",
                "reason": "Wrong-company and service-mismatch labels should tighten tests before scoring logic changes.",
                "examples": [
                    {
                        "provider_id": row["provider_id"],
                        "provider_name": row["provider_name"],
                        "candidate_url": row["candidate_url"],
                        "candidate_domain": domain_from_url(row["candidate_url"]),
                        "agent_b_decision": row["manual_decision"],
                        "evidence_score": row["confidence"],
                        "counter_evidence": "; ".join(row["note_tags"]),
                        "reason_for_unsure": "human_review_identity_or_service_gap",
                    }
                    for row in normalized
                    if {"wrong_company", "service_mismatch", "name_or_logo_mismatch", "region_mismatch"} & set(row["note_tags"])
                ][:50],
            }
        )
    if tag_counts.get("confirmed_no_official") or tag_counts.get("wrong_candidate_no_replacement"):
        examples = [
            row
            for row in normalized
            if {"confirmed_no_official", "wrong_candidate_no_replacement"} & set(row["note_tags"])
        ][:100]
        recommendations.append(
            {
                "type": "no_official_regression",
                "title": "Human review confirms no reliable official website for some providers",
                "count": len(examples),
                "safe_to_apply": False,
                "safe_artifact": True,
                "action": "write_no_official_regression_fixtures",
                "reason": "No-official labels should prevent same-name forced matches and become regression fixtures rather than excluded-domain rules.",
                "examples": examples,
            }
        )
    return recommendations


def _normalize_human_review_row(row: dict[str, str]) -> dict:
    decision = _decision(row)
    candidate_url = _first(row, "current_or_candidate_url", "official_url", "candidate_url", "top_candidate_url")
    manual_url = _first(row, "your_true_official_url", "manual_url", "true_official_url")
    notes = _first(row, "your_notes", "notes", "manual_notes")
    error_type = _first(row, "error_type", "your_error_type", "error_reason")
    if decision == "reject" and manual_url:
        decision = "replace"
    return {
        "provider_id": _first(row, "provider_id"),
        "provider_name": _first(row, "provider_name"),
        "provider_detail_url": _first(row, "amazon_detail_url", "provider_detail_url"),
        "candidate_url": candidate_url,
        "manual_decision": decision,
        "manual_url": manual_url,
        "confidence": _first(row, "confidence"),
        "error_type": error_type,
        "notes": notes,
        "note_tags": _note_tags(notes, candidate_url, manual_url, decision, error_type),
        "expected_outcome": _expected_outcome(decision),
    }


def _expected_outcome(decision: str) -> str:
    if decision == "accept":
        return "accept_current"
    if decision == "replace":
        return "replace_with_manual_url"
    if decision == "reject":
        return "reject_candidate"
    return "needs_identity_review"


def _note_tags(notes: str, candidate_url: str, manual_url: str, decision: str, error_type: str = "") -> list[str]:
    text = f"{notes} {error_type}".casefold()
    tags = []
    if any(marker in text for marker in ["实际无官网", "no official", "no reliable official"]):
        if candidate_url:
            tags.append("wrong_candidate_no_replacement")
        tags.append("confirmed_no_official")
    if any(marker in text for marker in ["错误官网", "wrong official", "wrong website"]):
        tags.append("wrong_company")
    if any(marker in text for marker in ["没找到官网", "未找到官网", "not found official", "recall gap"]):
        tags.append("recall_gap")
    if any(marker in text for marker in ["无法打开", "打不开", "无法访问", "not open", "not working", "does not open", "timeout"]):
        tags.append("candidate_unreachable")
    if manual_url and candidate_url and domain_from_url(manual_url) == domain_from_url(candidate_url):
        tags.append("domain_variant_fix")
    if any(marker in text for marker in ["平台 profile", "平台profile", "平台页", "店铺页", "linkedin", "indiamart", "amazon listing", "社媒"]):
        tags.append("platform_profile_only")
    if any(marker in text for marker in ["不是", "另一个", "无关", "公司不同", "同名"]):
        tags.append("wrong_company")
    if any(marker in text for marker in ["服务内容不匹配", "服务类型", "业务类型", "具体内容不一致", "内容完全不一致", "行业", "主营业务"]):
        tags.append("service_mismatch")
    if any(marker in text for marker in ["地区不匹配", "国家", "provider location", "tld", "域名后缀"]):
        tags.append("region_mismatch")
    if any(marker in text for marker in ["名字不一致", "名称不一致", "logo不一致", "logo 和名字", "logo和名字", "品牌不匹配"]):
        tags.append("name_or_logo_mismatch")
    if any(marker in text for marker in ["不完全确定", "不确定", "缺少证据", "无法确认", "证据"]):
        tags.append("insufficient_evidence")
    if decision == "unsure" and "insufficient_evidence" not in tags:
        tags.append("insufficient_evidence")
    return _dedupe(tags)


def _decision(row: dict[str, str]) -> str:
    raw = _first(row, "manual_decision", "your_decision", "decision").strip().casefold()
    aliases = {
        "accept": "accept",
        "accpet": "accept",
        "approve": "accept",
        "approved": "accept",
        "接受": "accept",
        "确认": "accept",
        "replace": "replace",
        "替换": "replace",
        "reject": "reject",
        "rejected": "reject",
        "拒绝": "reject",
        "unsure": "unsure",
        "不确定": "unsure",
    }
    return aliases.get(raw, raw)


def _first(row: dict[str, str], *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _write_markdown(path: Path, summary: dict) -> None:
    lines = [
        "# Check and Suggestion Optimization Suggestions",
        "",
        "## Overall",
        "",
        "| Metric | Value |",
        "|---|---:|",
    ]
    for key, value in summary["overall"].items():
        lines.append(f"| {key} | {value} |")
    lines.extend(["", "## Recommendations", ""])
    for item in summary["recommendations"]:
        safe = "yes" if item.get("safe_to_apply") else "no"
        lines.append(f"- **{item.get('title')}** (`{item.get('type')}`, safe_to_apply={safe}): {item.get('reason')}")
        if item.get("domain"):
            lines.append(f"  - domain: `{item['domain']}`")
        if item.get("queries"):
            lines.append(f"  - queries: {', '.join(f'`{query}`' for query in item['queries'])}")
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def _read_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _read_table(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    if path.suffix.casefold() == ".xlsx":
        return _read_xlsx(path)
    return _read_rows(path)


def _read_xlsx(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return _read_simple_xlsx(path)
    workbook = load_workbook(path, data_only=False, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows())
    if not rows:
        return []
    headers = [_cell_text(cell) for cell in rows[0]]
    out = []
    for cells in rows[1:]:
        row = {
            headers[idx]: _cell_text(cells[idx])
            for idx in range(len(headers))
            if headers[idx]
        }
        if any(row.values()):
            out.append(row)
    return out


def _cell_text(cell) -> str:
    value = cell.value
    if isinstance(value, str) and value.startswith("="):
        url = _formula_to_url(value)
        if url:
            return url
    return str(value or "").strip()


def _formula_to_url(value: str) -> str:
    match = re.search(r'HYPERLINK\("([^"]+)"', value, flags=re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return ""


def _read_simple_xlsx(path: Path) -> list[dict[str, str]]:
    with zipfile.ZipFile(path) as z:
        shared = _read_shared_strings(z)
        sheet_names = sorted(name for name in z.namelist() if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"))
        if not sheet_names:
            return []
        text = z.read(sheet_names[0]).decode("utf-8", errors="replace")
    rows = []
    for row_xml in re.findall(r"<row\b[^>]*>(.*?)</row>", text):
        cells = []
        for cell_xml in re.findall(r"<c\b([^>]*)>(.*?)</c>", row_xml):
            attrs, body = cell_xml
            value_match = re.search(r"<v>(.*?)</v>", body)
            inline_match = re.search(r"<t[^>]*>(.*?)</t>", body)
            if 't="s"' in attrs and value_match:
                idx = int(value_match.group(1))
                cells.append(shared[idx] if idx < len(shared) else "")
            elif inline_match:
                cells.append(_xml_unescape(inline_match.group(1)))
            elif value_match:
                cells.append(_xml_unescape(value_match.group(1)))
            else:
                cells.append("")
        rows.append(cells)
    if not rows:
        return []
    headers = [value.strip() for value in rows[0]]
    return [
        {headers[idx]: values[idx] if idx < len(values) else "" for idx in range(len(headers)) if headers[idx]}
        for values in rows[1:]
        if any(values)
    ]


def _read_shared_strings(z: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in z.namelist():
        return []
    text = z.read("xl/sharedStrings.xml").decode("utf-8", errors="replace")
    return [_xml_unescape("".join(re.findall(r"<t[^>]*>(.*?)</t>", item))) for item in re.findall(r"<si\b[^>]*>(.*?)</si>", text)]


def _xml_unescape(value: str) -> str:
    return (
        value.replace("&quot;", '"')
        .replace("&apos;", "'")
        .replace("&lt;", "<")
        .replace("&gt;", ">")
        .replace("&amp;", "&")
    )


def _read_json(path: Path) -> dict:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def _looks_like_bad_domain(domain: str) -> bool:
    markers = {
        "linkedin.com",
        "facebook.com",
        "instagram.com",
        "youtube.com",
        "crunchbase.com",
        "trustpilot.com",
        "clutch.co",
        "goodfirms.co",
        "opencorporates.com",
    }
    return domain in markers or any(domain.endswith(f".{marker}") for marker in markers)


def _to_int(value: object) -> int:
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return 0


def _dedupe(values: list[str]) -> list[str]:
    out = []
    for value in values:
        if value and value not in out:
            out.append(value)
    return out


def _update_manifest(path: Path, summary: dict) -> None:
    if not path.exists():
        return
    manifest = json.loads(path.read_text(encoding="utf-8"))
    manifest["check_suggestion_suggestions"] = summary
    manifest.setdefault("outputs", {}).update(
        {f"check_suggestion_suggestions_{key}": value for key, value in summary["outputs"].items()}
    )
    path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    raise SystemExit(main())
