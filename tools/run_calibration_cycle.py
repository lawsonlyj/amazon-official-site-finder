from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from tools.build_calibration_review_sample import build_calibration_review_sample
from tools.build_balance_report import build_balance_report
from tools.build_calibration_label_gap_task import build_calibration_label_gap_task
from tools.build_calibration_regression_cases import build_calibration_regression_cases
from tools.build_calibration_status_report import build_calibration_status_report
from tools.check_calibration_application_gate import APPLICATION_GATES, check_calibration_application_gate
from tools.evaluate_calibration_review_sample import evaluate_calibration_review_sample
from tools.mine_evidence_patterns import mine_evidence_patterns
from tools.run_calibration_regression_gate import run_calibration_regression_gate
from tools.simulate_pattern_release import simulate_pattern_release
from tools.build_threshold_boundary_report import build_threshold_boundary_report


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Run the repeatable calibration-material generation cycle.")
    parser.add_argument("--labeled-eval-json", required=True, help="Labeled balance JSON from evaluate_workflow_balance.py.")
    parser.add_argument("--labeled-agent-b-csv", required=True, help="AgentB check.csv for the labeled calibration run.")
    parser.add_argument("--review-csv", required=True, help="Target batch review_task.csv.")
    parser.add_argument("--batch-agent-b-csv", required=True, help="Target batch agent_b/check.csv.")
    parser.add_argument("--batch-total-rows", type=int, default=0, help="Total rows in the target batch.")
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--sample-prefix", default="pattern_validation_sample_50")
    parser.add_argument("--max-rows", type=int, default=50)
    parser.add_argument("--max-per-reason", type=int, default=12)
    parser.add_argument("--max-per-pattern", type=int, default=5)
    parser.add_argument("--min-support", type=int, default=2)
    parser.add_argument("--max-pattern-size", type=int, default=3)
    parser.add_argument(
        "--filled-sample",
        action="append",
        default=[],
        help="Optional filled calibration sample CSV/XLSX to evaluate in the same cycle. Repeatable.",
    )
    parser.add_argument(
        "--pattern-release-json",
        action="append",
        default=[],
        help="Optional already-validated pattern release simulation JSON. Repeatable.",
    )
    parser.add_argument(
        "--policy-report-json",
        help="Optional final release policy JSON from tools/build_release_policy_report.py.",
    )
    parser.add_argument(
        "--candidate-final-csv",
        help="Optional candidate official_sites.csv/provider_final CSV to validate against filled calibration regression cases.",
    )
    args = parser.parse_args(argv)

    report = run_calibration_cycle(
        labeled_eval_json=args.labeled_eval_json,
        labeled_agent_b_csv=args.labeled_agent_b_csv,
        review_csv=args.review_csv,
        batch_agent_b_csv=args.batch_agent_b_csv,
        batch_total_rows=args.batch_total_rows,
        output_dir=args.output_dir,
        sample_prefix=args.sample_prefix,
        max_rows=args.max_rows,
        max_per_reason=args.max_per_reason,
        max_per_pattern=args.max_per_pattern,
        min_support=args.min_support,
        max_pattern_size=args.max_pattern_size,
        filled_sample=args.filled_sample,
        pattern_release_jsons=args.pattern_release_json,
        policy_report_json=args.policy_report_json,
        candidate_final_csv=args.candidate_final_csv,
    )
    print(json.dumps(report["summary"], ensure_ascii=False, indent=2))
    return 0


def run_calibration_cycle(
    *,
    labeled_eval_json: str | Path,
    labeled_agent_b_csv: str | Path,
    review_csv: str | Path,
    batch_agent_b_csv: str | Path,
    output_dir: str | Path,
    batch_total_rows: int = 0,
    sample_prefix: str = "pattern_validation_sample_50",
    max_rows: int = 50,
    max_per_reason: int = 12,
    max_per_pattern: int = 5,
    min_support: int = 2,
    max_pattern_size: int = 3,
    filled_sample: str | Path | list[str | Path] | None = None,
    pattern_release_jsons: list[str | Path] | None = None,
    policy_report_json: str | Path | None = None,
    candidate_final_csv: str | Path | None = None,
) -> dict:
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    recall_json = out_dir / "evidence_patterns_recall.json"
    recall_md = out_dir / "evidence_patterns_recall.md"
    precision_json = out_dir / "evidence_patterns_precision.json"
    precision_md = out_dir / "evidence_patterns_precision.md"
    release_sim_json = out_dir / "pattern_release_simulation.json"
    release_sim_md = out_dir / "pattern_release_simulation.md"
    balance_report_json = out_dir / "balance_report.json"
    balance_report_md = out_dir / "balance_report.md"
    threshold_boundary_json = out_dir / "threshold_boundary_report.json"
    threshold_boundary_md = out_dir / "threshold_boundary_report.md"
    sample_csv = out_dir / f"{sample_prefix}.csv"
    sample_xlsx = out_dir / f"{sample_prefix}.xlsx"
    eval_json = out_dir / f"{sample_prefix}_eval_empty.json"
    eval_md = out_dir / f"{sample_prefix}_eval_empty.md"
    eval_csv = out_dir / f"{sample_prefix}_eval_empty_details.csv"
    filled_eval_json = out_dir / f"{sample_prefix}_eval_filled.json"
    filled_eval_md = out_dir / f"{sample_prefix}_eval_filled.md"
    filled_eval_csv = out_dir / f"{sample_prefix}_eval_filled_details.csv"
    filled_samples_merged_csv = out_dir / f"{sample_prefix}_filled_samples_merged.csv"
    rule_candidates_json = out_dir / "pattern_rule_candidates.json"
    rule_candidates_md = out_dir / "pattern_rule_candidates.md"
    regression_cases_csv = out_dir / "calibration_regression_cases.csv"
    regression_cases_json = out_dir / "calibration_regression_cases.json"
    regression_cases_md = out_dir / "calibration_regression_cases.md"
    regression_gate_csv = out_dir / "calibration_regression_gate.csv"
    regression_gate_json = out_dir / "calibration_regression_gate.json"
    regression_gate_md = out_dir / "calibration_regression_gate.md"
    label_gap_csv = out_dir / "label_gap_task.csv"
    label_gap_xlsx = out_dir / "label_gap_task.xlsx"
    label_gap_high_csv = out_dir / "label_gap_high_priority_task.csv"
    label_gap_high_xlsx = out_dir / "label_gap_high_priority_task.xlsx"
    summary_json = out_dir / "calibration_cycle_summary.json"
    summary_md = out_dir / "calibration_cycle_summary.md"
    status_json = out_dir / "calibration_status.json"
    status_md = out_dir / "calibration_status.md"
    application_gates_json = out_dir / "calibration_application_gates.json"
    application_gates_md = out_dir / "calibration_application_gates.md"

    recall_report = mine_evidence_patterns(
        balance_json=labeled_eval_json,
        agent_b_csv=labeled_agent_b_csv,
        scope="recall",
        max_pattern_size=max_pattern_size,
        min_support=min_support,
        output_json=recall_json,
        output_md=recall_md,
    )
    precision_report = mine_evidence_patterns(
        balance_json=labeled_eval_json,
        agent_b_csv=labeled_agent_b_csv,
        scope="precision",
        max_pattern_size=max_pattern_size,
        min_support=min_support,
        output_json=precision_json,
        output_md=precision_md,
    )
    release_simulation = simulate_pattern_release(
        balance_json=labeled_eval_json,
        agent_b_csv=labeled_agent_b_csv,
        pattern_jsons=[recall_json],
        scope="recall",
        min_support=min_support,
        output_json=release_sim_json,
        output_md=release_sim_md,
    )
    pattern_release_inputs = [release_sim_json, *[Path(path) for path in (pattern_release_jsons or [])]]
    preferred_pattern_release = _preferred_pattern_release_json(release_sim_json, pattern_release_jsons or [])
    balance_report = build_balance_report(
        labeled_eval_json=labeled_eval_json,
        batch_review_csv=review_csv,
        batch_agent_b_csv=batch_agent_b_csv,
        batch_total_rows=batch_total_rows,
        pattern_release_jsons=pattern_release_inputs,
        output_json=balance_report_json,
        output_md=balance_report_md,
    )
    threshold_boundary = build_threshold_boundary_report(
        labeled_eval_json=labeled_eval_json,
        pattern_release_json=preferred_pattern_release,
        policy_report_json=policy_report_json,
        output_json=threshold_boundary_json,
        output_md=threshold_boundary_md,
    )
    sample_summary = build_calibration_review_sample(
        review_csv=review_csv,
        agent_b_csv=batch_agent_b_csv,
        output_csv=sample_csv,
        output_xlsx=sample_xlsx,
        max_rows=max_rows,
        max_per_reason=max_per_reason,
        max_per_pattern=max_per_pattern,
        pattern_jsons=[*pattern_release_inputs, recall_json, precision_json],
    )
    empty_eval = evaluate_calibration_review_sample(
        sample=sample_xlsx,
        output_json=eval_json,
        output_md=eval_md,
        output_csv=eval_csv,
    )
    filled_sample_paths = _filled_sample_paths(filled_sample)
    filled_eval_sample = _filled_eval_sample_path(filled_sample_paths, filled_samples_merged_csv)
    filled_eval = {}
    regression_cases = {}
    regression_gate = {}
    if filled_eval_sample:
        filled_eval = evaluate_calibration_review_sample(
            sample=filled_eval_sample,
            output_json=filled_eval_json,
            output_md=filled_eval_md,
            output_csv=filled_eval_csv,
        )
        _write_rule_candidates(
            filled_eval.get("pattern_rule_candidates", {}),
            rule_candidates_json,
            rule_candidates_md,
        )
        regression_cases = build_calibration_regression_cases(
            sample_eval_json=filled_eval_json,
            output_csv=regression_cases_csv,
            output_json=regression_cases_json,
            output_md=regression_cases_md,
        )
        if candidate_final_csv:
            regression_gate = run_calibration_regression_gate(
                cases_csv=regression_cases_csv,
                candidate_final_csv=candidate_final_csv,
                output_csv=regression_gate_csv,
                output_json=regression_gate_json,
                output_md=regression_gate_md,
            )
    pattern_recommendation_counts = _pattern_recommendation_counts(filled_eval)
    lane_recommendation_counts = _lane_recommendation_counts(filled_eval)
    report = {
        "summary": {
            "recall_durable_safe_patterns": recall_report["summary"].get("durable_safe_patterns"),
            "precision_durable_safe_patterns": precision_report["summary"].get("durable_safe_patterns"),
            "release_safe_patterns": release_simulation["summary"].get("safe_pattern_count"),
            "release_actionable_safe_patterns": release_simulation["summary"].get("actionable_safe_pattern_count"),
            "best_actionable_release_pattern": release_simulation["summary"].get("best_actionable_safe_pattern"),
            "best_actionable_release_correct_rows": release_simulation["summary"].get("best_actionable_safe_correct_recovery_rows"),
            "best_actionable_release_wrong_rows": release_simulation["summary"].get("best_actionable_safe_wrong_release_rows"),
            "best_actionable_release_accuracy": release_simulation["summary"].get("best_actionable_safe_accuracy"),
            "selected_actionable_release_patterns": release_simulation["summary"].get("selected_actionable_pattern_count"),
            "selected_actionable_release_correct_rows": release_simulation["summary"].get("selected_actionable_correct_recovery_rows"),
            "selected_actionable_release_wrong_rows": release_simulation["summary"].get("selected_actionable_wrong_release_rows"),
            "selected_actionable_release_accuracy": release_simulation["summary"].get("selected_actionable_accuracy"),
            "recommended_global_accept_threshold": threshold_boundary["summary"].get(
                "recommended_global_accept_threshold"
            ),
            "recommended_second_pass_threshold": threshold_boundary["summary"].get("recommended_second_pass_threshold"),
            "precision_watch_min": threshold_boundary["summary"].get("precision_watch_min"),
            "precision_watch_max": threshold_boundary["summary"].get("precision_watch_max"),
            "recommended_matched_review_confidence_below": threshold_boundary["summary"].get(
                "recommended_matched_review_confidence_below"
            ),
            "raw_agent_b_recall_release": threshold_boundary["summary"].get("raw_agent_b_recall_release"),
            "calibrated_pattern_release": threshold_boundary["summary"].get("calibrated_pattern_release"),
            "recommended_pattern_release": balance_report["summary"].get("recommended_pattern_release"),
            "recommended_pattern_release_source_path": balance_report["summary"].get("pattern_release_source_path", ""),
            "recommended_pattern_release_source_kind": _pattern_release_source_kind(
                balance_report["summary"].get("pattern_release_source_path", ""),
                release_sim_json,
                pattern_release_jsons or [],
            ),
            "pattern_release_correct_rows": balance_report["summary"].get("pattern_release_correct_rows"),
            "pattern_release_wrong_rows": balance_report["summary"].get("pattern_release_wrong_rows"),
            "protected_review_lane_count": balance_report["summary"].get("protected_review_lane_count"),
            "protected_review_lane_rows": balance_report["summary"].get("protected_review_lane_rows"),
            "spot_check_candidate_lanes": balance_report["summary"].get("spot_check_candidate_lanes"),
            "more_label_review_lanes": balance_report["summary"].get("more_label_review_lanes"),
            "sample_rows": sample_summary.get("sample_rows"),
            "actionable_release_validation_rows": sample_summary.get("sample_reason_counts", {}).get(
                "actionable_release_validation", 0
            ),
            "pattern_validation_rows": sample_summary.get("sample_reason_counts", {}).get("pattern_candidate_validation", 0),
            "pattern_control_rows": sample_summary.get("sample_reason_counts", {}).get("pattern_control_validation", 0),
            "timeout_rows": sample_summary.get("sample_reason_counts", {}).get("timeout_needs_manual", 0),
            "pattern_count": len(sample_summary.get("pattern_match_counts", {})),
            "max_per_pattern": sample_summary.get("max_per_pattern"),
            "empty_eval_labeled_rows": empty_eval["summary"].get("labeled_rows"),
            "filled_eval_labeled_rows": filled_eval.get("summary", {}).get("labeled_rows") if filled_eval else None,
            "filled_eval_decisive_rows": filled_eval.get("summary", {}).get("decisive_rows") if filled_eval else None,
            "filled_pattern_recommendation_counts": pattern_recommendation_counts,
            "filled_lane_recommendation_counts": lane_recommendation_counts,
            "filled_lane_candidate_for_change_count": filled_eval.get("summary", {}).get("lane_candidate_for_change_rows")
            if filled_eval
            else None,
            "filled_lane_keep_review_count": filled_eval.get("summary", {}).get("lane_keep_review_rows")
            if filled_eval
            else None,
            "filled_rule_candidate_count": len(
                filled_eval.get("pattern_rule_candidates", {}).get("candidate_for_rule", [])
            )
            if filled_eval
            else None,
            "filled_rejected_pattern_count": len(
                filled_eval.get("pattern_rule_candidates", {}).get("reject_pattern", [])
            )
            if filled_eval
            else None,
            "filled_regression_case_rows": regression_cases.get("summary", {}).get("case_rows") if regression_cases else None,
            "filled_precision_blocking_fixture_rows": regression_cases.get("summary", {}).get("precision_blocking_fixture_rows")
            if regression_cases
            else None,
            "filled_recall_blocking_fixture_rows": regression_cases.get("summary", {}).get("recall_blocking_fixture_rows")
            if regression_cases
            else None,
            "filled_positive_fixture_rows": regression_cases.get("summary", {}).get("positive_fixture_rows")
            if regression_cases
            else None,
            "regression_gate_next_step": _regression_gate_next_step(regression_cases_csv) if regression_cases else "",
            "regression_gate_status": regression_gate.get("summary", {}).get("gate_status") if regression_gate else "",
            "regression_gate_fail_rows": regression_gate.get("summary", {}).get("fail_rows") if regression_gate else None,
            "regression_gate_unverified_rows": regression_gate.get("summary", {}).get("unverified_rows")
            if regression_gate
            else None,
        },
        "inputs": {
            "labeled_eval_json": str(labeled_eval_json),
            "labeled_agent_b_csv": str(labeled_agent_b_csv),
            "review_csv": str(review_csv),
            "batch_agent_b_csv": str(batch_agent_b_csv),
            "batch_total_rows": str(batch_total_rows),
            "pattern_release_jsons": [str(path) for path in (pattern_release_jsons or [])],
            "preferred_pattern_release_json": str(preferred_pattern_release),
            "filled_sample": str(filled_sample_paths[0]) if len(filled_sample_paths) == 1 else "",
            "filled_samples": [str(path) for path in filled_sample_paths],
            "policy_report_json": str(policy_report_json or ""),
            "candidate_final_csv": str(candidate_final_csv or ""),
        },
        "outputs": {
            "recall_json": str(recall_json),
            "recall_md": str(recall_md),
            "precision_json": str(precision_json),
            "precision_md": str(precision_md),
            "release_simulation_json": str(release_sim_json),
            "release_simulation_md": str(release_sim_md),
            "balance_report_json": str(balance_report_json),
            "balance_report_md": str(balance_report_md),
            "threshold_boundary_json": str(threshold_boundary_json),
            "threshold_boundary_md": str(threshold_boundary_md),
            "sample_csv": str(sample_csv),
            "sample_xlsx": str(sample_xlsx),
            "eval_json": str(eval_json),
            "eval_md": str(eval_md),
            "eval_csv": str(eval_csv),
            "filled_samples_merged_csv": str(filled_samples_merged_csv) if len(filled_sample_paths) > 1 else "",
            "filled_eval_json": str(filled_eval_json) if filled_eval_sample else "",
            "filled_eval_md": str(filled_eval_md) if filled_eval_sample else "",
            "filled_eval_csv": str(filled_eval_csv) if filled_eval_sample else "",
            "rule_candidates_json": str(rule_candidates_json) if filled_eval_sample else "",
            "rule_candidates_md": str(rule_candidates_md) if filled_eval_sample else "",
            "regression_cases_csv": str(regression_cases_csv) if filled_eval_sample else "",
            "regression_cases_json": str(regression_cases_json) if filled_eval_sample else "",
            "regression_cases_md": str(regression_cases_md) if filled_eval_sample else "",
            "regression_gate_csv": str(regression_gate_csv) if regression_gate else "",
            "regression_gate_json": str(regression_gate_json) if regression_gate else "",
            "regression_gate_md": str(regression_gate_md) if regression_gate else "",
            "label_gap_csv": str(label_gap_csv),
            "label_gap_xlsx": str(label_gap_xlsx),
            "label_gap_high_priority_csv": str(label_gap_high_csv),
            "label_gap_high_priority_xlsx": str(label_gap_high_xlsx),
            "summary_json": str(summary_json),
            "summary_md": str(summary_md),
            "status_json": str(status_json),
            "status_md": str(status_md),
            "application_gates_json": str(application_gates_json),
            "application_gates_md": str(application_gates_md),
        },
        "recall_recommendations": recall_report.get("recommendations", []),
        "precision_recommendations": precision_report.get("recommendations", []),
        "release_simulation_summary": release_simulation.get("summary", {}),
        "pattern_release_inputs": [str(path) for path in pattern_release_inputs],
        "balance_report": balance_report,
        "threshold_boundary": threshold_boundary,
        "actionable_release_patterns": release_simulation.get("actionable_safe_patterns", []),
        "selected_actionable_release_patterns": release_simulation.get("selected_actionable_pattern_set", []),
        "sample": sample_summary,
        "empty_evaluation_summary": empty_eval.get("summary", {}),
        "filled_evaluation_summary": filled_eval.get("summary", {}) if filled_eval else {},
        "filled_lane_recommendations": filled_eval.get("lane_recommendations", []) if filled_eval else [],
        "filled_pattern_recommendations": filled_eval.get("pattern_recommendations", []) if filled_eval else [],
        "filled_pattern_rule_candidates": filled_eval.get("pattern_rule_candidates", {}) if filled_eval else {},
        "filled_regression_cases": regression_cases,
        "regression_gate": regression_gate,
    }
    summary_json.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    summary_md.write_text(_render_markdown(report), encoding="utf-8")
    status_report = build_calibration_status_report(
        calibration_cycle_json=summary_json,
        balance_report_json=balance_report_json,
        threshold_boundary_json=threshold_boundary_json,
        sample_eval_json=filled_eval_json if filled_eval_sample else eval_json,
        output_json=status_json,
        output_md=status_md,
    )
    label_gap_task = build_calibration_label_gap_task(
        status_json=status_json,
        sample_csv=sample_csv,
        filled_sample=filled_eval_sample,
        output_csv=label_gap_csv,
        output_xlsx=label_gap_xlsx,
    )
    high_priority_label_gap_task = build_calibration_label_gap_task(
        status_json=status_json,
        sample_csv=sample_csv,
        filled_sample=filled_eval_sample,
        output_csv=label_gap_high_csv,
        output_xlsx=label_gap_high_xlsx,
        priorities=["high"],
    )
    report["calibration_status"] = status_report
    report["label_gap_task"] = label_gap_task
    report["label_gap_high_priority_task"] = high_priority_label_gap_task
    report["summary"]["label_gap_task_rows"] = label_gap_task.get("task_rows")
    report["summary"]["label_gap_high_priority_rows"] = label_gap_task.get("priority_counts", {}).get("high", 0)
    report["summary"]["label_gap_high_priority_task_rows"] = high_priority_label_gap_task.get("task_rows")
    summary_json.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    status_report = build_calibration_status_report(
        calibration_cycle_json=summary_json,
        balance_report_json=balance_report_json,
        threshold_boundary_json=threshold_boundary_json,
        sample_eval_json=filled_eval_json if filled_eval_sample else eval_json,
        output_json=status_json,
        output_md=status_md,
    )
    report["calibration_status"] = status_report
    application_gate_checks = _write_application_gate_checks(status_json, application_gates_json, application_gates_md)
    report["application_gate_checks"] = application_gate_checks
    report["summary"]["application_gate_allowed_count"] = application_gate_checks["summary"]["allowed_gate_count"]
    report["summary"]["application_gate_not_allowed_count"] = application_gate_checks["summary"]["not_allowed_gate_count"]
    report["summary"]["application_gate_candidate_count"] = application_gate_checks["summary"]["candidate_gate_count"]
    summary_json.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    summary_md.write_text(_render_markdown(report), encoding="utf-8")
    return report


def _render_markdown(report: dict) -> str:
    summary = report["summary"]
    lines = [
        "# Calibration Cycle Summary",
        "",
        "## Summary",
        "",
        f"- Recall durable safe patterns: {summary['recall_durable_safe_patterns']}",
        f"- Precision durable safe patterns: {summary['precision_durable_safe_patterns']}",
        f"- Recall release safe patterns: {summary['release_safe_patterns']}",
        f"- Recall release actionable safe patterns: {summary['release_actionable_safe_patterns']}",
        f"- Best actionable release pattern: {summary['best_actionable_release_pattern'] or 'None'}",
        f"- Best actionable release correct/wrong rows: {summary['best_actionable_release_correct_rows']}/{summary['best_actionable_release_wrong_rows']}",
        f"- Best actionable release accuracy: {summary['best_actionable_release_accuracy']}",
        f"- Selected actionable release patterns: {summary['selected_actionable_release_patterns']}",
        f"- Selected actionable release correct/wrong rows: {summary['selected_actionable_release_correct_rows']}/{summary['selected_actionable_release_wrong_rows']}",
        f"- Selected actionable release accuracy: {summary['selected_actionable_release_accuracy']}",
        f"- Recommended global accept threshold: {summary['recommended_global_accept_threshold']}",
        f"- Recommended second-pass threshold: {summary['recommended_second_pass_threshold']}",
        f"- Precision watch score band: {summary['precision_watch_min']}-{summary['precision_watch_max']}",
        f"- Matched-review confidence cutoff: <{summary['recommended_matched_review_confidence_below']}",
        f"- Raw AgentB recall release: {summary['raw_agent_b_recall_release']}",
        f"- Calibrated pattern release: {summary['calibrated_pattern_release']}",
        f"- Recommended pattern release: {summary['recommended_pattern_release']}",
        f"- Recommended pattern release source: {summary.get('recommended_pattern_release_source_path') or 'not_evaluated'}",
        f"- Recommended pattern release source kind: {summary.get('recommended_pattern_release_source_kind') or 'not_evaluated'}",
        f"- Pattern release correct/wrong rows: {summary['pattern_release_correct_rows']}/{summary['pattern_release_wrong_rows']}",
        f"- Protected review lanes: {summary['protected_review_lane_count']}",
        f"- Protected review lane rows: {summary['protected_review_lane_rows']}",
        f"- Spot-check candidate lanes: {', '.join(summary.get('spot_check_candidate_lanes') or []) or 'None'}",
        f"- More-label review lanes: {', '.join(summary.get('more_label_review_lanes') or []) or 'None'}",
        f"- Sample rows: {summary['sample_rows']}",
        f"- Label-gap task rows: {summary.get('label_gap_task_rows')}",
        f"- Label-gap high-priority rows: {summary.get('label_gap_high_priority_rows')}",
        f"- Label-gap high-priority task rows: {summary.get('label_gap_high_priority_task_rows')}",
        f"- Actionable release validation rows: {summary['actionable_release_validation_rows']}",
        f"- Pattern candidate validation rows: {summary['pattern_validation_rows']}",
        f"- Pattern control validation rows: {summary['pattern_control_rows']}",
        f"- Timeout rows: {summary['timeout_rows']}",
        f"- Pattern count: {summary['pattern_count']}",
        f"- Max per pattern: {summary['max_per_pattern']}",
        f"- Empty evaluation labeled rows: {summary['empty_eval_labeled_rows']}",
        f"- Filled evaluation labeled rows: {summary['filled_eval_labeled_rows']}",
        f"- Filled evaluation decisive rows: {summary['filled_eval_decisive_rows']}",
        f"- Filled lane candidate-for-change count: {summary['filled_lane_candidate_for_change_count']}",
        f"- Filled lane keep-review count: {summary['filled_lane_keep_review_count']}",
        f"- Filled candidate-for-rule patterns: {summary['filled_rule_candidate_count']}",
        f"- Filled rejected patterns: {summary['filled_rejected_pattern_count']}",
        f"- Filled regression case rows: {summary.get('filled_regression_case_rows')}",
        f"- Filled precision blocking fixtures: {summary.get('filled_precision_blocking_fixture_rows')}",
        f"- Filled recall blocking fixtures: {summary.get('filled_recall_blocking_fixture_rows')}",
        f"- Filled positive fixtures: {summary.get('filled_positive_fixture_rows')}",
        f"- Regression gate status: {summary.get('regression_gate_status') or 'not_run'}",
        f"- Regression gate fail/unverified rows: {summary.get('regression_gate_fail_rows')}/{summary.get('regression_gate_unverified_rows')}",
        f"- Application gates allowed/not allowed/candidate: {summary.get('application_gate_allowed_count')}/{summary.get('application_gate_not_allowed_count')}/{summary.get('application_gate_candidate_count')}",
        "",
        "## Outputs",
        "",
    ]
    for key, value in report["outputs"].items():
        lines.append(f"- {key}: {value}")
    lines.extend(["", "## Recall Recommendations", ""])
    for item in report.get("recall_recommendations", []):
        lines.append(f"- {item}")
    lines.extend(["", "## Precision Recommendations", ""])
    for item in report.get("precision_recommendations", []):
        lines.append(f"- {item}")
    lane_policy = (report.get("balance_report") or {}).get("manual_review_lane_policy") or {}
    if lane_policy:
        lines.extend(["", "## Review Lane Policy", ""])
        protected = lane_policy.get("protected") or []
        if protected:
            lines.append("### Protected")
            for row in protected:
                lines.append(
                    "- {reason}: rows={rows}, risk={risk}, {why}".format(
                        reason=row.get("review_reason"),
                        rows=row.get("review_task_rows"),
                        risk=row.get("risk_rows"),
                        why=row.get("protection_reason"),
                    )
                )
        spot_check = lane_policy.get("spot_check_candidates") or []
        if spot_check:
            lines.append("### Spot-Check Candidates")
            for row in spot_check:
                lines.append(
                    "- {reason}: rows={rows}, labeled={labeled}".format(
                        reason=row.get("review_reason"),
                        rows=row.get("review_task_rows"),
                        labeled=row.get("labeled_rows"),
                    )
                )
        needs_more = lane_policy.get("needs_more_labels") or []
        if needs_more:
            lines.append("### Needs More Labels")
            for row in needs_more:
                lines.append(
                    "- {reason}: rows={rows}, labeled={labeled}".format(
                        reason=row.get("review_reason"),
                        rows=row.get("review_task_rows"),
                        labeled=row.get("labeled_rows"),
                    )
                )
    if report.get("actionable_release_patterns"):
        lines.extend(["", "## Actionable Recall Release Patterns", ""])
        for item in report["actionable_release_patterns"][:10]:
            lines.append(
                "- correct={correct}, wrong={wrong}, accuracy={accuracy}: {pattern}".format(
                    correct=item.get("correct_recovery_rows"),
                    wrong=item.get("wrong_release_rows"),
                    accuracy=item.get("simulated_overall", {}).get("overall_accuracy"),
                    pattern=item.get("pattern"),
                )
            )
    if report.get("selected_actionable_release_patterns"):
        lines.extend(["", "## Selected Actionable Recall Release Set", ""])
        for item in report["selected_actionable_release_patterns"][:10]:
            lines.append(
                "- correct={correct}, wrong={wrong}, accuracy={accuracy}: {pattern}".format(
                    correct=item.get("correct_recovery_rows"),
                    wrong=item.get("wrong_release_rows"),
                    accuracy=item.get("simulated_overall", {}).get("overall_accuracy"),
                    pattern=item.get("pattern"),
                )
            )
    if report.get("filled_lane_recommendations"):
        lines.extend(["", "## Filled Lane Recommendations", ""])
        counts = summary.get("filled_lane_recommendation_counts", {})
        for key, value in sorted(counts.items()):
            lines.append(f"- {key}: {value}")
        lines.append("")
        for item in report["filled_lane_recommendations"][:20]:
            lines.append(
                "- {recommendation}: rows={rows}, decisive={decisive}, good={good}, bad={bad}, recall_useful={recall_useful}, recall_bad={recall_bad} :: {reason}".format(
                    recommendation=item.get("recommendation"),
                    rows=item.get("rows"),
                    decisive=item.get("decisive_rows"),
                    good=item.get("candidate_correct_rows"),
                    bad=item.get("candidate_incorrect_rows"),
                    recall_useful=item.get("recall_useful_rows"),
                    recall_bad=item.get("recall_not_useful_rows"),
                    reason=item.get("review_reason"),
                )
            )
    if report.get("filled_pattern_recommendations"):
        lines.extend(["", "## Filled Pattern Recommendations", ""])
        counts = summary.get("filled_pattern_recommendation_counts", {})
        for key, value in sorted(counts.items()):
            lines.append(f"- {key}: {value}")
        lines.append("")
        for item in report["filled_pattern_recommendations"][:20]:
            lines.append(
                "- {recommendation}: rows={rows}, decisive={decisive}, support={support}, block={block} :: {pattern}".format(
                    recommendation=item.get("recommendation"),
                    rows=item.get("rows"),
                    decisive=item.get("decisive_rows"),
                    support=item.get("supporting_rows"),
                    block=item.get("blocking_rows"),
                    pattern=item.get("pattern"),
                )
            )
    rule_candidates = report.get("filled_pattern_rule_candidates", {})
    if rule_candidates:
        lines.extend(["", "## Filled Candidate Rule Export", ""])
        for key, title in [
            ("candidate_for_rule", "Candidate For Rule"),
            ("needs_more_labels", "Needs More Labels"),
            ("reject_pattern", "Rejected Pattern"),
        ]:
            items = rule_candidates.get(key) or []
            lines.append(f"### {title}")
            if not items:
                lines.append("- None")
                continue
            for item in items[:10]:
                lines.append(
                    "- scope={scope}, support={support}, block={block}: {pattern} -- {action}".format(
                        scope=item.get("pattern_scope", ""),
                        support=item.get("supporting_rows"),
                        block=item.get("blocking_rows"),
                        pattern=item.get("pattern"),
                        action=item.get("required_action", ""),
                    )
                )
    regression_cases = report.get("filled_regression_cases") or {}
    if regression_cases.get("summary"):
        lines.extend(["", "## Filled Regression Cases", ""])
        case_summary = regression_cases["summary"]
        lines.append(f"- Case rows: {case_summary.get('case_rows')}")
        lines.append(f"- Precision blocking fixtures: {case_summary.get('precision_blocking_fixture_rows')}")
        lines.append(f"- Recall blocking fixtures: {case_summary.get('recall_blocking_fixture_rows')}")
        lines.append(f"- Positive fixtures: {case_summary.get('positive_fixture_rows')}")
        if summary.get("regression_gate_next_step"):
            lines.append(f"- Regression gate next step: {summary.get('regression_gate_next_step')}")
    regression_gate = report.get("regression_gate") or {}
    if regression_gate.get("summary"):
        lines.extend(["", "## Regression Gate", ""])
        gate_summary = regression_gate["summary"]
        lines.append(f"- Gate status: {gate_summary.get('gate_status')}")
        lines.append(f"- Pass/fail/unverified: {gate_summary.get('pass_rows')}/{gate_summary.get('fail_rows')}/{gate_summary.get('unverified_rows')}")
    application_gate_checks = report.get("application_gate_checks") or {}
    if application_gate_checks.get("checks"):
        lines.extend(["", "## Application Gates", ""])
        for item in application_gate_checks["checks"]:
            lines.append(
                "- {gate}: allowed={allowed}, status={status}, blockers={blockers} :: {reason}".format(
                    gate=item.get("gate"),
                    allowed=str(item.get("allowed")).lower(),
                    status=item.get("gate_status"),
                    blockers=", ".join(item.get("blockers") or []) or "none",
                    reason=item.get("reason") or item.get("decision_reason") or "",
                )
            )
    lines.append("")
    return "\n".join(lines)


def _write_application_gate_checks(status_json: Path, output_json: Path, output_md: Path) -> dict:
    checks = [
        check_calibration_application_gate(status_json=status_json, gate=gate)["summary"]
        for gate in sorted(APPLICATION_GATES)
    ]
    allowed = [row["gate"] for row in checks if row.get("allowed")]
    not_allowed = [row["gate"] for row in checks if not row.get("allowed")]
    candidates = [row["gate"] for row in checks if row.get("gate_status") == "candidate"]
    report = {
        "summary": {
            "gate_count": len(checks),
            "allowed_gate_count": len(allowed),
            "not_allowed_gate_count": len(not_allowed),
            "candidate_gate_count": len(candidates),
            "allowed_gates": allowed,
            "not_allowed_gates": not_allowed,
            "candidate_gates": candidates,
        },
        "inputs": {"status_json": str(status_json)},
        "checks": checks,
    }
    output_json.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    output_md.write_text(_render_application_gate_checks_markdown(report), encoding="utf-8")
    return report


def _render_application_gate_checks_markdown(report: dict) -> str:
    summary = report["summary"]
    lines = [
        "# Calibration Application Gates",
        "",
        f"- Gate count: {summary['gate_count']}",
        f"- Allowed gates: {summary['allowed_gate_count']}",
        f"- Not allowed gates: {summary['not_allowed_gate_count']}",
        f"- Candidate gates: {summary['candidate_gate_count']}",
        "",
        "## Checks",
        "",
    ]
    for row in report.get("checks", []):
        blockers = ", ".join(row.get("blockers") or []) or "none"
        lines.extend(
            [
                f"### {row.get('gate')}",
                f"- Allowed: {str(row.get('allowed')).lower()}",
                f"- Gate status: {row.get('gate_status')}",
                f"- Can apply now: {str(row.get('can_apply_now')).lower()}",
                f"- Blockers: {blockers}",
                f"- Decision reason: {row.get('decision_reason')}",
                f"- Reason: {row.get('reason') or 'not recorded'}",
                f"- Required action: {row.get('required_action') or 'not recorded'}",
                "",
            ]
        )
    return "\n".join(lines)


def _write_rule_candidates(rule_candidates: dict, output_json: Path, output_md: Path) -> None:
    output_json.write_text(json.dumps(rule_candidates, ensure_ascii=False, indent=2), encoding="utf-8")
    output_md.write_text(_render_rule_candidates_markdown(rule_candidates), encoding="utf-8")


def _regression_gate_next_step(regression_cases_csv: Path) -> str:
    return (
        "Run tools/run_calibration_regression_gate.py with "
        f"--cases-csv {regression_cases_csv} and --candidate-final-csv <candidate official_sites.csv> before applying threshold or routing changes."
    )


def _render_rule_candidates_markdown(rule_candidates: dict) -> str:
    lines = [
        "# Pattern Rule Candidates",
        "",
        "These are advisory outputs from filled calibration labels. Production workflow rules should only change after matching regression tests are added.",
        "",
    ]
    for key, title in [
        ("candidate_for_rule", "Candidate For Rule"),
        ("needs_more_labels", "Needs More Labels"),
        ("reject_pattern", "Rejected Pattern"),
    ]:
        lines.extend([f"## {title}", ""])
        items = rule_candidates.get(key) or []
        if not items:
            lines.append("- None")
            lines.append("")
            continue
        for item in items:
            lines.append(
                "- scope={scope}, rows={rows}, decisive={decisive}, support={support}, block={block}: {pattern}".format(
                    scope=item.get("pattern_scope", ""),
                    rows=item.get("rows"),
                    decisive=item.get("decisive_rows"),
                    support=item.get("supporting_rows"),
                    block=item.get("blocking_rows"),
                    pattern=item.get("pattern"),
                )
            )
            lines.append(f"  Action: {item.get('required_action', '')}")
        lines.append("")
    return "\n".join(lines)


def _pattern_recommendation_counts(filled_eval: dict) -> dict[str, int]:
    counts: dict[str, int] = {}
    for item in filled_eval.get("pattern_recommendations", []) if filled_eval else []:
        recommendation = str(item.get("recommendation") or "")
        if recommendation:
            counts[recommendation] = counts.get(recommendation, 0) + 1
    return counts


def _lane_recommendation_counts(filled_eval: dict) -> dict[str, int]:
    counts: dict[str, int] = {}
    for item in filled_eval.get("lane_recommendations", []) if filled_eval else []:
        recommendation = str(item.get("recommendation") or "")
        if recommendation:
            counts[recommendation] = counts.get(recommendation, 0) + 1
    return counts


def _filled_sample_paths(filled_sample: str | Path | list[str | Path] | None) -> list[Path]:
    if not filled_sample:
        return []
    if isinstance(filled_sample, (str, Path)):
        return [Path(filled_sample)]
    return [Path(path) for path in filled_sample if str(path)]


def _filled_eval_sample_path(paths: list[Path], merged_output: Path) -> Path | None:
    existing = [path for path in paths if path.exists()]
    if not existing:
        return None
    if len(existing) == 1:
        return existing[0]
    rows_by_key: dict[tuple[str, str], dict[str, str]] = {}
    ordered_keys: list[tuple[str, str]] = []
    unkeyed_rows = []
    fields: list[str] = []
    for path in existing:
        for row in _read_table(path):
            key = _filled_row_key(row)
            if key:
                if key not in rows_by_key:
                    ordered_keys.append(key)
                    rows_by_key[key] = row
                elif _should_replace_filled_row(rows_by_key[key], row):
                    rows_by_key[key] = row
            else:
                unkeyed_rows.append(row)
            for field in row:
                if field not in fields:
                    fields.append(field)
    rows = [rows_by_key[key] for key in ordered_keys if key in rows_by_key]
    rows.extend(unkeyed_rows)
    _write_rows(merged_output, rows, fields)
    return merged_output


def _filled_row_key(row: dict[str, str]) -> tuple[str, str]:
    provider_id = str(row.get("provider_id") or "").strip()
    review_reason = str(row.get("review_reason") or "").strip()
    if not provider_id or not review_reason:
        return ("", "")
    return (provider_id, review_reason)


def _should_replace_filled_row(existing: dict[str, str], incoming: dict[str, str]) -> bool:
    if _has_manual_decision(incoming):
        return True
    if _has_manual_decision(existing):
        return False
    return True


def _has_manual_decision(row: dict[str, str]) -> bool:
    decision = str(
        row.get("manual_decision")
        or row.get("your_decision")
        or row.get("decision")
        or ""
    ).strip()
    return bool(decision)


def _read_table(path: Path) -> list[dict[str, str]]:
    if path.suffix.casefold() == ".xlsx":
        return _read_xlsx(path)
    return _read_rows(path)


def _read_xlsx(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    workbook = load_workbook(path, data_only=False, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows())
    if not rows:
        return []
    headers = [_cell_text(cell.value) for cell in rows[0]]
    out = []
    for cells in rows[1:]:
        row = {headers[idx]: _cell_text(cells[idx].value) for idx in range(len(headers)) if headers[idx]}
        if any(row.values()):
            out.append(row)
    return out


def _read_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _write_rows(path: Path, rows: list[dict[str, str]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _cell_text(value: object) -> str:
    return str(value or "").strip()


def _pattern_release_source_kind(source_path: str, generated: Path, extras: list[str | Path]) -> str:
    if not source_path:
        return "not_evaluated"
    source = Path(source_path)
    if _same_path(source, generated):
        return "current_cycle"
    if any(_same_path(source, Path(extra)) for extra in extras):
        return "supplied_prior"
    return "unknown"


def _same_path(left: Path, right: Path) -> bool:
    try:
        return left.resolve() == right.resolve()
    except OSError:
        return str(left) == str(right)


def _preferred_pattern_release_json(generated: Path, extras: list[str | Path]) -> Path:
    for path_value in extras:
        path = Path(path_value)
        if path.exists():
            return path
    return generated


if __name__ == "__main__":
    raise SystemExit(main())
