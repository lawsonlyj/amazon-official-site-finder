from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from tools.evaluate_policy_validation_task import evaluate_policy_validation_task
from tools.run_calibration_cycle import run_calibration_cycle
from tools.verify_protected_lane_review_task import verify_protected_lane_review_task


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Rerun calibration from a previous cycle summary after one or more label-gap files are filled."
    )
    parser.add_argument("--previous-summary-json", required=True, help="Previous calibration_cycle_summary.json.")
    parser.add_argument(
        "--filled-sample",
        action="append",
        default=[],
        help="Filled calibration sample or label-gap CSV/XLSX. Repeatable.",
    )
    parser.add_argument(
        "--filled-policy-validation",
        action="append",
        default=[],
        help="Filled policy_validation_task CSV/XLSX. Repeatable.",
    )
    parser.add_argument("--output-dir", help="Output directory. Defaults to the previous summary directory.")
    parser.add_argument("--candidate-final-csv", help="Candidate official_sites.csv for regression-gate validation.")
    parser.add_argument(
        "--no-reuse-previous-filled",
        action="store_true",
        help="Use only the newly supplied filled samples instead of also reusing previous filled samples.",
    )
    args = parser.parse_args(argv)

    decision = run_calibration_followup(
        previous_summary_json=args.previous_summary_json,
        filled_sample=args.filled_sample,
        filled_policy_validation=args.filled_policy_validation,
        output_dir=args.output_dir,
        candidate_final_csv=args.candidate_final_csv,
        reuse_previous_filled=not args.no_reuse_previous_filled,
    )
    print(json.dumps(decision["summary"], ensure_ascii=False, indent=2))
    return 0


def run_calibration_followup(
    *,
    previous_summary_json: str | Path,
    filled_sample: str | Path | list[str | Path] | None = None,
    filled_policy_validation: str | Path | list[str | Path] | None = None,
    output_dir: str | Path | None = None,
    candidate_final_csv: str | Path | None = None,
    reuse_previous_filled: bool = True,
) -> dict:
    previous_path = Path(previous_summary_json)
    previous = json.loads(previous_path.read_text(encoding="utf-8"))
    inputs = previous.get("inputs") or {}
    outputs = previous.get("outputs") or {}
    summary = previous.get("summary") or {}
    out_dir = Path(output_dir) if output_dir else previous_path.parent
    out_dir.mkdir(parents=True, exist_ok=True)
    filled_samples = _merged_filled_samples(inputs, outputs, filled_sample, reuse_previous_filled)
    filled_policy_validations = _merged_filled_policy_validations(
        inputs,
        filled_policy_validation,
        reuse_previous_filled,
    )
    filled_sample_verifications = _verify_filled_protected_samples(filled_samples, outputs, out_dir)
    policy_validation_evaluations = _evaluate_filled_policy_validations(filled_policy_validations, out_dir)
    candidate_final = str(candidate_final_csv or inputs.get("candidate_final_csv") or "")
    sample_prefix = str(inputs.get("sample_prefix") or _sample_prefix(outputs.get("sample_csv")))
    report = run_calibration_cycle(
        labeled_eval_json=_required_input(inputs, "labeled_eval_json"),
        labeled_agent_b_csv=_required_input(inputs, "labeled_agent_b_csv"),
        review_csv=_required_input(inputs, "review_csv"),
        batch_agent_b_csv=_required_input(inputs, "batch_agent_b_csv"),
        batch_total_rows=_to_int(inputs.get("batch_total_rows")),
        output_dir=out_dir,
        sample_prefix=sample_prefix,
        max_rows=_to_int(inputs.get("max_rows"), _to_int(summary.get("sample_rows"), 50)),
        max_per_reason=_to_int(inputs.get("max_per_reason"), 12),
        max_per_pattern=_to_int(inputs.get("max_per_pattern"), _to_int(summary.get("max_per_pattern"), 5)),
        min_support=_to_int(inputs.get("min_support"), 2),
        max_pattern_size=_to_int(inputs.get("max_pattern_size"), 3),
        filled_sample=filled_samples,
        pattern_release_jsons=inputs.get("pattern_release_jsons") or [],
        policy_report_json=inputs.get("policy_report_json") or None,
        candidate_final_csv=candidate_final or None,
    )
    decision_json = out_dir / "calibration_followup_decision.json"
    decision_md = out_dir / "calibration_followup_decision.md"
    decision = _build_decision(
        report=report,
        previous_summary_json=previous_path,
        filled_samples=filled_samples,
        filled_policy_validations=filled_policy_validations,
        filled_sample_verifications=filled_sample_verifications,
        policy_validation_evaluations=policy_validation_evaluations,
        candidate_final_csv=candidate_final,
        output_json=decision_json,
        output_md=decision_md,
    )
    decision_json.write_text(json.dumps(decision, ensure_ascii=False, indent=2), encoding="utf-8")
    decision_md.write_text(_render_decision_markdown(decision), encoding="utf-8")
    return decision


def _build_decision(
    *,
    report: dict,
    previous_summary_json: Path,
    filled_samples: list[Path],
    filled_policy_validations: list[Path],
    filled_sample_verifications: list[dict],
    policy_validation_evaluations: list[dict],
    candidate_final_csv: str,
    output_json: Path,
    output_md: Path,
) -> dict:
    status = (report.get("calibration_status") or {}).get("summary") or {}
    gates = report.get("application_gate_checks") or {}
    gate_summary = gates.get("summary") or {}
    checks = gates.get("checks") or []
    convergence = report.get("convergence_audit") or {}
    convergence_summary = convergence.get("summary") or {}
    convergence_next_actions = convergence.get("next_actions") or []
    allowed = gate_summary.get("allowed_gates") or []
    candidates = gate_summary.get("candidate_gates") or []
    not_allowed = gate_summary.get("not_allowed_gates") or []
    blocked = [gate for gate in not_allowed if gate not in candidates]
    next_actions = (report.get("calibration_status") or {}).get("next_actions") or []
    effective_next_actions = list(convergence_next_actions or next_actions)
    effective_next_actions.extend(_policy_validation_next_actions(policy_validation_evaluations))
    policy_summary = _policy_validation_summary(policy_validation_evaluations)
    policy_decision = _policy_validation_decision(policy_summary)
    summary = {
        "workflow_status": status.get("workflow_status", ""),
        "convergence_state": convergence_summary.get("convergence_state", ""),
        "threshold_decision": convergence_summary.get("threshold_decision", ""),
        "review_lane_decision": convergence_summary.get("review_lane_decision", ""),
        "pattern_release_decision": convergence_summary.get("pattern_release_decision", ""),
        "current_threshold_ties_best_accuracy": convergence_summary.get("current_threshold_ties_best_accuracy"),
        "recommended_global_accept_threshold": status.get("recommended_global_accept_threshold"),
        "recommended_second_pass_threshold": status.get("recommended_second_pass_threshold"),
        "filled_labeled_rows": status.get("filled_labeled_rows"),
        "filled_decisive_rows": status.get("filled_decisive_rows"),
        "filled_protected_sample_verification_count": len(filled_sample_verifications),
        "filled_protected_sample_verification_passed": all(
            bool((item.get("summary") or {}).get("passed")) for item in filled_sample_verifications
        )
        if filled_sample_verifications
        else None,
        "filled_policy_validation_file_count": len(filled_policy_validations),
        "filled_policy_validation_labeled_rows": policy_summary["labeled_rows"],
        "filled_policy_validation_decisive_rows": policy_summary["decisive_rows"],
        "filled_policy_validation_support_rows": policy_summary["support_rows"],
        "filled_policy_validation_blocking_rows": policy_summary["blocking_rows"],
        "filled_policy_candidate_for_rule_count": policy_summary["candidate_for_rule_rows"],
        "filled_policy_needs_more_labels_count": policy_summary["needs_more_labels_rows"],
        "filled_policy_reject_pattern_count": policy_summary["reject_pattern_rows"],
        "policy_validation_decision": policy_decision["decision"],
        "policy_validation_gate_status": policy_decision["gate_status"],
        "policy_validation_required_action": policy_decision["required_action"],
        "filled_lane_candidate_for_change_count": (report.get("summary") or {}).get(
            "filled_lane_candidate_for_change_count"
        ),
        "filled_lane_keep_review_count": (report.get("summary") or {}).get("filled_lane_keep_review_count"),
        "filled_rule_candidate_count": (report.get("summary") or {}).get("filled_rule_candidate_count"),
        "filled_rejected_pattern_count": (report.get("summary") or {}).get("filled_rejected_pattern_count"),
        "protected_lanes_next_review_task_rows": convergence_summary.get("protected_lanes_next_review_task_rows"),
        "protected_lanes_priority_task_rows": (report.get("summary") or {}).get("protected_lanes_priority_task_rows"),
        "regression_gate_status": status.get("regression_gate_status"),
        "allowed_gate_count": gate_summary.get("allowed_gate_count", 0),
        "candidate_gate_count": gate_summary.get("candidate_gate_count", 0),
        "not_allowed_gate_count": gate_summary.get("not_allowed_gate_count", 0),
        "blocked_gate_count": len(blocked),
        "allowed_gates": allowed,
        "candidate_gates": candidates,
        "not_allowed_gates": not_allowed,
        "blocked_gates": blocked,
        "ready_to_apply_any_change": bool(allowed),
        "candidate_changes_need_controlled_rollout": bool(candidates and not allowed),
        "next_action": _primary_next_action(effective_next_actions)
        or _default_next_action(allowed, candidates, blocked),
    }
    return {
        "summary": summary,
        "inputs": {
            "previous_summary_json": str(previous_summary_json),
            "filled_samples": [str(path) for path in filled_samples],
            "filled_policy_validations": [str(path) for path in filled_policy_validations],
            "candidate_final_csv": candidate_final_csv,
        },
        "outputs": {
            "decision_json": str(output_json),
            "decision_md": str(output_md),
            "calibration_status_json": (report.get("outputs") or {}).get("status_json", ""),
            "application_gates_json": (report.get("outputs") or {}).get("application_gates_json", ""),
            "convergence_audit_json": (report.get("outputs") or {}).get("convergence_audit_json", ""),
            "convergence_audit_md": (report.get("outputs") or {}).get("convergence_audit_md", ""),
            "label_gap_high_priority_xlsx": (report.get("outputs") or {}).get("label_gap_high_priority_xlsx", ""),
            "protected_lanes_next_review_task_xlsx": (report.get("outputs") or {}).get(
                "protected_lanes_next_review_task_xlsx", ""
            ),
            "protected_lanes_priority_task_xlsx": (report.get("outputs") or {}).get(
                "protected_lanes_priority_task_xlsx", ""
            ),
            "protected_lanes_priority_task_handoff_md": (report.get("outputs") or {}).get(
                "protected_lanes_priority_task_handoff_md", ""
            ),
            "protected_lanes_next_review_task_verification_json": (report.get("outputs") or {}).get(
                "protected_lanes_next_review_task_verification_json", ""
            ),
            "protected_lanes_priority_task_verification_json": (report.get("outputs") or {}).get(
                "protected_lanes_priority_task_verification_json", ""
            ),
            "protected_lanes_next_review_task_verification_md": (report.get("outputs") or {}).get(
                "protected_lanes_next_review_task_verification_md", ""
            ),
            "protected_lanes_priority_task_verification_md": (report.get("outputs") or {}).get(
                "protected_lanes_priority_task_verification_md", ""
            ),
            "regression_cases_csv": (report.get("outputs") or {}).get("regression_cases_csv", ""),
            "filled_eval_json": (report.get("outputs") or {}).get("filled_eval_json", ""),
            "filled_eval_md": (report.get("outputs") or {}).get("filled_eval_md", ""),
            "filled_eval_csv": (report.get("outputs") or {}).get("filled_eval_csv", ""),
            "pattern_rule_candidates_json": (report.get("outputs") or {}).get("rule_candidates_json", ""),
            "pattern_rule_candidates_md": (report.get("outputs") or {}).get("rule_candidates_md", ""),
            "regression_gate_json": (report.get("outputs") or {}).get("regression_gate_json", ""),
            "filled_protected_sample_verification_json": str(output_json.with_name("filled_protected_sample_verification.json"))
            if filled_sample_verifications
            else "",
            "filled_protected_sample_verification_md": str(output_md.with_name("filled_protected_sample_verification.md"))
            if filled_sample_verifications
            else "",
            "policy_validation_eval_json": str(output_json.with_name("filled_policy_validation_evaluation.json"))
            if policy_validation_evaluations
            else "",
            "policy_validation_eval_md": str(output_md.with_name("filled_policy_validation_evaluation.md"))
            if policy_validation_evaluations
            else "",
        },
        "filled_protected_sample_verifications": filled_sample_verifications,
        "filled_policy_validation_evaluations": policy_validation_evaluations,
        "filled_policy_rule_candidates": _policy_validation_rule_candidates(policy_validation_evaluations),
        "filled_lane_recommendations": report.get("filled_lane_recommendations") or [],
        "filled_pattern_rule_candidates": report.get("filled_pattern_rule_candidates") or {},
        "application_gate_checks": checks,
        "convergence_audit": convergence,
        "next_actions": effective_next_actions,
    }


def _render_decision_markdown(decision: dict) -> str:
    summary = decision["summary"]
    lines = [
        "# Calibration Follow-Up Decision",
        "",
        f"- Workflow status: {summary.get('workflow_status')}",
        f"- Convergence state: {summary.get('convergence_state') or 'not_audited'}",
        f"- Threshold decision: {summary.get('threshold_decision') or 'not_audited'}",
        f"- Review-lane decision: {summary.get('review_lane_decision') or 'not_audited'}",
        f"- Pattern-release decision: {summary.get('pattern_release_decision') or 'not_audited'}",
        f"- Recommended thresholds: {summary.get('recommended_global_accept_threshold')}/{summary.get('recommended_second_pass_threshold')}",
        f"- Current threshold ties best accuracy: {str(summary.get('current_threshold_ties_best_accuracy')).lower()}",
        f"- Protected-lane next review rows: {summary.get('protected_lanes_next_review_task_rows')}",
        f"- Protected-lane priority review rows: {summary.get('protected_lanes_priority_task_rows')}",
        f"- Filled protected sample verification: {summary.get('filled_protected_sample_verification_count')}/{summary.get('filled_protected_sample_verification_passed')}",
        f"- Filled policy validation files: {summary.get('filled_policy_validation_file_count')}",
        f"- Filled policy validation labeled/decisive rows: {summary.get('filled_policy_validation_labeled_rows')}/{summary.get('filled_policy_validation_decisive_rows')}",
        f"- Filled policy validation support/block rows: {summary.get('filled_policy_validation_support_rows')}/{summary.get('filled_policy_validation_blocking_rows')}",
        f"- Filled policy candidate/needs-more/reject patterns: {summary.get('filled_policy_candidate_for_rule_count')}/{summary.get('filled_policy_needs_more_labels_count')}/{summary.get('filled_policy_reject_pattern_count')}",
        f"- Policy validation decision: {summary.get('policy_validation_decision')}",
        f"- Policy validation gate status: {summary.get('policy_validation_gate_status')}",
        f"- Policy validation required action: {summary.get('policy_validation_required_action')}",
        f"- Filled lane candidate/keep-review counts: {summary.get('filled_lane_candidate_for_change_count')}/{summary.get('filled_lane_keep_review_count')}",
        f"- Filled pattern candidate/rejected counts: {summary.get('filled_rule_candidate_count')}/{summary.get('filled_rejected_pattern_count')}",
        f"- Filled labeled/decisive rows: {summary.get('filled_labeled_rows')}/{summary.get('filled_decisive_rows')}",
        f"- Regression gate status: {summary.get('regression_gate_status')}",
        f"- Allowed gates: {', '.join(summary.get('allowed_gates') or []) or 'None'}",
        f"- Candidate gates: {', '.join(summary.get('candidate_gates') or []) or 'None'}",
        f"- Not allowed gates: {', '.join(summary.get('not_allowed_gates') or []) or 'None'}",
        f"- Blocked gates: {', '.join(summary.get('blocked_gates') or []) or 'None'}",
        f"- Next action: {summary.get('next_action') or 'None'}",
        "",
        "## Filled Samples",
        "",
    ]
    for path in decision.get("inputs", {}).get("filled_samples", []):
        lines.append(f"- {path}")
    if not decision.get("inputs", {}).get("filled_samples"):
        lines.append("- None")
    lines.extend(["", "## Filled Protected Sample Verification", ""])
    for item in decision.get("filled_protected_sample_verifications", []):
        summary = item.get("summary") or {}
        lines.append(
            "- {path}: passed={passed}, rows={rows}, failures={failures}".format(
                path=(item.get("inputs") or {}).get("csv") or "",
                passed=str(summary.get("passed")).lower(),
                rows=summary.get("row_count"),
                failures=summary.get("failure_count"),
            )
        )
    if not decision.get("filled_protected_sample_verifications"):
        lines.append("- None")
    lines.extend(["", "## Filled Policy Validation", ""])
    policy_evals = decision.get("filled_policy_validation_evaluations") or []
    for item in policy_evals:
        item_summary = item.get("summary") or {}
        lines.append(
            "- {path}: rows={rows}, decisive={decisive}, support={support}, blocking={blocking}, candidate={candidate}, needs_more={needs_more}, reject={reject}".format(
                path=(item.get("inputs") or {}).get("task") or "",
                rows=item_summary.get("task_rows"),
                decisive=item_summary.get("decisive_rows"),
                support=item_summary.get("support_rows"),
                blocking=item_summary.get("blocking_rows"),
                candidate=item_summary.get("candidate_for_rule_rows"),
                needs_more=item_summary.get("needs_more_labels_rows"),
                reject=item_summary.get("reject_pattern_rows"),
            )
        )
    if not policy_evals:
        lines.append("- None")
    lines.extend(["", "## Filled Policy Rule Candidates", ""])
    policy_candidates = decision.get("filled_policy_rule_candidates") or {}
    for bucket in ["candidate_for_rule", "reject_pattern", "needs_more_labels"]:
        rows = policy_candidates.get(bucket) or []
        lines.append(f"- {bucket}: {len(rows)}")
        for row in rows[:10]:
            lines.append(
                "  - support={support}, block={block}, pattern={pattern}".format(
                    support=row.get("supporting_rows"),
                    block=row.get("blocking_rows"),
                    pattern=row.get("pattern"),
                )
            )
    lines.extend(["", "## Filled Lane Recommendations", ""])
    for row in decision.get("filled_lane_recommendations", []):
        lines.append(
            "- {recommendation}: {review_reason}, decisive={decisive}, support={support}, blocking={blocking}, action={action}".format(
                recommendation=row.get("recommendation"),
                review_reason=row.get("review_reason"),
                decisive=row.get("decisive_rows"),
                support=row.get("support_rows"),
                blocking=row.get("blocking_rows"),
                action=row.get("required_action") or "",
            )
        )
    if not decision.get("filled_lane_recommendations"):
        lines.append("- None")
    lines.extend(["", "## Filled Pattern Rule Candidates", ""])
    candidates = decision.get("filled_pattern_rule_candidates") or {}
    for bucket in ["candidate_for_rule", "reject_pattern", "needs_more_labels"]:
        rows = candidates.get(bucket) or []
        lines.append(f"- {bucket}: {len(rows)}")
        for row in rows[:10]:
            lines.append(
                "  - {recommendation}: support={support}, block={block}, pattern={pattern}".format(
                    recommendation=row.get("recommendation"),
                    support=row.get("support_rows"),
                    block=row.get("blocking_rows"),
                    pattern=row.get("pattern"),
                )
            )
    if not candidates:
        lines.append("- None")
    lines.extend(["", "## Gate Details", ""])
    for row in decision.get("application_gate_checks", []):
        blockers = ", ".join(row.get("blockers") or []) or "none"
        lines.append(
            "- {gate}: allowed={allowed}, status={status}, blockers={blockers}, reason={reason}".format(
                gate=row.get("gate"),
                allowed=str(row.get("allowed")).lower(),
                status=row.get("gate_status"),
                blockers=blockers,
                reason=row.get("reason") or row.get("decision_reason") or "",
            )
        )
    lines.extend(["", "## Next Actions", ""])
    for action in decision.get("next_actions", []):
        lines.append(f"- {action}")
    if not decision.get("next_actions"):
        lines.append("- None")
    lines.append("")
    return "\n".join(lines)


def _evaluate_filled_policy_validations(paths: list[Path], out_dir: Path) -> list[dict]:
    reports = []
    for path in paths:
        reports.append(evaluate_policy_validation_task(task=path))
    if reports:
        _write_policy_validation_evaluation(reports, out_dir)
    return reports


def _write_policy_validation_evaluation(reports: list[dict], out_dir: Path) -> None:
    output_json = out_dir / "filled_policy_validation_evaluation.json"
    output_md = out_dir / "filled_policy_validation_evaluation.md"
    payload = {
        "summary": _policy_validation_summary(reports),
        "evaluations": reports,
        "policy_rule_candidates": _policy_validation_rule_candidates(reports),
    }
    output_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    output_md.write_text(_render_policy_validation_evaluation(payload), encoding="utf-8")


def _policy_validation_summary(reports: list[dict]) -> dict:
    keys = [
        "task_rows",
        "labeled_rows",
        "decisive_rows",
        "decision_quality_issue_rows",
        "support_rows",
        "blocking_rows",
        "candidate_for_rule_rows",
        "needs_more_labels_rows",
        "reject_pattern_rows",
    ]
    summary = {key: 0 for key in keys}
    summary["file_count"] = len(reports)
    for report in reports:
        item = report.get("summary") or {}
        for key in keys:
            summary[key] += _to_int(item.get(key))
    return summary


def _policy_validation_rule_candidates(reports: list[dict]) -> dict[str, list[dict]]:
    buckets = {"candidate_for_rule": [], "needs_more_labels": [], "reject_pattern": []}
    seen: set[tuple[str, str]] = set()
    for report in reports:
        source = (report.get("inputs") or {}).get("task") or ""
        candidates = report.get("policy_rule_candidates") or {}
        for bucket in buckets:
            for row in candidates.get(bucket) or []:
                key = (bucket, row.get("pattern", ""))
                if key in seen:
                    continue
                seen.add(key)
                enriched = dict(row)
                enriched["source_task"] = source
                buckets[bucket].append(enriched)
    return buckets


def _policy_validation_next_actions(reports: list[dict]) -> list[str]:
    if not reports:
        return []
    summary = _policy_validation_summary(reports)
    actions = []
    if summary["decision_quality_issue_rows"]:
        actions.append("Fix policy-validation fill-quality issues before using those rows for rule decisions.")
    if summary["reject_pattern_rows"] or summary["blocking_rows"]:
        actions.append("Block policy patterns rejected by filled policy-validation labels and add regression fixtures.")
    if summary["candidate_for_rule_rows"]:
        actions.append("Add regression tests for candidate policy-validation patterns before enabling them.")
    if summary["needs_more_labels_rows"]:
        actions.append("Keep supported but thin policy-validation patterns in the next targeted validation task.")
    return actions


def _policy_validation_decision(summary: dict) -> dict[str, str]:
    if not summary.get("file_count"):
        return {
            "decision": "not_evaluated",
            "gate_status": "not_evaluated",
            "required_action": "No filled policy-validation workbook was supplied.",
        }
    if summary.get("decision_quality_issue_rows"):
        return {
            "decision": "fix_fill_quality",
            "gate_status": "blocked",
            "required_action": "Fix invalid manual decisions or missing manual_url values before using policy-validation evidence.",
        }
    if summary.get("reject_pattern_rows") or summary.get("blocking_rows"):
        return {
            "decision": "blocked_by_policy_validation",
            "gate_status": "blocked",
            "required_action": "Do not automate blocked policy patterns; add blocking regression fixtures for those provider/candidate rows.",
        }
    if summary.get("candidate_for_rule_rows"):
        return {
            "decision": "candidate_for_rule",
            "gate_status": "candidate",
            "required_action": "Add regression tests for every supporting row, then consider applying only the exact validated pattern.",
        }
    if summary.get("needs_more_labels_rows") or summary.get("support_rows"):
        return {
            "decision": "needs_more_labels",
            "gate_status": "blocked",
            "required_action": "Keep this exact pattern in targeted policy-validation tasks until it reaches enough clean decisive labels.",
        }
    return {
        "decision": "not_evaluated",
        "gate_status": "not_evaluated",
        "required_action": "No decisive policy-validation evidence is available yet.",
    }


def _render_policy_validation_evaluation(payload: dict) -> str:
    summary = payload.get("summary") or {}
    lines = [
        "# Filled Policy Validation Evaluation",
        "",
        f"- Files: {summary.get('file_count')}",
        f"- Rows: {summary.get('task_rows')}",
        f"- Labeled/decisive rows: {summary.get('labeled_rows')}/{summary.get('decisive_rows')}",
        f"- Support/block rows: {summary.get('support_rows')}/{summary.get('blocking_rows')}",
        f"- Candidate/needs-more/reject patterns: {summary.get('candidate_for_rule_rows')}/{summary.get('needs_more_labels_rows')}/{summary.get('reject_pattern_rows')}",
        "",
        "## Pattern Buckets",
        "",
    ]
    candidates = payload.get("policy_rule_candidates") or {}
    for bucket in ["candidate_for_rule", "reject_pattern", "needs_more_labels"]:
        rows = candidates.get(bucket) or []
        lines.append(f"### {bucket}")
        if not rows:
            lines.append("- None")
            continue
        for row in rows[:20]:
            lines.append(
                "- support={support}, block={block}: {pattern}".format(
                    support=row.get("supporting_rows"),
                    block=row.get("blocking_rows"),
                    pattern=row.get("pattern"),
                )
            )
    lines.append("")
    return "\n".join(lines)


def _verify_filled_protected_samples(filled_samples: list[Path], outputs: dict, out_dir: Path) -> list[dict]:
    verifications = []
    for path in filled_samples:
        if not _looks_like_protected_lane_task(path):
            continue
        summary_json = _summary_for_filled_sample(path, outputs)
        report = verify_protected_lane_review_task(
            csv_path=path,
            summary_json=summary_json or None,
            xlsx_path=path if path.suffix.casefold() == ".xlsx" else None,
            allow_filled=True,
            require_filled=True,
        )
        verifications.append(report)
        if not report["summary"].get("passed"):
            _write_filled_protected_verifications(verifications, out_dir)
            raise ValueError(
                f"Filled protected-lane sample failed verification: {path}. "
                f"Failures: {report.get('failures') or []}"
            )
    if verifications:
        _write_filled_protected_verifications(verifications, out_dir)
    return verifications


def _looks_like_protected_lane_task(path: Path) -> bool:
    lowered = path.name.casefold()
    if "protected_lanes" in lowered:
        return True
    headers = _table_headers(path)
    return (
        {"provider_id", "review_reason", "manual_decision", "optimization_use"}.issubset(headers)
        and ("protected_lane_priority" in headers or "priority_rank" in headers)
    )


def _summary_for_filled_sample(path: Path, outputs: dict) -> str:
    lowered = path.name.casefold()
    if "priority" in lowered:
        return str(outputs.get("protected_lanes_priority_task_summary_json") or "")
    return str(outputs.get("protected_lanes_next_review_task_summary_json") or "")


def _write_filled_protected_verifications(verifications: list[dict], out_dir: Path) -> None:
    output_json = out_dir / "filled_protected_sample_verification.json"
    output_md = out_dir / "filled_protected_sample_verification.md"
    payload = {"verifications": verifications, "summary": _verification_summary(verifications)}
    output_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    output_md.write_text(_render_filled_protected_verifications(payload), encoding="utf-8")


def _verification_summary(verifications: list[dict]) -> dict:
    return {
        "verification_count": len(verifications),
        "passed": all(bool((item.get("summary") or {}).get("passed")) for item in verifications),
        "row_count": sum(_to_int((item.get("summary") or {}).get("row_count")) for item in verifications),
        "failure_count": sum(_to_int((item.get("summary") or {}).get("failure_count")) for item in verifications),
    }


def _render_filled_protected_verifications(payload: dict) -> str:
    summary = payload.get("summary") or {}
    lines = [
        "# Filled Protected Sample Verification",
        "",
        f"- Passed: {str(summary.get('passed')).lower()}",
        f"- Verifications: {summary.get('verification_count')}",
        f"- Rows: {summary.get('row_count')}",
        f"- Failures: {summary.get('failure_count')}",
        "",
        "## Files",
        "",
    ]
    for item in payload.get("verifications") or []:
        item_summary = item.get("summary") or {}
        lines.append(
            "- {path}: passed={passed}, rows={rows}, failures={failures}".format(
                path=(item.get("inputs") or {}).get("csv") or "",
                passed=str(item_summary.get("passed")).lower(),
                rows=item_summary.get("row_count"),
                failures=item_summary.get("failure_count"),
            )
        )
    lines.append("")
    return "\n".join(lines)


def _table_headers(path: Path) -> set[str]:
    if not path.exists():
        return set()
    if path.suffix.casefold() == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError:
            return set()
        workbook = load_workbook(path, read_only=True, data_only=False)
        sheet = workbook.active
        first_row = next(sheet.iter_rows(max_row=1), [])
        return {str(cell.value or "").strip() for cell in first_row if str(cell.value or "").strip()}
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return {str(item or "").strip() for item in (reader.fieldnames or []) if str(item or "").strip()}


def _merged_filled_samples(
    inputs: dict,
    outputs: dict,
    filled_sample: str | Path | list[str | Path] | None,
    reuse_previous: bool,
) -> list[Path]:
    paths: list[Path] = []
    if reuse_previous:
        paths.extend(
            path
            for path in (Path(path) for path in inputs.get("filled_samples") or [] if str(path))
            if _previous_filled_sample_can_reuse(path, outputs)
        )
        previous_single = str(inputs.get("filled_sample") or "")
        if previous_single:
            previous_path = Path(previous_single)
            if _previous_filled_sample_can_reuse(previous_path, outputs):
                paths.append(previous_path)
    if filled_sample:
        if isinstance(filled_sample, (str, Path)):
            paths.append(Path(filled_sample))
        else:
            paths.extend(Path(path) for path in filled_sample if str(path))
    deduped: list[Path] = []
    seen = set()
    for path in paths:
        key = str(path)
        if key and key not in seen:
            seen.add(key)
            deduped.append(path)
    return deduped


def _previous_filled_sample_can_reuse(path: Path, outputs: dict) -> bool:
    if not path.exists():
        return False
    if _looks_like_protected_lane_task(path):
        try:
            report = verify_protected_lane_review_task(
                csv_path=path,
                summary_json=_summary_for_filled_sample(path, outputs) or None,
                xlsx_path=path if path.suffix.casefold() == ".xlsx" else None,
                allow_filled=True,
                require_filled=True,
            )
        except Exception:
            return False
        return bool((report.get("summary") or {}).get("passed"))
    headers = _table_headers(path)
    if "manual_decision" in headers:
        return _table_has_any_manual_decision(path)
    return True


def _table_has_any_manual_decision(path: Path) -> bool:
    for row in _table_rows(path):
        if str(row.get("manual_decision") or "").strip():
            return True
    return False


def _table_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    if path.suffix.casefold() == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError:
            return []
        workbook = load_workbook(path, read_only=True, data_only=False)
        sheet = workbook.active
        rows = list(sheet.iter_rows())
        if not rows:
            return []
        headers = [str(cell.value or "").strip() for cell in rows[0]]
        out = []
        for cells in rows[1:]:
            row = {
                headers[idx]: str(cells[idx].value or "").strip()
                for idx in range(len(headers))
                if headers[idx]
            }
            if any(row.values()):
                out.append(row)
        return out
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _merged_filled_policy_validations(
    inputs: dict,
    filled_policy_validation: str | Path | list[str | Path] | None,
    reuse_previous: bool,
) -> list[Path]:
    paths: list[Path] = []
    if reuse_previous:
        paths.extend(Path(path) for path in inputs.get("filled_policy_validations") or [] if str(path))
        previous_single = str(inputs.get("filled_policy_validation") or "")
        if previous_single:
            paths.append(Path(previous_single))
    if filled_policy_validation:
        if isinstance(filled_policy_validation, (str, Path)):
            paths.append(Path(filled_policy_validation))
        else:
            paths.extend(Path(path) for path in filled_policy_validation if str(path))
    deduped: list[Path] = []
    seen = set()
    for path in paths:
        key = str(path)
        if key and key not in seen:
            seen.add(key)
            deduped.append(path)
    return deduped


def _required_input(inputs: dict, key: str) -> str:
    value = str(inputs.get(key) or "").strip()
    if not value:
        raise ValueError(f"Previous calibration summary is missing inputs.{key}")
    return value


def _sample_prefix(sample_csv: str | None) -> str:
    if not sample_csv:
        return "pattern_validation_sample_50"
    return Path(sample_csv).stem or "pattern_validation_sample_50"


def _to_int(value: object, default: int = 0) -> int:
    try:
        if value is None or value == "":
            return default
        return int(value)
    except (TypeError, ValueError):
        return default


def _default_next_action(allowed: list[str], candidates: list[str], blocked: list[str]) -> str:
    if allowed:
        return "Apply only the allowed calibration changes and keep regression coverage."
    if candidates:
        return "Review candidate gates, run controlled rollout with regression coverage, then rerun the application gate."
    if blocked:
        return "Fill the remaining label-gap task before changing thresholds or routing."
    return "No calibration action is available."


def _primary_next_action(actions: list[str]) -> str:
    if not actions:
        return ""
    for action in actions:
        lowered = action.casefold()
        if "fill " in lowered or "protected-lane" in lowered or "label" in lowered:
            return str(action)
    return str(actions[0])


if __name__ == "__main__":
    raise SystemExit(main())
