from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import zipfile
from collections import Counter
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from finder.finalize import finalize_results
from finder.scoring import load_config
from finder.text import domain_from_url
from tools.build_linked_workbook import build_workbook
from tools.evaluate_labeled_results import read_rows as read_csv_rows
from tools.output_layout import first_existing, reviewed_paths as canonical_reviewed_paths
from tools.quality_gate import evaluate_quality_gate, write_markdown as write_quality_markdown


DECISION_ALIASES = {
    "accept": "accept",
    "accpet": "accept",
    "approve": "accept",
    "approved": "accept",
    "confirm": "accept",
    "confirmed": "accept",
    "yes": "accept",
    "接受": "accept",
    "确认": "accept",
    "replace": "replace",
    "override": "replace",
    "替换": "replace",
    "reject": "reject",
    "rejected": "reject",
    "not_found": "reject",
    "no": "reject",
    "拒绝": "reject",
    "否": "reject",
    "unsure": "unsure",
    "不确定": "unsure",
}

COMBINED_REVIEW_FIELDS = [
    "provider_id",
    "provider_name",
    "official_url",
    "candidate_1_url",
    "manual_decision",
    "manual_url",
    "notes",
    "confidence",
    "source_status",
    "evidence_summary",
    "candidate_count",
    "scored_candidate_count",
    "service_apis",
    "provider_locations",
    "error_type",
]

LABEL_FIELDS = ["provider_id", "provider_name", "expected_url", "expected_domain", "label_source", "notes"]


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Apply filled manual review feedback and generate workflow optimization notes.")
    parser.add_argument("--run-dir", required=True)
    parser.add_argument("--review", required=True, help="Filled manual review CSV or XLSX.")
    parser.add_argument("--labels")
    parser.add_argument("--config", default="config/scoring.json")
    parser.add_argument("--write-xlsx", action="store_true")
    parser.add_argument("--update-config", action="store_true", help="Apply only safe repeated excluded-domain additions.")
    parser.add_argument("--min-domain-accuracy", type=float)
    parser.add_argument("--min-auto-precision", type=float)
    parser.add_argument("--min-official-url-rate", type=float)
    parser.add_argument("--max-unresolved-rate", type=float)
    args = parser.parse_args(argv)

    summary = run_review_learning(
        run_dir=args.run_dir,
        review_path=args.review,
        labels_csv=args.labels,
        config_path=args.config,
        write_xlsx=args.write_xlsx,
        update_config=args.update_config,
        min_domain_accuracy=args.min_domain_accuracy,
        min_auto_precision=args.min_auto_precision,
        min_official_url_rate=args.min_official_url_rate,
        max_unresolved_rate=args.max_unresolved_rate,
    )
    print(json.dumps(summary["overall"], ensure_ascii=False, indent=2))
    return 0 if summary["quality_overall"].get("passed") else 1


def run_review_learning(
    *,
    run_dir: str | Path,
    review_path: str | Path,
    labels_csv: str | Path | None = None,
    config_path: str | Path = "config/scoring.json",
    write_xlsx: bool = True,
    update_config: bool = False,
    min_domain_accuracy: float | None = None,
    min_auto_precision: float | None = None,
    min_official_url_rate: float | None = None,
    max_unresolved_rate: float | None = None,
) -> dict[str, Any]:
    run_dir = Path(run_dir)
    paths = review_learning_paths(run_dir)
    manifest = _read_manifest(run_dir / "manifest.json")
    source_results = _source_results_path(run_dir, manifest)
    if not source_results.exists():
        raise FileNotFoundError(f"source result CSV not found: {source_results}")

    raw_review_rows = _read_table(review_path)
    manual_rows, skipped_rows = _normalize_manual_review_rows(raw_review_rows)
    base_review_rows = _base_review_rows(run_dir)
    combined_rows = _combine_review_rows(base_review_rows, manual_rows)
    _write_rows(paths["combined_review"], combined_rows, COMBINED_REVIEW_FIELDS)

    final_summary = finalize_results(
        source_results,
        paths["final"],
        review_csv=paths["combined_review"],
        unresolved_csv=paths["unresolved"],
    )
    manual_labels = _manual_labels(manual_rows)
    _write_rows(paths["manual_labels"], manual_labels, LABEL_FIELDS)
    labels = _combined_labels(labels_csv, manual_labels)
    params = manifest.get("parameters", {})
    optimization = _optimization_summary(manual_rows, raw_review_rows, base_review_rows)
    config_update = _maybe_update_config(config_path, optimization, update_config=update_config)
    config = load_config(config_path)
    quality = evaluate_quality_gate(
        results_csv=paths["final"],
        config=config,
        labels=labels if labels else None,
        expected_rows=_pick_int(params.get("total_to_run")) or None,
        min_domain_accuracy=_pick_float(min_domain_accuracy, params.get("min_domain_accuracy"), 0.8),
        min_auto_precision=_pick_float(min_auto_precision, params.get("min_auto_precision"), 0.95),
        min_official_url_rate=_pick_float(min_official_url_rate, params.get("min_official_url_rate"), 0.0),
        max_unresolved_rate=_pick_float(max_unresolved_rate, params.get("max_unresolved_rate"), 1.0),
    )
    write_quality_markdown(quality, paths["quality_md"])
    paths["quality_json"].write_text(json.dumps(quality, ensure_ascii=False, indent=2), encoding="utf-8")

    xlsx = {}
    if write_xlsx:
        xlsx = build_workbook(
            [
                ("Reviewed_Final", paths["final"]),
                ("Reviewed_Unresolved", paths["unresolved"]),
                ("Manual_Labels", paths["manual_labels"]),
            ],
            paths["xlsx"],
        )

    summary = {
        "overall": {
            "raw_review_rows": len(raw_review_rows),
            "applied_manual_rows": len(manual_rows),
            "skipped_manual_rows": len(skipped_rows),
            "base_review_rows": len(base_review_rows),
            "combined_review_rows": len(combined_rows),
            "final_rows": final_summary["final_rows"],
            "official_url_rows": final_summary["official_url_rows"],
            "unresolved_rows": final_summary["unresolved_rows"],
            "quality_passed": quality["overall"]["passed"],
        },
        "finalize": final_summary,
        "quality_overall": quality["overall"],
        "optimization": optimization,
        "config_update": config_update,
        "outputs": {name: str(path) for name, path in paths.items()},
        "xlsx": xlsx,
    }
    paths["summary"].write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    _write_report(paths["report_md"], summary)
    _update_manifest(run_dir / "manifest.json", summary)
    return summary


def review_learning_paths(run_dir: str | Path) -> dict[str, Path]:
    return canonical_reviewed_paths(run_dir)


def _read_table(path: str | Path) -> list[dict[str, str]]:
    path = Path(path)
    if path.suffix.casefold() == ".xlsx":
        return _read_xlsx(path)
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _read_xlsx(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return _read_simple_xlsx(path)
    workbook = load_workbook(path, data_only=False, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    out = []
    for values in rows[1:]:
        row = {headers[idx]: _cell_text(values[idx] if idx < len(values) else "") for idx in range(len(headers)) if headers[idx]}
        if any(value for value in row.values()):
            out.append(row)
    return out


def _read_simple_xlsx(path: Path) -> list[dict[str, str]]:
    with zipfile.ZipFile(path) as z:
        sheet_names = sorted(name for name in z.namelist() if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"))
        if not sheet_names:
            return []
        text = z.read(sheet_names[0]).decode("utf-8", errors="replace")
    rows = []
    for row_xml in re.findall(r"<row\b[^>]*>(.*?)</row>", text):
        cells = []
        for cell_xml in re.findall(r"<c\b[^>]*>(.*?)</c>", row_xml):
            formula = re.search(r"<f>(.*?)</f>", cell_xml)
            if formula:
                cells.append(_formula_to_url(_xml_unescape(formula.group(1))))
                continue
            match = re.search(r"<t[^>]*>(.*?)</t>", cell_xml)
            cells.append(_xml_unescape(match.group(1)) if match else "")
        rows.append(cells)
    if not rows:
        return []
    headers = [value.strip() for value in rows[0]]
    return [
        {headers[idx]: values[idx] if idx < len(values) else "" for idx in range(len(headers)) if headers[idx]}
        for values in rows[1:]
        if any(values)
    ]


def _normalize_manual_review_rows(rows: list[dict[str, str]]) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    normalized = []
    skipped = []
    for row in rows:
        decision = _decision(row)
        manual_url = _first(row, "manual_url", "your_true_official_url", "true_official_url")
        if decision == "reject" and manual_url:
            decision = "replace"
        if not decision and manual_url:
            decision = "replace"
        if decision == "unsure" or (not decision and not manual_url):
            skipped.append(row)
            continue
        out = dict(row)
        out["manual_decision"] = decision
        out["manual_url"] = manual_url
        out["notes"] = _first(row, "notes", "your_notes", "manual_notes")
        out["error_type"] = _first(row, "error_type", "your_error_type", "error_reason")
        if not out.get("official_url"):
            out["official_url"] = _first(row, "top_candidate_url", "candidate_1_url")
        if not out.get("candidate_1_url"):
            out["candidate_1_url"] = _first(row, "top_candidate_url", "official_url")
        normalized.append({field: out.get(field, "") for field in COMBINED_REVIEW_FIELDS})
    return normalized, skipped


def _decision(row: dict[str, str]) -> str:
    raw = _first(row, "manual_decision", "your_decision", "decision").strip().casefold()
    return DECISION_ALIASES.get(raw, raw)


def _first(row: dict[str, str], *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip():
            return _cell_text(value)
    return ""


def _cell_text(value: object) -> str:
    text = str(value or "").strip()
    if text.startswith("="):
        return _formula_to_url(text[1:])
    if text.startswith("HYPERLINK("):
        return _formula_to_url(text)
    return text


def _formula_to_url(value: str) -> str:
    match = re.search(r'HYPERLINK\("([^"]+)"', value)
    return match.group(1).replace('""', '"') if match else value


def _base_review_rows(run_dir: Path) -> list[dict[str, str]]:
    for path in [
        first_existing(run_dir, "details/second_pass/decisions.csv", "unresolved_second_pass_review_decisions.csv"),
        first_existing(run_dir, "details/first_pass/review_queue.csv", "provider_review_queue.csv"),
    ]:
        if path and path.exists():
            return _read_csv_rows(path)
    return []


def _combine_review_rows(base_rows: list[dict[str, str]], manual_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    combined: dict[str, dict[str, str]] = {}
    for row in base_rows:
        key = _row_key(row)
        if key:
            combined[key] = {field: row.get(field, "") for field in COMBINED_REVIEW_FIELDS}
    for row in manual_rows:
        key = _row_key(row)
        if key:
            combined[key] = {field: row.get(field, "") for field in COMBINED_REVIEW_FIELDS}
    return list(combined.values())


def _manual_labels(manual_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    labels = []
    for row in manual_rows:
        decision = row.get("manual_decision", "")
        url = row.get("manual_url") or row.get("official_url") or row.get("candidate_1_url")
        if decision not in {"accept", "replace"} or not url:
            continue
        labels.append(
            {
                "provider_id": row.get("provider_id", ""),
                "provider_name": row.get("provider_name", ""),
                "expected_url": _normalize_url(url),
                "expected_domain": domain_from_url(url),
                "label_source": "manual_review",
                "notes": row.get("notes", ""),
            }
        )
    return labels


def _combined_labels(labels_csv: str | Path | None, manual_labels: list[dict[str, str]]) -> list[dict[str, str]]:
    labels = read_csv_rows(labels_csv) if labels_csv and Path(labels_csv).exists() else []
    by_key = {_row_key(label): label for label in labels if _row_key(label)}
    for label in manual_labels:
        by_key[_row_key(label)] = label
    return list(by_key.values())


def _optimization_summary(
    manual_rows: list[dict[str, str]],
    raw_review_rows: list[dict[str, str]],
    base_review_rows: list[dict[str, str]],
) -> dict[str, Any]:
    raw_by_key = {_row_key(row): row for row in raw_review_rows if _row_key(row)}
    base_by_key = {_row_key(row): row for row in base_review_rows if _row_key(row)}
    decision_counts = Counter(row.get("manual_decision", "") for row in manual_rows)
    rejected_domains = Counter()
    accepted_domains = Counter()
    accepted_low_confidence = 0
    rejected_auto_accepted = 0
    recall_additions = 0
    no_official_rows = 0
    wrong_candidate_rows = 0
    examples = []
    for row in manual_rows:
        raw = raw_by_key.get(_row_key(row), {})
        base = base_by_key.get(_row_key(row), {})
        decision = row.get("manual_decision", "")
        error_type = row.get("error_type", "")
        current_url = raw.get("official_url") or raw.get("top_candidate_url") or base.get("manual_url") or base.get("official_url") or ""
        current_domain = domain_from_url(current_url)
        final_url = row.get("manual_url") or raw.get("official_url") or raw.get("candidate_1_url") or ""
        final_domain = domain_from_url(final_url)
        confidence = _to_int(raw.get("confidence") or row.get("confidence"))
        if decision in {"accept", "replace"} and final_domain:
            accepted_domains[final_domain] += 1
            if confidence and confidence < 70:
                accepted_low_confidence += 1
            if not current_domain:
                recall_additions += 1
        if decision == "reject" and current_domain:
            rejected_domains[current_domain] += 1
            if base.get("manual_decision") or raw.get("status") == "manual_accepted":
                rejected_auto_accepted += 1
        if _is_no_official_error(error_type, row.get("notes", "")):
            no_official_rows += 1
            if current_domain:
                wrong_candidate_rows += 1
        examples.append(
            {
                "provider_id": row.get("provider_id", ""),
                "provider_name": row.get("provider_name", ""),
                "decision": decision,
                "previous_domain": current_domain,
                "final_domain": final_domain,
                "confidence": confidence,
                "error_type": error_type,
            }
        )
    exclude_candidates = [
        domain
        for domain, count in rejected_domains.items()
        if domain not in accepted_domains and (count >= 2 or _looks_like_directory_domain(domain))
    ]
    recommendations = _recommendations(
        decision_counts=decision_counts,
        rejected_auto_accepted=rejected_auto_accepted,
        accepted_low_confidence=accepted_low_confidence,
        recall_additions=recall_additions,
        no_official_rows=no_official_rows,
        wrong_candidate_rows=wrong_candidate_rows,
        exclude_candidates=exclude_candidates,
    )
    return {
        "manual_decision_counts": dict(decision_counts),
        "rejected_auto_accepted_rows": rejected_auto_accepted,
        "accepted_low_confidence_rows": accepted_low_confidence,
        "recall_additions_from_manual_review": recall_additions,
        "confirmed_no_official_rows": no_official_rows,
        "wrong_candidate_no_official_rows": wrong_candidate_rows,
        "rejected_domains": dict(rejected_domains),
        "accepted_domains": dict(accepted_domains),
        "safe_excluded_domain_candidates": exclude_candidates,
        "recommendations": recommendations,
        "examples": examples[:50],
    }


def _recommendations(
    *,
    decision_counts: Counter,
    rejected_auto_accepted: int,
    accepted_low_confidence: int,
    recall_additions: int,
    no_official_rows: int,
    wrong_candidate_rows: int,
    exclude_candidates: list[str],
) -> list[str]:
    recommendations = []
    if rejected_auto_accepted:
        recommendations.append(
            "Inspect rejected auto/second-pass accepted rows before lowering thresholds; add repeated platform or directory domains to excluded_domains."
        )
    if accepted_low_confidence:
        recommendations.append(
            "Low-confidence manual accepts are useful recall seeds; keep them as labels and compare future second-pass changes against manual_review_labels.csv."
        )
    if recall_additions:
        recommendations.append(
            "Manual replacements on unresolved rows indicate recall gaps; inspect their query patterns and add targeted second-pass queries only when several examples share the same pattern."
        )
    if no_official_rows:
        recommendations.append(
            "No-official labels are precision regression seeds; prevent forced same-name matches and keep these as no_official fixtures."
        )
    if wrong_candidate_rows:
        recommendations.append(
            "Rejected candidates marked as actual no-official should tighten identity, country, and service gates rather than become excluded domains."
        )
    if exclude_candidates:
        recommendations.append(
            "Safe excluded-domain candidates were found. Review them before running with --update-config."
        )
    if not recommendations and decision_counts:
        recommendations.append("No repeated safe rule change was detected; keep the review labels for regression evaluation.")
    return recommendations


def _maybe_update_config(config_path: str | Path, optimization: dict[str, Any], *, update_config: bool) -> dict[str, Any]:
    if not update_config:
        return {"updated": False, "reason": "not_requested"}
    path = Path(config_path)
    config = load_config(path)
    existing = set(config.get("excluded_domains", []))
    additions = [domain for domain in optimization.get("safe_excluded_domain_candidates", []) if domain not in existing]
    if additions:
        config["excluded_domains"] = list(config.get("excluded_domains", [])) + additions
        path.write_text(json.dumps(config, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return {"updated": bool(additions), "added_excluded_domains": additions, "config_path": str(path)}


def _is_no_official_error(error_type: str, notes: str) -> bool:
    text = f"{error_type} {notes}".casefold()
    return any(marker in text for marker in ["实际无官网", "no official", "no reliable official"])


def _write_report(path: Path, summary: dict[str, Any]) -> None:
    overall = summary["overall"]
    optimization = summary["optimization"]
    lines = [
        "# Manual Review Learning Report",
        "",
        "## Overall",
        "",
        "| Metric | Value |",
        "|---|---:|",
    ]
    for key, value in overall.items():
        lines.append(f"| {key} | {value} |")
    lines.extend(["", "## Manual Decisions", "", "| Decision | Rows |", "|---|---:|"])
    for decision, count in sorted(optimization["manual_decision_counts"].items()):
        lines.append(f"| {decision or '(blank)'} | {count} |")
    lines.extend(["", "## Recommendations", ""])
    for item in optimization["recommendations"]:
        lines.append(f"- {item}")
    if optimization["safe_excluded_domain_candidates"]:
        lines.extend(["", "## Safe Excluded-Domain Candidates", ""])
        for domain in optimization["safe_excluded_domain_candidates"]:
            lines.append(f"- `{domain}`")
    lines.extend(["", "## Outputs", ""])
    for name, output_path in summary["outputs"].items():
        lines.append(f"- `{name}`: `{output_path}`")
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def _source_results_path(run_dir: Path, manifest: dict[str, Any]) -> Path:
    outputs = manifest.get("outputs", {})
    for value in [
        run_dir / "details/first_pass/enriched.csv",
        outputs.get("results_enriched"),
        run_dir / "provider_official_websites_enriched.csv",
        outputs.get("results"),
        run_dir / "provider_official_websites.csv",
    ]:
        if value and Path(value).exists():
            return Path(value)
    return run_dir / "provider_official_websites_enriched.csv"


def _read_manifest(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"parameters": {}, "outputs": {}}
    return json.loads(path.read_text(encoding="utf-8"))


def _update_manifest(path: Path, summary: dict[str, Any]) -> None:
    if not path.exists():
        return
    manifest = _read_manifest(path)
    manifest["review_learning"] = summary
    manifest.setdefault("summary", {}).update(
        {
            "manual_review_applied_rows": summary["overall"]["applied_manual_rows"],
            "reviewed_official_url_rows": summary["overall"]["official_url_rows"],
            "reviewed_unresolved_rows": summary["overall"]["unresolved_rows"],
            "reviewed_quality_passed": summary["overall"]["quality_passed"],
        }
    )
    manifest.setdefault("outputs", {}).update({f"review_{name}": value for name, value in summary["outputs"].items()})
    path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _write_rows(path: str | Path, rows: list[dict[str, str]], fields: list[str]) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _row_key(row: dict[str, str]) -> str:
    provider_id = (row.get("provider_id") or "").strip()
    if provider_id:
        return f"id:{provider_id}"
    return f"name:{(row.get('provider_name') or '').strip().casefold()}"


def _normalize_url(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return ""
    if "://" not in value:
        return f"https://{value}"
    return value


def _looks_like_directory_domain(domain: str) -> bool:
    markers = [
        "linkedin.com",
        "facebook.com",
        "instagram.com",
        "youtube.com",
        "crunchbase.com",
        "trustpilot.com",
        "clutch.co",
        "goodfirms.co",
        "kompass.com",
        "exportersindia.com",
        "opencorporates.com",
    ]
    return domain in markers or any(domain.endswith(f".{marker}") for marker in markers)


def _pick_float(explicit: float | None, manifest_value: object, default: float) -> float:
    if explicit is not None:
        return explicit
    try:
        return float(manifest_value)
    except (TypeError, ValueError):
        return default


def _pick_int(value: object) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _to_int(value: object) -> int:
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return 0


def _xml_unescape(value: str) -> str:
    return (
        value.replace("&quot;", '"')
        .replace("&apos;", "'")
        .replace("&lt;", "<")
        .replace("&gt;", ">")
        .replace("&amp;", "&")
    )


if __name__ == "__main__":
    raise SystemExit(main())
